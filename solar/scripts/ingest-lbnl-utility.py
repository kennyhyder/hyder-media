#!/usr/bin/env python3
"""
LBNL Utility-Scale Solar Data Ingestion Script

Downloads and imports the Lawrence Berkeley National Lab (LBNL)
Utility-Scale Solar database (~1,760 projects, all >5 MW AC) including:
- Project name, location, state, county
- AC and DC capacity
- Tracking type, mount configuration
- Developer, owner, operator names
- Cost data (installed cost $/W_DC)
- Commercial operation date (COD)

Source: https://data.openei.org/submissions/8541
"""

import os
import sys
import json
import re
import uuid
import urllib.request
import urllib.parse
from pathlib import Path
from datetime import datetime

import openpyxl
from dotenv import load_dotenv

# Load env vars
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

# The LBNL data file URL on OpenEI (2025 edition)
# The file naming follows the pattern: "2025 Utility-Scale Solar Data Update.xlsx"
# hosted under data.openei.org/files/8541/
DOWNLOAD_URL = "https://data.openei.org/files/8541/2025%20Utility-Scale%20Solar%20Data%20Update%20(1).xlsx"
DATA_DIR = Path(__file__).parent.parent / "data" / "lbnl_utility"

MIN_SIZE_KW = 1000  # Utility-scale only (>= 1 MW)
BATCH_SIZE = 50


def supabase_request(method, table, data=None, params=None, headers_extra=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(f"{k}={urllib.parse.quote(str(v), safe='.*,')}" for k, v in params.items())

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates",
    }
    if headers_extra:
        headers.update(headers_extra)

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else []
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()[:200]
        print(f"  Supabase error ({e.code}): {error_body}")
        return None


def get_or_create_data_source():
    """Get or create the LBNL Utility-Scale data source record."""
    params = {"name": "eq.lbnl_utility_scale", "select": "id"}
    existing = supabase_request("GET", "solar_data_sources", params=params)
    if existing:
        return existing[0]["id"]

    ds_id = str(uuid.uuid4())
    supabase_request("POST", "solar_data_sources", {
        "id": ds_id,
        "name": "lbnl_utility_scale",
        "description": "LBNL Utility-Scale Solar - Empirical plant-level data from U.S. ground-mounted PV, PV+battery, and CSP plants (>5 MW AC)",
        "url": "https://data.openei.org/submissions/8541",
        "record_count": 0,
    })
    return ds_id


def get_existing_source_ids():
    """Load existing lbnl_ source_record_ids from Supabase."""
    existing = set()
    offset = 0
    while True:
        params = {
            "select": "source_record_id",
            "source_record_id": "like.lbnl_*",
            "order": "source_record_id",
            "offset": str(offset),
            "limit": "1000",
        }
        batch = supabase_request("GET", "solar_installations", params=params)
        if not batch:
            break
        for r in batch:
            existing.add(r["source_record_id"])
        if len(batch) < 1000:
            break
        offset += len(batch)
    return existing


def download_data():
    """Download the LBNL Utility-Scale Solar Excel file."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = DATA_DIR / "lbnl_utility_scale_solar.xlsx"

    if xlsx_path.exists():
        size_mb = xlsx_path.stat().st_size / 1024 / 1024
        print(f"  Found existing Excel file ({size_mb:.1f} MB), skipping download")
        return xlsx_path

    print(f"  Downloading from {DOWNLOAD_URL}...")
    try:
        req = urllib.request.Request(DOWNLOAD_URL, headers={
            "User-Agent": "Mozilla/5.0 (SolarTrack Data Ingestion)"
        })
        with urllib.request.urlopen(req) as resp:
            with open(xlsx_path, "wb") as f:
                f.write(resp.read())
        size_mb = xlsx_path.stat().st_size / 1024 / 1024
        print(f"  Downloaded {size_mb:.1f} MB")
    except urllib.error.HTTPError as e:
        print(f"  Download failed ({e.code}): {e.reason}")
        print(f"  URL: {DOWNLOAD_URL}")
        print(f"\n  The download URL may have changed. Please:")
        print(f"  1. Visit https://data.openei.org/submissions/8541")
        print(f"  2. Download the main data file (Excel/XLSX)")
        print(f"  3. Save it as: {xlsx_path}")
        sys.exit(1)

    return xlsx_path


def safe_str(val):
    """Convert value to string, handling None and empty."""
    if val is None or val == "" or val == "N/A" or val == "n/a":
        return None
    s = str(val).strip()
    return s if s else None


def safe_float(val):
    """Convert value to float, handling None and empty."""
    if val is None or val == "" or val == " ":
        return None
    try:
        return float(str(val).replace(",", "").replace("$", ""))
    except (ValueError, TypeError):
        return None


def parse_date(val):
    """Parse date value to ISO format string."""
    if val is None or val == "" or val == " ":
        return None

    # Handle datetime objects from openpyxl
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    val = str(val).strip()

    # Try numeric year only (e.g., 2023 or 2023.0)
    try:
        year = int(float(val))
        if 1990 <= year <= 2030:
            return f"{year}-01-01"
    except (ValueError, TypeError):
        pass

    # Try various date formats
    for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S", "%Y"]:
        try:
            dt = datetime.strptime(val.split(" ")[0], fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


def normalize_tracking(val):
    """Normalize tracking type to standard values."""
    if not val:
        return None
    v = str(val).strip().lower()
    if "single" in v or "1-axis" in v or "one" in v:
        return "single-axis"
    if "dual" in v or "2-axis" in v or "two" in v:
        return "dual-axis"
    if "fixed" in v:
        return "fixed-tilt"
    return safe_str(val)


def normalize_mount_type(val, tracking_val=None):
    """Normalize mount type to standard values."""
    if not val:
        return "ground_fixed"  # LBNL data is all ground-mounted utility-scale
    v = str(val).strip().lower()
    if "roof" in v:
        return "rooftop"
    if "carport" in v or "canop" in v:
        return "carport"
    if "float" in v:
        return "floating"
    # Check tracking to differentiate ground_fixed vs ground_single_axis
    if tracking_val:
        t = str(tracking_val).strip().lower()
        if "single" in t or "1-axis" in t or "one" in t:
            return "ground_single_axis"
        if "dual" in t or "2-axis" in t or "two" in t:
            return "ground_single_axis"
    if "fixed" in v:
        return "ground_fixed"
    if "track" in v:
        return "ground_single_axis"
    return "ground_fixed"


def find_project_sheet(wb):
    """Find the sheet containing project-level data.

    The LBNL workbook has ~57 tabs. The main project list is typically on a
    sheet named something like 'Project List', 'Projects', 'Plant-Level Data',
    'Data', or 'USS20XX'. We try several heuristics.
    """
    sheet_names = wb.sheetnames
    print(f"  Workbook has {len(sheet_names)} sheets")

    # Priority patterns to match (case-insensitive)
    priority_patterns = [
        "individual",       # "Individual_Project_Data" (LBNL 2025 format)
        "plant-level",      # "Plant-Level Data"
        "plant data",       # "Plant Data"
        "project list",     # "Project List"
        "project data",     # "Project Data"
        "site list",        # "Site List"
        "site data",        # "Site Data"
        "install",          # "Installation Data"
    ]

    # Skip patterns (summary/chart sheets we don't want)
    skip_patterns = [
        "chart", "figure", "graph", "summary", "notes", "source",
        "about", "readme", "contents", "toc", "cover", "glossary",
        "ppa", "price", "value", "performance", "generation",
        "queue", "csp", "battery", "hybrid", "cost trend",
        "o&m", "lcoe", "market", "curtailment",
    ]

    # First pass: look for exact priority pattern matches
    for pattern in priority_patterns:
        for name in sheet_names:
            name_lower = name.strip().lower()
            # Skip if it matches a skip pattern
            if any(skip in name_lower for skip in skip_patterns):
                continue
            if pattern in name_lower:
                print(f"  Using sheet: '{name}' (matched pattern '{pattern}')")
                return name

    # Second pass: look for sheets with many rows (likely data sheets)
    # Try each sheet and pick the one with the most rows containing what looks
    # like project data (has state abbreviations, MW values, etc.)
    print("  No obvious project sheet found. Scanning sheets for project data...")
    best_sheet = None
    best_row_count = 0

    for name in sheet_names:
        name_lower = name.strip().lower()
        if any(skip in name_lower for skip in skip_patterns):
            continue
        try:
            ws = wb[name]
            # Sample first 5 data rows to see if this looks like project data
            row_count = 0
            for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
                if row and any(safe_str(cell) for cell in row):
                    row_count += 1
            if row_count > best_row_count:
                best_row_count = row_count
                best_sheet = name
        except Exception:
            continue

    if best_sheet:
        print(f"  Falling back to sheet: '{best_sheet}' ({best_row_count} rows sampled)")
        return best_sheet

    # Last resort: first sheet
    print(f"  WARNING: Using first sheet: '{sheet_names[0]}'")
    return sheet_names[0]


def find_header_row(ws, max_scan=20):
    """Find the header row in the worksheet.

    LBNL files sometimes have title rows, notes, or blank rows before
    the actual column headers. We scan the first N rows looking for the
    one that has the most non-empty cells and contains expected keywords.
    """
    expected_keywords = {
        "project", "name", "state", "capacity", "mw", "kw", "ac",
        "dc", "developer", "owner", "operator", "cod", "date",
        "tracking", "county", "latitude", "longitude", "cost",
        "eia", "status", "technology",
    }

    best_row_idx = 0
    best_score = 0

    rows = list(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True))
    for i, row in enumerate(rows):
        if not row:
            continue
        cells = [str(c).strip().lower() for c in row if c is not None and str(c).strip()]
        if not cells:
            continue

        # Score: number of cells matching expected keywords
        score = 0
        for cell in cells:
            for kw in expected_keywords:
                if kw in cell:
                    score += 1
                    break

        # Bonus for having many non-empty cells (header rows are typically full)
        score += len(cells) * 0.1

        if score > best_score:
            best_score = score
            best_row_idx = i

    return best_row_idx, rows


def build_column_map(headers):
    """Build a mapping from our target fields to column indices.

    The LBNL file may use various column header names. We map them to
    standardized field names with fuzzy matching.
    """
    col_map = {}

    # Define target fields and their possible header patterns
    # Each entry: target_field -> list of (priority, pattern_substring)
    field_patterns = {
        "project_name": [
            "project name", "project", "plant name", "plant", "site name", "name",
        ],
        "state": [
            "state",
        ],
        "county": [
            "county",
        ],
        "city": [
            "city",
        ],
        "latitude": [
            "latitude", "lat",
        ],
        "longitude": [
            "longitude", "lng", "lon",
        ],
        "capacity_ac_mw": [
            "capacity (mw ac)", "capacity ac", "mw ac", "ac capacity",
            "nameplate capacity (mw)", "nameplate capacity",
            "capacity_mw_ac", "ac (mw)", "mwac", "mw-ac",
        ],
        "capacity_dc_mw": [
            "capacity (mw dc)", "capacity dc", "mw dc", "dc capacity",
            "capacity_mw_dc", "dc (mw)", "mwdc", "mw-dc",
        ],
        "cod_date": [
            "cod", "commercial operation date", "operation date",
            "cod date", "cod year", "online date", "in-service",
        ],
        "developer": [
            "developer",
        ],
        "owner": [
            "owner",
        ],
        "operator": [
            "operator", "utility",
        ],
        "tracking": [
            "tracking", "tracker", "axis",
        ],
        "mount_type": [
            "mount", "mounting",
        ],
        "technology": [
            "technology", "tech", "module type", "panel type", "pv type",
        ],
        "eia_id": [
            "eia plant", "eia id", "eia code", "plant code", "plant id",
            "eia_id", "eia plant code",
        ],
        "cost_per_watt": [
            "installed cost", "cost ($/w", "$/w", "cost per watt",
            "cost_per_watt", "$/wdc", "$/watt",
        ],
        "total_cost": [
            "total cost", "total installed cost", "project cost",
        ],
        "status": [
            "status", "operational status",
        ],
        "zip_code": [
            "zip", "zipcode", "zip code", "postal",
        ],
        "tilt": [
            "tilt", "tilt angle",
        ],
        "azimuth": [
            "azimuth",
        ],
        "num_modules": [
            "number of modules", "module count", "num modules", "modules",
        ],
        "num_inverters": [
            "number of inverters", "inverter count", "num inverters", "inverters",
        ],
        "module_manufacturer": [
            "module manufacturer", "panel manufacturer", "module mfr",
        ],
        "module_model": [
            "module model", "panel model",
        ],
        "inverter_manufacturer": [
            "inverter manufacturer", "inverter mfr",
        ],
        "inverter_model": [
            "inverter model",
        ],
        "battery_storage": [
            "battery", "storage", "bess",
        ],
    }

    # Normalize headers for comparison
    normalized_headers = []
    for h in headers:
        if h is None:
            normalized_headers.append("")
        else:
            normalized_headers.append(str(h).strip().lower())

    # Short patterns that must match as whole words to avoid false positives
    # e.g., "city" must not match "capacity", "county" must not match "accounting"
    SHORT_PATTERNS = {"city", "county", "state", "lat", "lon", "lng", "cod"}

    # Match each field to a column index
    for field, patterns in field_patterns.items():
        for pattern in patterns:
            for col_idx, header in enumerate(normalized_headers):
                if not header:
                    continue
                if pattern in SHORT_PATTERNS:
                    # Use word boundary matching for short patterns
                    if re.search(r'\b' + re.escape(pattern) + r'\b', header):
                        if field not in col_map:
                            col_map[field] = col_idx
                        break
                elif pattern in header:
                    # Don't overwrite a more specific match
                    if field not in col_map:
                        col_map[field] = col_idx
                    break
            if field in col_map:
                break

    return col_map


def get_cell(row, col_map, field):
    """Safely get a cell value from a row using the column map."""
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def process_excel(xlsx_path, data_source_id):
    """Process the LBNL Utility-Scale Solar Excel file."""
    print(f"\n  Loading {xlsx_path.name}...")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find the right sheet
    sheet_name = find_project_sheet(wb)
    ws = wb[sheet_name]

    # Find header row
    header_row_idx, scanned_rows = find_header_row(ws)
    headers = scanned_rows[header_row_idx] if header_row_idx < len(scanned_rows) else []
    headers = list(headers) if headers else []

    print(f"  Header row: {header_row_idx + 1}")
    print(f"  Found {len(headers)} columns")

    # Print discovered headers for debugging
    non_empty = [str(h).strip() for h in headers if h is not None and str(h).strip()]
    print(f"  Sample headers: {non_empty[:15]}")

    # Build column mapping
    col_map = build_column_map(headers)
    print(f"  Mapped {len(col_map)} fields: {list(col_map.keys())}")

    if not col_map:
        print("  ERROR: Could not map any columns. Check the sheet structure.")
        print(f"  All sheet names: {wb.sheetnames}")
        wb.close()
        return 0, 0, 0

    # Verify we have at minimum a capacity or project name column
    if "capacity_ac_mw" not in col_map and "capacity_dc_mw" not in col_map and "project_name" not in col_map:
        print("  ERROR: No capacity or project name column found. Wrong sheet?")
        print(f"  Headers found: {non_empty}")
        wb.close()
        return 0, 0, 0

    # Load existing IDs for dedup
    existing_ids = get_existing_source_ids()
    print(f"  Existing LBNL records: {len(existing_ids)}")

    total = 0
    utility = 0
    created = 0
    errors = 0
    equipment_count = 0
    skipped_small = 0
    skipped_dup = 0

    inst_batch = []
    eq_batch = []

    # Iterate data rows (skip header rows)
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        total += 1

        # Get capacity - prefer AC, fall back to DC
        capacity_ac_mw = safe_float(get_cell(row, col_map, "capacity_ac_mw"))
        capacity_dc_mw = safe_float(get_cell(row, col_map, "capacity_dc_mw"))

        # Convert MW to kW
        capacity_ac_kw = round(capacity_ac_mw * 1000, 3) if capacity_ac_mw else None
        capacity_dc_kw = round(capacity_dc_mw * 1000, 3) if capacity_dc_mw else None

        # Use whichever capacity is available for size filtering
        size_kw = capacity_ac_kw or capacity_dc_kw
        if not size_kw or size_kw < MIN_SIZE_KW:
            skipped_small += 1
            continue

        utility += 1

        # Project identifier for source_record_id
        project_name = safe_str(get_cell(row, col_map, "project_name"))
        eia_id = safe_str(get_cell(row, col_map, "eia_id"))
        state = safe_str(get_cell(row, col_map, "state"))

        # Build a unique source_record_id
        if eia_id:
            source_record_id = f"lbnl_{eia_id}"
        elif project_name:
            # Sanitize project name for use as ID
            clean_name = project_name.lower().replace(" ", "_").replace("/", "_")[:80]
            # Add state to help differentiate projects with same name
            if state:
                source_record_id = f"lbnl_{state.lower()}_{clean_name}"
            else:
                source_record_id = f"lbnl_{clean_name}"
        else:
            source_record_id = f"lbnl_row_{total}"

        # Skip existing records
        if source_record_id in existing_ids:
            skipped_dup += 1
            continue

        inst_id = str(uuid.uuid4())

        # Location
        county = safe_str(get_cell(row, col_map, "county"))
        city = safe_str(get_cell(row, col_map, "city"))
        zip_code = safe_str(get_cell(row, col_map, "zip_code"))
        if zip_code:
            zip_code = str(zip_code).strip()[:10]

        lat = safe_float(get_cell(row, col_map, "latitude"))
        lon = safe_float(get_cell(row, col_map, "longitude"))

        # Validate lat/lon
        if lat is not None and (lat < 18 or lat > 72):
            lat = None
        if lon is not None and (lon < -180 or lon > -60):
            lon = None

        # Date (COD = Commercial Operation Date)
        install_date = parse_date(get_cell(row, col_map, "cod_date"))

        # Entities
        developer_name = safe_str(get_cell(row, col_map, "developer"))
        owner_name = safe_str(get_cell(row, col_map, "owner"))
        operator_name = safe_str(get_cell(row, col_map, "operator"))

        # Tracking
        tracking_raw = safe_str(get_cell(row, col_map, "tracking"))
        tracking_type = normalize_tracking(tracking_raw)

        # Mount type (LBNL is all ground-mounted utility-scale)
        mount_raw = safe_str(get_cell(row, col_map, "mount_type"))
        mount_type = normalize_mount_type(mount_raw, tracking_raw)

        # Cost
        cost_per_watt = safe_float(get_cell(row, col_map, "cost_per_watt"))
        total_cost = safe_float(get_cell(row, col_map, "total_cost"))

        # If we have cost_per_watt but not total_cost, calculate it
        if cost_per_watt and not total_cost and capacity_dc_kw:
            total_cost = round(cost_per_watt * capacity_dc_kw * 1000, 2)
        elif cost_per_watt and not total_cost and capacity_ac_kw:
            total_cost = round(cost_per_watt * capacity_ac_kw * 1000, 2)

        # If we have total_cost but not cost_per_watt, calculate it
        if total_cost and not cost_per_watt:
            watt_capacity = (capacity_dc_kw or capacity_ac_kw)
            if watt_capacity:
                cost_per_watt = round(total_cost / (watt_capacity * 1000), 3)

        # Status
        status_raw = safe_str(get_cell(row, col_map, "status"))
        site_status = "active"
        if status_raw:
            sl = status_raw.lower()
            if "retire" in sl or "decommission" in sl:
                site_status = "decommissioned"
            elif "construct" in sl or "develop" in sl:
                site_status = "under_construction"

        # Battery storage
        battery_raw = safe_str(get_cell(row, col_map, "battery_storage"))
        has_battery = False
        if battery_raw:
            bl = battery_raw.lower()
            if bl in ("yes", "y", "true", "1") or "battery" in bl or "storage" in bl:
                has_battery = True

        # Capacity MW for convenience column
        capacity_mw = None
        if capacity_ac_mw:
            capacity_mw = round(capacity_ac_mw, 6)
        elif capacity_dc_mw:
            capacity_mw = round(capacity_dc_mw, 6)

        # Validate state (should be 2-char code)
        if state and len(state) > 2:
            # Might be a full state name - try to keep first 2 chars or skip
            state = state[:2].upper() if len(state) >= 2 else state

        installation = {
            "id": inst_id,
            "source_record_id": source_record_id,
            "data_source_id": data_source_id,
            "site_name": project_name,
            "state": state[:2] if state else None,
            "county": county,
            "city": city,
            "zip_code": zip_code,
            "latitude": lat,
            "longitude": lon,
            "capacity_dc_kw": capacity_dc_kw,
            "capacity_ac_kw": capacity_ac_kw,
            "capacity_mw": capacity_mw,
            "mount_type": mount_type,
            "tracking_type": tracking_type,
            "install_date": install_date,
            "site_type": "utility",
            "site_status": site_status,
            "owner_name": owner_name,
            "developer_name": developer_name,
            "operator_name": operator_name,
            "total_cost": total_cost,
            "cost_per_watt": cost_per_watt,
            "has_battery_storage": has_battery,
        }

        inst_batch.append(installation)

        # Equipment records
        # Module
        module_mfr = safe_str(get_cell(row, col_map, "module_manufacturer"))
        module_model = safe_str(get_cell(row, col_map, "module_model"))
        technology = safe_str(get_cell(row, col_map, "technology"))
        num_modules = safe_float(get_cell(row, col_map, "num_modules"))

        if module_mfr or module_model or technology:
            eq_record = {
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "equipment_type": "module",
                "equipment_status": "active" if site_status == "active" else "removed",
                "data_source_id": data_source_id,
            }
            if module_mfr:
                eq_record["manufacturer"] = module_mfr
            if module_model:
                eq_record["model"] = module_model
            if technology:
                eq_record["module_technology"] = technology
            if num_modules:
                eq_record["quantity"] = int(num_modules)
            eq_batch.append(eq_record)

        # Inverter
        inverter_mfr = safe_str(get_cell(row, col_map, "inverter_manufacturer"))
        inverter_model = safe_str(get_cell(row, col_map, "inverter_model"))
        num_inverters = safe_float(get_cell(row, col_map, "num_inverters"))

        if inverter_mfr or inverter_model:
            eq_record = {
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "equipment_type": "inverter",
                "equipment_status": "active" if site_status == "active" else "removed",
                "data_source_id": data_source_id,
            }
            if inverter_mfr:
                eq_record["manufacturer"] = inverter_mfr
            if inverter_model:
                eq_record["model"] = inverter_model
            if num_inverters:
                eq_record["quantity"] = int(num_inverters)
            eq_batch.append(eq_record)

        # Flush batches
        if len(inst_batch) >= BATCH_SIZE:
            res = supabase_request("POST", "solar_installations", inst_batch)
            if res is not None:
                created += len(inst_batch)
            else:
                # Retry individually on batch failure (handles stray duplicates)
                for rec in inst_batch:
                    res2 = supabase_request("POST", "solar_installations", [rec])
                    if res2 is not None:
                        created += 1
                    else:
                        errors += 1
            inst_batch = []

            if eq_batch:
                for i in range(0, len(eq_batch), BATCH_SIZE):
                    chunk = eq_batch[i:i + BATCH_SIZE]
                    res = supabase_request("POST", "solar_equipment", chunk)
                    if res is not None:
                        equipment_count += len(chunk)
                eq_batch = []

            if created % 100 == 0 and created > 0:
                print(f"    {created}/{utility} created, {errors} errors, {equipment_count} equipment")

    # Flush remaining
    if inst_batch:
        res = supabase_request("POST", "solar_installations", inst_batch)
        if res is not None:
            created += len(inst_batch)
        else:
            for rec in inst_batch:
                res2 = supabase_request("POST", "solar_installations", [rec])
                if res2 is not None:
                    created += 1
                else:
                    errors += 1

    if eq_batch:
        for i in range(0, len(eq_batch), BATCH_SIZE):
            chunk = eq_batch[i:i + BATCH_SIZE]
            res = supabase_request("POST", "solar_equipment", chunk)
            if res is not None:
                equipment_count += len(chunk)

    wb.close()

    print(f"\n  Results: {total} total rows scanned")
    print(f"    Utility-scale (>= {MIN_SIZE_KW} kW): {utility}")
    print(f"    Skipped (too small): {skipped_small}")
    print(f"    Skipped (duplicate): {skipped_dup}")
    print(f"    New: {created}, Errors: {errors}")
    print(f"    Equipment: {equipment_count}")

    return created, equipment_count, errors


def main():
    print("LBNL Utility-Scale Solar Data Ingestion Script")
    print("=" * 60)
    print(f"Source: https://data.openei.org/submissions/8541")
    print(f"Filter: >= {MIN_SIZE_KW} kW ({MIN_SIZE_KW / 1000:.0f} MW)")
    print()

    data_source_id = get_or_create_data_source()
    print(f"Data source ID: {data_source_id}")

    # Download data
    print("\nDownloading LBNL Utility-Scale Solar data...")
    xlsx_path = download_data()

    # Process
    print("\nProcessing Excel file...")
    created, equipment, errors = process_excel(xlsx_path, data_source_id)

    # Update data source record count
    supabase_request(
        "PATCH",
        "solar_data_sources",
        {"record_count": created},
        params={"name": "eq.lbnl_utility_scale"},
    )

    print("\n" + "=" * 60)
    print("LBNL Utility-Scale Solar ingestion complete!")
    print(f"  Installations created: {created}")
    print(f"  Equipment records: {equipment}")
    print(f"  Errors: {errors}")


if __name__ == "__main__":
    main()
