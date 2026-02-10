#!/usr/bin/env python3
"""
PJM-GATS Solar Generator Enrichment Script

Cross-references PJM-GATS (Generation Attribute Tracking System) generator export
with existing solar installations. Enriches owner_name using Plant Name field.

PJM-GATS covers 13+ PJM states: NJ, PA, MD, DE, DC, OH, VA, WV, IL, IN, MI, KY, NC.
Data source: https://gats.pjm-eis.com/gats2/PublicReports/GATSGenerators
Export: Filter Fuel Type = "Solar - Photovoltaic", download XLSX

The GATS export has these columns:
  Unit ID, Plant Name, Unit Name, Balancing Authority, Fuel Type, Owner?,
  New Jersey, Maryland, District of Columbia, Pennsylvania, Delaware,
  Illinois, Ohio, Virginia, EFEC Eligible, IL ZEC, IL CMC, OH SGF

Key limitations:
  - "Owner?" is just Y/N flag, NOT the actual owner name
  - Plant Name contains entity name (business/individual) or address
  - Unit Name sometimes contains capacity (e.g., "13.30 kW") but often doesn't
  - No latitude/longitude or structured address fields
  - 582K+ solar records, mostly residential

Strategy: Extract records >= 25 kW from Unit Name capacity, plus all MSET (metered
utility-scale) records. Cross-reference with existing installations by state + name
similarity to fill owner_name.

Usage:
  python3 -u scripts/enrich-pjm-gats.py                    # Full enrichment
  python3 -u scripts/enrich-pjm-gats.py --dry-run           # Report without patching
  python3 -u scripts/enrich-pjm-gats.py --file /path/to.xlsx # Use specific file
"""

import os
import sys
import json
import re
import argparse
import urllib.request
import urllib.parse
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

# Load env vars
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "data" / "pjm_gats"

# State columns in GATS export → state abbreviation
STATE_COLS = {
    "New Jersey": "NJ",
    "Maryland": "MD",
    "District of Columbia": "DC",
    "Pennsylvania": "PA",
    "Delaware": "DE",
    "Illinois": "IL",
    "Ohio": "OH",
    "Virginia": "VA",
}

# MSET utility prefixes → owner names
UTILITY_PREFIXES = {
    "AEP": "American Electric Power",
    "AP": "Allegheny Power",
    "DPL": "Delmarva Power & Light",
    "EKPC": "East Kentucky Power Cooperative",
    "JC": "Jersey Central Power & Light",
    "ME": "Metropolitan Edison",
    "PN": "PECO Energy",
    "PS": "Public Service Electric & Gas",
    "VP": "Virginia Power (Dominion Energy)",
}


def supabase_get(table, params):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode())


def supabase_patch(table, data, params):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    body = json.dumps(data).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="PATCH")
    try:
        with urllib.request.urlopen(req) as resp:
            return True
    except Exception as e:
        print(f"    PATCH error: {e}")
        return False


def parse_capacity_kw(unit_name):
    """Extract capacity in kW from Unit Name field."""
    if not unit_name:
        return None
    s = str(unit_name).strip()
    # Match "123.45 kW" or "123 kw" - avoid version numbers like "14.3.0"
    m = re.match(r'^([\d]+\.?\d*)\s*kw$', s, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


def extract_owner_from_mset(plant_name):
    """Extract owner name from MSET utility-scale plant names."""
    if not plant_name:
        return None
    # MSET names like "AEP MAMMOTH NORTH 1 SP" → "American Electric Power"
    for prefix, owner in UTILITY_PREFIXES.items():
        if plant_name.upper().startswith(prefix + " "):
            return owner
    return None


def extract_owner_from_non(plant_name):
    """Extract business/entity name from NON plant names."""
    if not plant_name:
        return None
    name = str(plant_name).strip()
    # Skip pure addresses (start with number)
    if re.match(r'^\d+\s', name) and not any(kw in name.lower() for kw in ['llc', 'inc', 'corp', 'solar', 'farm', 'school', 'church', 'university']):
        return None
    # Skip residential names (First Last pattern without business indicators)
    # Keep anything with LLC, Inc, Corp, School, Church, Farm, Solar, etc.
    business_indicators = ['llc', 'inc', 'corp', 'ltd', 'lp', 'solar', 'farm', 'school',
                          'church', 'university', 'college', 'hospital', 'county', 'city of',
                          'township', 'district', 'authority', 'commission', 'association',
                          'foundation', 'company', 'properties', 'enterprises', 'group']
    name_lower = name.lower()
    if any(ind in name_lower for ind in business_indicators):
        # Clean up suffixes like "– ABP_AppID_123"
        clean = re.sub(r'\s*[–-]\s*(ABP_AppID|PROJ)[-_]\d+.*$', '', name).strip()
        return clean if clean else None
    return None


def determine_state(row_dict):
    """Determine state from GATS state eligibility columns."""
    # State columns contain eligibility labels when the generator is in that state
    for col, abbr in STATE_COLS.items():
        val = row_dict.get(col)
        if val and str(val).strip():
            return abbr
    return None


def name_similarity(name1, name2):
    """Simple word overlap similarity score."""
    if not name1 or not name2:
        return 0
    words1 = set(re.findall(r'\w+', name1.lower()))
    words2 = set(re.findall(r'\w+', name2.lower()))
    # Remove common stop words
    stop = {'the', 'of', 'and', 'in', 'at', 'a', 'an', 'to', 'for', 'on', 'is', 'it', 'sp', '1', '2', '3'}
    words1 -= stop
    words2 -= stop
    if not words1 or not words2:
        return 0
    overlap = words1 & words2
    return len(overlap) / min(len(words1), len(words2))


def main():
    parser = argparse.ArgumentParser(description="PJM-GATS solar generator enrichment")
    parser.add_argument("--dry-run", action="store_true", help="Report without patching")
    parser.add_argument("--file", type=str, help="Path to GATS XLSX export")
    args = parser.parse_args()

    # Find GATS file
    if args.file:
        gats_file = Path(args.file)
    else:
        # Look in data dir or Downloads
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        candidates = list(DATA_DIR.glob("GATSGenerators*.xlsx"))
        if not candidates:
            # Check Downloads
            dl = Path.home() / "Downloads"
            candidates = sorted(dl.glob("GATSGenerators*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if candidates:
            gats_file = candidates[0]
        else:
            print("Error: No GATS file found. Download from https://gats.pjm-eis.com/gats2/PublicReports/GATSGenerators")
            print("  Filter Fuel Type = 'Solar - Photovoltaic', export XLSX")
            print("  Place in data/pjm_gats/ or use --file flag")
            sys.exit(1)

    print("PJM-GATS Solar Generator Enrichment")
    print("=" * 60)
    print(f"  File: {gats_file}")
    print(f"  Dry run: {args.dry_run}")
    print()

    # Copy to data dir if not already there
    if gats_file.parent != DATA_DIR:
        import shutil
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        dest = DATA_DIR / gats_file.name
        if not dest.exists():
            shutil.copy2(gats_file, dest)
            print(f"  Copied to {dest}")

    # -------------------------------------------------------------------------
    # Phase 1: Read GATS file, filter to commercial/utility solar
    # -------------------------------------------------------------------------
    print("Phase 1: Reading GATS export...")
    wb = openpyxl.load_workbook(gats_file, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    gats_records = []
    total_solar = 0
    skipped_small = 0
    skipped_no_state = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        if d.get("Fuel Type") != "Solar - Photovoltaic":
            continue
        total_solar += 1

        uid = str(d.get("Unit ID", ""))
        plant_name = str(d.get("Plant Name", "")).strip()
        unit_name = str(d.get("Unit Name", "")).strip()
        is_mset = uid.startswith("MSET")

        # Determine state
        state = determine_state(d)
        if not state:
            skipped_no_state += 1
            continue

        # Filter: MSET (utility-scale) OR >= 25 kW from Unit Name
        capacity_kw = parse_capacity_kw(unit_name)
        if not is_mset and (capacity_kw is None or capacity_kw < 25):
            skipped_small += 1
            continue

        # Extract owner name
        if is_mset:
            owner = extract_owner_from_mset(plant_name)
        else:
            owner = extract_owner_from_non(plant_name)

        gats_records.append({
            "unit_id": uid,
            "plant_name": plant_name,
            "unit_name": unit_name,
            "state": state,
            "capacity_kw": capacity_kw,
            "is_mset": is_mset,
            "owner": owner,
        })

    wb.close()

    print(f"  Total solar in GATS: {total_solar}")
    print(f"  Skipped (no state): {skipped_no_state}")
    print(f"  Skipped (< 25 kW or no capacity): {skipped_small}")
    print(f"  Qualifying records: {len(gats_records)}")
    print(f"  With extractable owner: {sum(1 for r in gats_records if r['owner'])}")

    mset_count = sum(1 for r in gats_records if r["is_mset"])
    non_count = len(gats_records) - mset_count
    print(f"  MSET (utility): {mset_count}, NON (commercial): {non_count}")

    # State breakdown
    state_counts = {}
    for r in gats_records:
        state_counts[r["state"]] = state_counts.get(r["state"], 0) + 1
    print(f"  By state: {', '.join(f'{s}: {c}' for s, c in sorted(state_counts.items(), key=lambda x: -x[1]))}")
    print()

    # -------------------------------------------------------------------------
    # Phase 2: Load existing installations from PJM states
    # -------------------------------------------------------------------------
    print("Phase 2: Loading existing installations from PJM states...")
    pjm_states = list(set(r["state"] for r in gats_records))

    existing = []
    for state in sorted(pjm_states):
        offset = 0
        while True:
            batch = supabase_get("solar_installations", {
                "select": "id,site_name,owner_name,state,capacity_mw,address,city",
                "state": f"eq.{state}",
                "order": "id",
                "offset": offset,
                "limit": 1000,
            })
            if not batch:
                break
            existing.extend(batch)
            offset += len(batch)
            if len(batch) < 1000:
                break

    print(f"  Loaded {len(existing)} existing installations in PJM states")
    need_owner = [r for r in existing if not r.get("owner_name")]
    print(f"  Missing owner_name: {len(need_owner)}")
    print()

    # -------------------------------------------------------------------------
    # Phase 3: Cross-reference GATS with existing installations
    # -------------------------------------------------------------------------
    print("Phase 3: Cross-referencing GATS records with existing installations...")

    # Index existing by state for faster matching
    by_state = {}
    for inst in need_owner:
        s = inst.get("state", "")
        if s not in by_state:
            by_state[s] = []
        by_state[s].append(inst)

    patches = []
    matched_gats = 0

    # For MSET records: match by state + name similarity with site_name
    for gats in gats_records:
        if not gats["owner"]:
            continue

        state = gats["state"]
        candidates = by_state.get(state, [])
        if not candidates:
            continue

        best_match = None
        best_score = 0

        for inst in candidates:
            # Try matching by site_name similarity
            score = name_similarity(gats["plant_name"], inst.get("site_name", ""))

            # Also try matching by address similarity
            addr_score = name_similarity(gats["plant_name"], inst.get("address", ""))
            city_score = name_similarity(gats["plant_name"], inst.get("city", ""))
            score = max(score, addr_score, city_score)

            # Capacity match bonus (if we have both)
            if gats["capacity_kw"] and inst.get("capacity_mw"):
                gats_mw = gats["capacity_kw"] / 1000
                inst_mw = inst["capacity_mw"]
                if inst_mw > 0:
                    ratio = gats_mw / inst_mw
                    if 0.8 <= ratio <= 1.2:
                        score += 0.3  # Capacity match bonus

            if score > best_score:
                best_score = score
                best_match = inst

        # Require minimum match threshold
        min_threshold = 0.4 if gats["is_mset"] else 0.5
        if best_match and best_score >= min_threshold:
            matched_gats += 1
            patches.append({
                "id": best_match["id"],
                "owner_name": gats["owner"],
                "gats_plant": gats["plant_name"],
                "gats_uid": gats["unit_id"],
                "score": best_score,
            })
            # Remove from candidates to avoid double-matching
            by_state[state] = [c for c in by_state[state] if c["id"] != best_match["id"]]

    print(f"  GATS records with owner: {sum(1 for r in gats_records if r['owner'])}")
    print(f"  Matched to existing: {matched_gats}")
    print(f"  Patches to apply: {len(patches)}")
    print()

    if patches:
        # Show sample patches
        print("  Sample patches:")
        for p in patches[:10]:
            print(f"    {p['id'][:8]}... → {p['owner_name']} (from: {p['gats_plant'][:50]}, score: {p['score']:.2f})")
        print()

    # -------------------------------------------------------------------------
    # Phase 4: Apply patches
    # -------------------------------------------------------------------------
    if args.dry_run:
        print(f"  [DRY RUN] Would apply {len(patches)} owner_name patches")
    elif patches:
        print(f"Phase 4: Applying {len(patches)} patches...")
        applied = 0
        errors = 0
        for p in patches:
            ok = supabase_patch("solar_installations",
                               {"owner_name": p["owner_name"]},
                               {"id": f"eq.{p['id']}"})
            if ok:
                applied += 1
            else:
                errors += 1
            if (applied + errors) % 100 == 0:
                print(f"    Progress: {applied} applied, {errors} errors ({applied + errors}/{len(patches)})")
        print(f"  Applied: {applied}")
        print(f"  Errors: {errors}")
    else:
        print("  No patches to apply.")

    print()
    print("Done!")


if __name__ == "__main__":
    main()
