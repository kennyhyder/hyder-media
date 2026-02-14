#!/usr/bin/env python3
"""
Minnesota PUC DER Data Ingestion

Downloads and ingests Minnesota Public Utilities Commission Distributed Energy
Resources (DER) data into solar_installations table.

Source: https://mn.gov/puc/assets/PUBLIC%20MN%20Utility%20Reported%20DER%20through%2012-31-_Elizabeth%20Ballor_tcm14-708148.xlsx

Usage:
  python3 -u scripts/ingest-mn-puc.py              # Full ingestion
  python3 -u scripts/ingest-mn-puc.py --dry-run     # Count without ingesting
"""

import os
import sys
import json
import uuid
import argparse
import urllib.request
import urllib.parse
import time
from pathlib import Path
from openpyxl import load_workbook

from dotenv import load_dotenv

# Load env vars
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

BATCH_SIZE = 50
DATA_FILE = Path(__file__).parent.parent / "data" / "mn_puc" / "mn_der_data.xlsx"
DATA_SOURCE_NAME = "mn_puc_der"


def supabase_get(table, params, retries=3):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode())
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as e:
            if attempt < retries - 1:
                wait = 2 ** (attempt + 1)
                print(f"    Retry {attempt+1}/{retries} after {e} (waiting {wait}s)")
                time.sleep(wait)
            else:
                raise


def supabase_post(table, records):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates,return=minimal",
    }
    body = json.dumps(records).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return True
    except urllib.error.HTTPError as e:
        err = e.read().decode()[:300]
        print(f"  POST error ({e.code}): {err}")
        return False


def get_data_source_id(name):
    rows = supabase_get("solar_data_sources", {
        "select": "id",
        "name": f"eq.{name}",
    })
    if rows:
        return rows[0]["id"]
    # Create it
    supabase_post("solar_data_sources", [{
        "id": str(uuid.uuid4()),
        "name": name,
        "url": "https://mn.gov/puc/data-reports/distributed-energy-resources/",
        "description": "Minnesota PUC DER data reported by utilities",
    }])
    rows = supabase_get("solar_data_sources", {
        "select": "id",
        "name": f"eq.{name}",
    })
    return rows[0]["id"] if rows else None


def get_existing_source_ids(prefix):
    existing = set()
    offset = 0
    while True:
        rows = supabase_get("solar_installations", {
            "select": "source_record_id",
            "source_record_id": f"like.{prefix}*",
            "limit": "1000",
            "offset": str(offset),
            "order": "source_record_id",
        })
        if not rows:
            break
        for r in rows:
            existing.add(r["source_record_id"])
        if len(rows) < 1000:
            break
        offset += 1000
    return existing


def safe_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f != f:  # NaN check
            return None
        return f
    except (ValueError, TypeError):
        return None


def main():
    parser = argparse.ArgumentParser(description="Ingest MN PUC DER data")
    parser.add_argument("--dry-run", action="store_true", help="Count without ingesting")
    args = parser.parse_args()

    print("Minnesota PUC DER Data Ingestion")
    print("=" * 60)

    if not DATA_FILE.exists():
        print(f"Error: Data file not found: {DATA_FILE}")
        print("Download from: https://mn.gov/puc/data-reports/distributed-energy-resources/")
        sys.exit(1)

    print(f"  Loading: {DATA_FILE}")
    wb = load_workbook(DATA_FILE, read_only=True, data_only=True)

    # Find the data sheet
    sheet = None
    for name in wb.sheetnames:
        if "data" in name.lower() or "der" in name.lower():
            sheet = wb[name]
            break
    if not sheet:
        sheet = wb[wb.sheetnames[0]]
    print(f"  Sheet: {sheet.title}")

    # Read headers
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("  No data found!")
        return

    # Find header row (look for DER.Identifier or similar)
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        row_str = " ".join(str(c) for c in row if c)
        if "DER" in row_str or "Utility" in row_str or "Capacity" in row_str:
            header_idx = i
            break

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]
    print(f"  Headers: {headers[:10]}...")
    print(f"  Total rows: {len(rows) - header_idx - 1}")

    # Map headers by exact name matching (dots/spaces/underscores normalized)
    header_map = {}
    for i, h in enumerate(headers):
        key = h.lower().replace(".", " ").replace("_", " ").replace("(", "").replace(")", "").strip()
        header_map[key] = i

    def find_col(exact_names):
        for name in exact_names:
            norm = name.lower().replace(".", " ").replace("_", " ").replace("(", "").replace(")", "").strip()
            if norm in header_map:
                return header_map[norm]
        return None

    col_der_id = find_col(["DER.Identifier", "DER Identifier"])
    col_utility = find_col(["Utility"])
    col_eia_id = find_col(["EIA.ID", "EIA ID"])
    col_capacity = find_col(["DER.Capacity.kW.AC", "DER Capacity kW AC"])
    col_der_type = find_col(["DER.Type", "DER Type"])
    col_status = find_col(["DER.Status", "DER Status"])
    col_city = find_col(["City"])
    col_zip = find_col(["Zip.Code", "Zip Code"])
    col_customer_type = find_col(["Customer.Type", "Customer Type"])
    col_cost = find_col(["Total.Installed.Cost.without.Incentives", "Total Installed Cost without Incentives"])
    col_year_interconnected = find_col(["Year.Interconnected", "Year Interconnected"])
    col_year_decommissioned = find_col(["Year.Decommissioned.(if.applicable)", "Year Decommissioned if applicable"])
    col_doc_id = find_col(["Document.ID", "Document ID"])

    print(f"  Column mapping:")
    print(f"    DER ID: col {col_der_id}, Utility: col {col_utility}, Capacity: col {col_capacity}")
    print(f"    Type: col {col_der_type}, City: col {col_city}, Zip: col {col_zip}")
    print(f"    Cost: col {col_cost}, Year: col {col_year_interconnected}")

    if col_der_id is None and col_doc_id is None:
        print("  ERROR: Cannot find DER Identifier column!")
        return

    # Parse records
    records = []
    skipped_type = 0
    skipped_small = 0

    for row in rows[header_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        der_type = str(row[col_der_type]).strip() if col_der_type is not None and row[col_der_type] else ""
        if "solar" not in der_type.lower():
            skipped_type += 1
            continue

        capacity_kw = safe_float(row[col_capacity] if col_capacity is not None else None)
        if capacity_kw is not None and capacity_kw < 25:
            skipped_small += 1
            continue

        der_id = str(row[col_der_id]).strip() if col_der_id is not None and row[col_der_id] else ""
        doc_id = str(row[col_doc_id]).strip() if col_doc_id is not None and row[col_doc_id] else ""
        unique_key = der_id or doc_id or str(uuid.uuid4())[:8]

        utility = str(row[col_utility]).strip() if col_utility is not None and row[col_utility] else ""
        eia_id = str(row[col_eia_id]).strip() if col_eia_id is not None and row[col_eia_id] else ""
        city = str(row[col_city]).strip() if col_city is not None and row[col_city] else ""
        zip_code = str(row[col_zip]).strip() if col_zip is not None and row[col_zip] else ""
        status_raw = str(row[col_status]).strip().lower() if col_status is not None and row[col_status] else ""
        cost = safe_float(row[col_cost] if col_cost is not None else None)
        year_ic = row[col_year_interconnected] if col_year_interconnected is not None else None
        year_dc = row[col_year_decommissioned] if col_year_decommissioned is not None else None
        customer_type = str(row[col_customer_type]).strip().lower() if col_customer_type is not None and row[col_customer_type] else ""

        site_status = "active"
        if "decommission" in status_raw:
            site_status = "retired"
        elif "cancel" in status_raw or "withdraw" in status_raw:
            site_status = "canceled"

        install_date = None
        if year_ic:
            try:
                install_date = f"{int(float(str(year_ic)))}-01-01"
            except (ValueError, TypeError):
                pass

        # Determine site type
        site_type = "commercial"
        if capacity_kw and capacity_kw >= 1000:
            site_type = "utility"
        if "residential" in customer_type:
            continue  # Skip residential

        # Clean zip code
        if zip_code:
            zip_code = zip_code.split(".")[0].split("-")[0].strip()
            if len(zip_code) < 5:
                zip_code = zip_code.zfill(5)

        source_id = f"mnpuc_{unique_key}"

        record = {
            "source_record_id": source_id,
            "site_name": None,
            "site_type": site_type,
            "address": None,
            "city": city if city and city.lower() not in ("none", "nan", "") else None,
            "state": "MN",
            "zip_code": zip_code if zip_code and len(zip_code) >= 5 else None,
            "county": None,
            "latitude": None,
            "longitude": None,
            "capacity_dc_kw": capacity_kw,
            "capacity_mw": round(capacity_kw / 1000, 3) if capacity_kw else None,
            "install_date": install_date,
            "site_status": site_status,
            "installer_name": None,
            "owner_name": None,
            "developer_name": None,
            "operator_name": utility if utility and utility.lower() not in ("none", "nan", "") else None,
            "total_cost": cost,
            "data_source_id": None,
            "has_battery_storage": False,
        }
        records.append(record)

    print(f"\n  Solar records >= 25 kW (non-residential): {len(records)}")
    print(f"  Skipped (non-solar): {skipped_type}")
    print(f"  Skipped (< 25 kW): {skipped_small}")

    # Show utility breakdown
    utilities = {}
    for r in records:
        u = r.get("operator_name") or "Unknown"
        utilities[u] = utilities.get(u, 0) + 1
    print(f"\n  Top utilities:")
    for u, c in sorted(utilities.items(), key=lambda x: -x[1])[:10]:
        print(f"    {u}: {c}")

    if args.dry_run:
        print(f"\n  [DRY RUN] No records created.")
        return

    # Get data source ID
    data_source_id = get_data_source_id(DATA_SOURCE_NAME)
    for r in records:
        r["data_source_id"] = data_source_id

    # Check existing
    existing = get_existing_source_ids("mnpuc_")
    print(f"\n  Existing records: {len(existing)}")
    new_records = [r for r in records if r["source_record_id"] not in existing]
    print(f"  New records to create: {len(new_records)}")

    if not new_records:
        print("  All records already exist!")
        return

    # Insert
    created = 0
    errors = 0
    for i in range(0, len(new_records), BATCH_SIZE):
        batch = new_records[i:i + BATCH_SIZE]
        ok = supabase_post("solar_installations", batch)
        if ok:
            created += len(batch)
        else:
            errors += len(batch)
        if (i // BATCH_SIZE) % 20 == 0:
            print(f"  Progress: {created + errors}/{len(new_records)}")

    print(f"\n  Created: {created}")
    print(f"  Errors: {errors}")
    print("Done!")


if __name__ == "__main__":
    main()
