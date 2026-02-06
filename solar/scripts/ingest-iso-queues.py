#!/usr/bin/env python3
"""
ISO Interconnection Queue Data Ingestion Script

Downloads interconnection queue data from multiple ISOs (Independent System Operators)
and imports solar project records. ISO queues are the best free source for
developer/owner names on proposed and active solar projects.

ISOs covered:
  - CAISO (California) - Direct Excel download
  - NYISO (New York) - Direct Excel download
  - PJM (Mid-Atlantic, 13 states) - Direct Excel download
  - ERCOT (Texas) - Manual download required (MIS portal login)
  - MISO (Central US) - Manual download required (interactive queue)
  - SPP (Central/South) - Manual download required
  - ISO-NE (New England) - Manual download required

Usage:
  python3 -u scripts/ingest-iso-queues.py               # All ISOs with auto-download
  python3 -u scripts/ingest-iso-queues.py --iso caiso    # CAISO only
  python3 -u scripts/ingest-iso-queues.py --iso caiso nyiso pjm  # Multiple ISOs
  python3 -u scripts/ingest-iso-queues.py --redownload   # Force re-download

Source: Multiple ISO interconnection queue reports
"""

import os
import sys
import json
import uuid
import argparse
import ssl
import urllib.request
import urllib.parse
from pathlib import Path
from datetime import datetime

import openpyxl
from dotenv import load_dotenv

# Load env vars
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "data" / "iso_queues"
BATCH_SIZE = 50
MIN_CAPACITY_MW = 1  # Utility-scale threshold


# ---------------------------------------------------------------------------
# ISO Configuration
# ---------------------------------------------------------------------------
# Each ISO config defines:
#   name: short identifier (used in source_record_id prefix)
#   label: human-readable name
#   url: direct download URL (None if manual download required)
#   filename: local filename to save/look for
#   sheet_names: dict mapping status -> sheet name in the Excel file
#   header_row: 1-indexed row number containing column headers
#   column_map: mapping from ISO-specific column names -> normalized names
#   fuel_filter: list of fuel type strings that indicate solar
#   status_skip: list of status strings to skip (withdrawn, cancelled, etc.)
#   state_column: whether the ISO provides a state column (some are single-state)
#   default_state: state to use if no state column exists
#   manual_download_instructions: instructions for ISOs requiring manual download

ISO_CONFIGS = {
    "caiso": {
        "name": "caiso",
        "label": "CAISO (California ISO)",
        "url": "http://www.caiso.com/PublishedDocuments/PublicQueueReport.xlsx",
        "filename": "caiso_queue.xlsx",
        "sheet_names": {
            "active": "Grid GenerationQueue",
            "completed": "Completed Generation Projects",
            # Skip withdrawn: "Withdrawn Generation Projects"
        },
        "header_row": 4,
        "column_map": {
            "queue_id": "Queue Position",
            "project_name": "Project Name",
            "fuel_1": "Fuel-1",
            "fuel_2": "Fuel-2",
            "fuel_3": "Fuel-3",
            "capacity_mw_1": "MW-1",
            "capacity_mw_2": "MW-2",
            "capacity_mw_total": "Net MWs to Grid",
            "county": "County",
            "state": "State",
            "status": "Application Status",
            "queue_date": "Queue Date",
            "proposed_cod": "Proposed\nOn-line Date\n(as filed with IR)",
            "actual_cod": ["Current\nOn-line Date", "Actual\nOn-line Date"],
            "poi": "Station or Transmission Line",
            "interconnection_date": "Interconnection Request\nReceive Date",
        },
        "fuel_filter": ["Solar", "Solar PV", "Photovoltaic", "Storage + Solar",
                        "Solar + Storage", "Hybrid: Solar"],
        "status_skip": ["Withdrawn", "WITHDRAWN"],
        "default_state": None,  # CAISO has state column
    },
    "nyiso": {
        "name": "nyiso",
        "label": "NYISO (New York ISO)",
        "url": "https://www.nyiso.com/documents/20142/1407078/NYISO-Interconnection-Queue.xlsx",
        "filename": "nyiso_queue.xlsx",
        "sheet_names": {
            "active": "Interconnection Queue",
        },
        "header_row": 1,
        "column_map": {
            "queue_id": "Queue Pos.",
            "project_name": "Project Name",
            "developer": "Owner/Developer",
            "fuel_1": "Type/ Fuel",
            "capacity_mw_sp": "SP (MW)",
            "capacity_mw_wp": "WP (MW)",
            "county": "County",
            "state": "State",
            "status": "Status",
            "queue_date": "Date of IR",
            "proposed_cod": "Proposed In-Service Date",
            "actual_cod": "In Service Date",
            "poi": "Interconnection Point",
        },
        "fuel_filter": ["S", "Solar", "ES", "Energy Storage + Solar", "PS",
                        "Photovoltaic", "Solar + Storage"],
        "status_skip": ["W", "Withdrawn", "Cancelled"],
        "default_state": "NY",
    },
    "pjm": {
        "name": "pjm",
        "label": "PJM Interconnection",
        # PJM moved to "Queue Scope" web app - no direct download URL.
        # Export from: https://www.pjm.com/planning/services-requests/interconnection-queues
        # Or use Queue Scope tool to export to Excel
        "url": None,
        "filename": "pjm_queue.xlsx",
        "sheet_names": {
            "active": None,  # PJM uses first/only sheet
        },
        "header_row": 1,
        "column_map": {
            "queue_id": "Queue Number",
            "project_name": "Project Name",
            "developer": "Developer Name",
            "fuel_1": "Fuel",
            "capacity_mw": "MFO (MW)",
            "county": "County",
            "state": "State",
            "status": "Status",
            "queue_date": "Queue Date",
            "proposed_cod": "Proposed In-Service Date",
            "actual_cod": "Actual In-Service Date",
        },
        "fuel_filter": ["Solar", "Photovoltaic", "Solar; Battery", "Battery; Solar"],
        "status_skip": ["Withdrawn", "Deactivated"],
        "default_state": None,  # PJM has state column
        "manual_download_instructions": (
            "PJM moved to Queue Scope web app - no direct download URL.\n"
            "1. Go to: https://www.pjm.com/planning/services-requests/interconnection-queues\n"
            "2. Use Queue Scope tool to filter and export to Excel\n"
            "3. Save to solar/data/iso_queues/pjm/pjm_queue.xlsx"
        ),
    },
    "ercot": {
        "name": "ercot",
        "label": "ERCOT (Texas)",
        # ERCOT GIS report requires MIS portal login - manual download required.
        # Download from: https://www.ercot.com/mp/data-products/data-product-details?id=PG7-200-ER
        # Or via MIS: http://mis.ercot.com/misapp/GetReports.do?reportTypeId=15933&reportTitle=GIS%20Report
        # Save the latest GIS Report Excel to solar/data/iso_queues/ercot/ercot_gis.xlsx
        "url": None,
        "filename": "ercot_gis.xlsx",
        "sheet_names": {
            "active": None,  # Use first sheet
        },
        "header_row": 1,
        "column_map": {
            "queue_id": "INR Number",
            "project_name": "Project Name",
            "developer": "Developer",
            "fuel_1": "Technology",
            "capacity_mw": "Capacity (MW)",
            "county": "County",
            "status": "Status",
            "queue_date": "Screening Date",
            "proposed_cod": "Projected COD",
        },
        "fuel_filter": ["PV", "Solar", "Photovoltaic", "PVGRN", "SUN"],
        "status_skip": ["Withdrawn", "Cancelled", "Suspended"],
        "default_state": "TX",
        "manual_download_instructions": (
            "ERCOT GIS Report requires login to ERCOT MIS portal.\n"
            "1. Go to: https://www.ercot.com/mp/data-products/data-product-details?id=PG7-200-ER\n"
            "2. Download the latest GIS Report Excel file\n"
            "3. Save as: solar/data/iso_queues/ercot/ercot_gis.xlsx"
        ),
    },
    "miso": {
        "name": "miso",
        "label": "MISO (Midcontinent ISO)",
        # MISO queue is interactive: https://www.misoenergy.org/planning/generator-interconnection/GI_Queue/
        # Export from the GI Interactive Queue to Excel manually.
        # Save to solar/data/iso_queues/miso/miso_queue.xlsx
        "url": None,
        "filename": "miso_queue.xlsx",
        "sheet_names": {
            "active": None,  # Use first sheet
        },
        "header_row": 1,
        "column_map": {
            "queue_id": "Queue Number",
            "project_name": "Project Name",
            "developer": "Developer Name",
            "fuel_1": "Fuel",
            "capacity_mw": "Capacity (MW)",
            "county": "County",
            "state": "State",
            "status": "Status",
            "queue_date": "Queue Date",
            "proposed_cod": "Proposed In-Service Date",
            "poi": "POI Name",
        },
        "fuel_filter": ["Solar", "Photovoltaic", "Solar + Storage", "Hybrid - Solar"],
        "status_skip": ["Withdrawn", "Cancelled"],
        "default_state": None,  # MISO spans multiple states
        "manual_download_instructions": (
            "MISO queue requires manual export from interactive tool.\n"
            "1. Go to: https://www.misoenergy.org/planning/generator-interconnection/GI_Queue/gi-interactive-queue/\n"
            "2. Export the full queue to Excel\n"
            "3. Save as: solar/data/iso_queues/miso/miso_queue.xlsx"
        ),
    },
    "spp": {
        "name": "spp",
        "label": "SPP (Southwest Power Pool)",
        # SPP GI queue: https://opsportal.spp.org/Studies/GIActiveQueue
        # Export to Excel manually.
        "url": None,
        "filename": "spp_queue.xlsx",
        "sheet_names": {
            "active": None,
        },
        "header_row": 1,
        "column_map": {
            "queue_id": "Queue Position",
            "project_name": "Project Name",
            "developer": "Developer",
            "fuel_1": "Fuel Type",
            "capacity_mw": "Capacity (MW)",
            "county": "County",
            "state": "State",
            "status": "Status",
            "queue_date": "Queue Date",
            "proposed_cod": "Commercial Operation Date",
        },
        "fuel_filter": ["Solar", "SUN", "Photovoltaic"],
        "status_skip": ["Withdrawn", "Cancelled"],
        "default_state": None,
        "manual_download_instructions": (
            "SPP queue requires manual export.\n"
            "1. Go to: https://opsportal.spp.org/Studies/GIActiveQueue\n"
            "2. Export to Excel\n"
            "3. Save as: solar/data/iso_queues/spp/spp_queue.xlsx"
        ),
    },
    "isone": {
        "name": "isone",
        "label": "ISO-NE (New England)",
        # ISO-NE queue: https://www.iso-ne.com/system-planning/interconnection-service/interconnection-request-queue/
        # Download the queue Excel file manually.
        "url": None,
        "filename": "isone_queue.xlsx",
        "sheet_names": {
            "active": None,
        },
        "header_row": 1,
        "column_map": {
            "queue_id": "Position",
            "project_name": "Alternative Name",
            "developer": "Developer",
            "fuel_1": "Fuel Type",
            "capacity_mw": "Net MW",
            "county": "County",
            "state": "State",
            "status": "Status",
            "queue_date": "Queue Date",
            "proposed_cod": "Proposed In-Service Date",
        },
        "fuel_filter": ["SUN", "Solar", "Photovoltaic"],
        "status_skip": ["Withdrawn", "Cancelled"],
        "default_state": None,
        "manual_download_instructions": (
            "ISO-NE queue requires manual download.\n"
            "1. Go to: https://www.iso-ne.com/system-planning/interconnection-service/interconnection-request-queue/\n"
            "2. Download the Interconnection Request Queue Excel file\n"
            "3. Save as: solar/data/iso_queues/isone/isone_queue.xlsx"
        ),
    },
}

# ISOs that support automatic download
AUTO_DOWNLOAD_ISOS = [name for name, cfg in ISO_CONFIGS.items() if cfg.get("url")]


# ---------------------------------------------------------------------------
# Supabase helpers (copied from ingest-ny-sun.py pattern)
# ---------------------------------------------------------------------------

def supabase_request(method, table, data=None, params=None, headers_extra=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(f"{k}={urllib.parse.quote(str(v), safe='.*,')}" for k, v in params.items())

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates",
    }
    if headers_extra:
        headers.update(headers_extra)

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else []
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()[:200]
        print(f"  Supabase error ({e.code}): {error_body}")
        return None


def get_or_create_data_source():
    """Get or create the iso_queues data source record."""
    params = {"name": "eq.iso_queues", "select": "id"}
    existing = supabase_request("GET", "solar_data_sources", params=params)
    if existing:
        return existing[0]["id"]

    ds_id = str(uuid.uuid4())
    supabase_request("POST", "solar_data_sources", {
        "id": ds_id,
        "name": "iso_queues",
        "description": "ISO Interconnection Queues - Developer/owner data for proposed and active solar projects from regional grid operators (CAISO, NYISO, PJM, ERCOT, MISO, SPP, ISO-NE)",
        "url": "https://www.caiso.com/library/interconnection-queue-reports",
        "record_count": 0,
    })
    return ds_id


# ---------------------------------------------------------------------------
# Data helpers (copied from ingest-ny-sun.py pattern)
# ---------------------------------------------------------------------------

def safe_str(val):
    """Convert value to string, handling None and empty."""
    if val is None or val == "" or val == "N/A" or val == "n/a":
        return None
    s = str(val).strip()
    return s if s else None


def safe_float(val):
    """Convert value to float, handling None and empty."""
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace(",", "").replace("$", ""))
    except (ValueError, TypeError):
        return None


def parse_date(val):
    """Parse date string or datetime object to ISO format."""
    if not val:
        return None
    # Handle datetime objects from openpyxl
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    val = str(val).strip()
    if not val or val.lower() in ("n/a", "tbd", "none", "na", ""):
        return None
    for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S",
                "%m-%d-%Y", "%d-%b-%Y", "%b %d, %Y", "%B %d, %Y",
                "%Y/%m/%d", "%d/%m/%Y"]:
        try:
            dt = datetime.strptime(val.split(" ")[0], fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_file(url, dest_path, label=""):
    """Download a file using urllib, handling SSL and redirects."""
    if dest_path.exists():
        size_mb = dest_path.stat().st_size / 1024 / 1024
        print(f"  Found existing file ({size_mb:.1f} MB), skipping download")
        return True

    print(f"  Downloading {label} from {url}...")

    # Create SSL context that handles common certificate issues
    ctx = ssl.create_default_context()

    try:
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) SolarTrack/1.0",
            "Accept": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet, application/octet-stream, */*",
        })
        with urllib.request.urlopen(req, context=ctx, timeout=120) as resp:
            data = resp.read()
            dest_path.parent.mkdir(parents=True, exist_ok=True)
            with open(dest_path, "wb") as f:
                f.write(data)
            size_mb = len(data) / 1024 / 1024
            print(f"  Downloaded {size_mb:.1f} MB")
            return True
    except Exception as e:
        print(f"  Download failed: {e}")
        return False


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------

def read_excel_rows(filepath, sheet_name=None, header_row=1):
    """
    Read an Excel sheet and return list of dicts keyed by header names.
    If sheet_name is None, uses the first (active) sheet.
    header_row is 1-indexed.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)

    if sheet_name:
        # Try exact match first, then case-insensitive
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            # Case-insensitive fallback
            sheet_lower = sheet_name.lower()
            found = None
            for sn in wb.sheetnames:
                if sn.lower() == sheet_lower:
                    found = sn
                    break
            if found:
                ws = wb[found]
            else:
                print(f"  WARNING: Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
                wb.close()
                return []
    else:
        ws = wb.active

    rows_data = []
    headers = None
    row_num = 0

    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if row_num < header_row:
            continue
        if row_num == header_row:
            # Clean header names: strip whitespace, normalize newlines
            headers = []
            for i, h in enumerate(row):
                if h is not None:
                    hstr = str(h).strip().replace("\r\n", "\n").replace("\r", "\n")
                    headers.append(hstr)
                else:
                    headers.append(f"col_{i}")
            continue

        if headers is None:
            continue

        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        rows_data.append(record)

    wb.close()
    return rows_data


def get_column_value(row, column_map, field_name):
    """Get a value from a row dict using the column mapping.
    Supports list of column names (tries each in order)."""
    col_name = column_map.get(field_name)
    if not col_name:
        return None
    if isinstance(col_name, list):
        for cn in col_name:
            val = row.get(cn)
            if val is not None:
                return val
        return None
    return row.get(col_name)


# ---------------------------------------------------------------------------
# Fuel type matching
# ---------------------------------------------------------------------------

def is_solar_fuel(row, column_map, fuel_filter):
    """
    Check if any fuel column in the row matches solar fuel types.
    Handles CAISO-style multi-fuel columns (Fuel-1, Fuel-2) and
    single-fuel columns from other ISOs.
    """
    fuel_filter_lower = [f.lower() for f in fuel_filter]

    # Check all possible fuel columns
    for field in ["fuel_1", "fuel_2", "fuel_3"]:
        fuel_val = safe_str(get_column_value(row, column_map, field))
        if fuel_val:
            fuel_lower = fuel_val.lower()
            for pattern in fuel_filter_lower:
                if pattern in fuel_lower or fuel_lower in pattern:
                    return True
            # Also check for common solar keywords in the value
            if any(kw in fuel_lower for kw in ["solar", "photovoltaic", "pv"]):
                return True

    return False


def should_skip_status(row, column_map, status_skip):
    """Check if the project status indicates it should be skipped."""
    status = safe_str(get_column_value(row, column_map, "status"))
    if not status:
        return False
    status_lower = status.lower()
    for skip in status_skip:
        if skip.lower() in status_lower:
            return True
    return False


# ---------------------------------------------------------------------------
# Capacity extraction
# ---------------------------------------------------------------------------

def get_capacity_mw(row, column_map):
    """
    Extract capacity in MW from the row. Handles different column layouts:
    - Single MW column: capacity_mw
    - CAISO multi-fuel: MW-1, MW-2, MW Total
    - NYISO: SP (MW), WP (MW)
    Returns the best capacity value in MW.
    """
    # Try direct MW column first
    mw = safe_float(get_column_value(row, column_map, "capacity_mw"))
    if mw and mw > 0:
        return mw

    # Try total MW (CAISO)
    mw_total = safe_float(get_column_value(row, column_map, "capacity_mw_total"))
    if mw_total and mw_total > 0:
        return mw_total

    # Try MW-1 (CAISO primary fuel)
    mw_1 = safe_float(get_column_value(row, column_map, "capacity_mw_1"))
    if mw_1 and mw_1 > 0:
        return mw_1

    # Try SP (MW) for NYISO summer peak
    mw_sp = safe_float(get_column_value(row, column_map, "capacity_mw_sp"))
    if mw_sp and mw_sp > 0:
        return mw_sp

    # Try WP (MW) for NYISO winter peak
    mw_wp = safe_float(get_column_value(row, column_map, "capacity_mw_wp"))
    if mw_wp and mw_wp > 0:
        return mw_wp

    return None


# ---------------------------------------------------------------------------
# Status mapping
# ---------------------------------------------------------------------------

def map_site_status(row, column_map, sheet_status=None):
    """
    Map ISO queue status to solar_installations.site_status.
    Returns: 'proposed', 'under_construction', 'active', or None (skip).
    """
    status = safe_str(get_column_value(row, column_map, "status"))
    if not status:
        status = ""
    status_lower = status.lower()

    # Sheet-level status (CAISO completed projects sheet)
    if sheet_status == "completed":
        return "active"

    # In service / operational
    if any(kw in status_lower for kw in [
        "in service", "operational", "commercial operation",
        "completed", "energized", "active", "in-service",
        "operating",
    ]):
        return "active"

    # Under construction
    if any(kw in status_lower for kw in [
        "under construction", "construction", "building",
        "engineering", "procurement",
    ]):
        return "under_construction"

    # Active in queue / proposed (default for queue projects)
    if any(kw in status_lower for kw in [
        "active", "feasibility", "system impact", "facilities study",
        "phase", "cluster", "ia pending", "ia in progress",
        "engineering & procurement", "ready for construction",
        "queued", "pending", "study", "in progress",
    ]):
        return "proposed"

    # Default: treat as proposed (most queue projects are)
    return "proposed"


# ---------------------------------------------------------------------------
# Process a single ISO
# ---------------------------------------------------------------------------

def process_iso(iso_config, data_source_id, redownload=False):
    """Process all sheets for a single ISO and return stats."""
    name = iso_config["name"]
    label = iso_config["label"]
    iso_dir = DATA_DIR / name

    print(f"\n{'=' * 60}")
    print(f"Processing {label}")
    print(f"{'=' * 60}")

    # Download or locate the file
    iso_dir.mkdir(parents=True, exist_ok=True)
    filepath = iso_dir / iso_config["filename"]

    if iso_config.get("url"):
        if redownload and filepath.exists():
            filepath.unlink()
            print(f"  Removed old file for re-download")
        if not download_file(iso_config["url"], filepath, label):
            print(f"  SKIPPING {label} - download failed")
            return {"created": 0, "skipped": 0, "errors": 0, "solar_total": 0}
    else:
        if not filepath.exists():
            instructions = iso_config.get("manual_download_instructions", "")
            print(f"  File not found: {filepath}")
            if instructions:
                print(f"\n  Manual download required:")
                for line in instructions.split("\n"):
                    print(f"    {line}")
            print(f"\n  SKIPPING {label}")
            return {"created": 0, "skipped": 0, "errors": 0, "solar_total": 0}

    # Validate file is a real Excel file (not an iCloud placeholder)
    if filepath.stat().st_size < 1000:
        print(f"  File too small ({filepath.stat().st_size} bytes) - may be iCloud placeholder")
        print(f"  Try: brctl download {filepath}")
        return {"created": 0, "skipped": 0, "errors": 0, "solar_total": 0}

    column_map = iso_config["column_map"]
    fuel_filter = iso_config["fuel_filter"]
    status_skip = iso_config["status_skip"]
    default_state = iso_config.get("default_state")
    header_row = iso_config.get("header_row", 1)

    total_created = 0
    total_skipped = 0
    total_errors = 0
    total_solar = 0

    # Process each sheet
    for sheet_status, sheet_name in iso_config["sheet_names"].items():
        sheet_label = sheet_name or "(first sheet)"
        print(f"\n  Reading sheet: {sheet_label} (status: {sheet_status})...")

        try:
            rows = read_excel_rows(filepath, sheet_name, header_row)
        except Exception as e:
            print(f"  ERROR reading sheet: {e}")
            total_errors += 1
            continue

        if not rows:
            print(f"  No data rows found in sheet")
            continue

        # Print sample columns for debugging
        sample_cols = list(rows[0].keys())[:12]
        print(f"  Found {len(rows)} rows, {len(rows[0])} columns")
        print(f"  Sample columns: {sample_cols}")

        # Process rows
        inst_batch = []
        sheet_solar = 0
        sheet_skipped = 0
        sheet_created = 0
        sheet_errors = 0

        for row_idx, row in enumerate(rows):
            # Filter: solar fuel types only
            if not is_solar_fuel(row, column_map, fuel_filter):
                continue

            sheet_solar += 1

            # Filter: skip withdrawn/cancelled
            if should_skip_status(row, column_map, status_skip):
                sheet_skipped += 1
                continue

            # Get capacity
            capacity_mw = get_capacity_mw(row, column_map)
            if not capacity_mw or capacity_mw < MIN_CAPACITY_MW:
                sheet_skipped += 1
                continue

            # Queue ID (required for source_record_id)
            queue_id = safe_str(get_column_value(row, column_map, "queue_id"))
            if not queue_id:
                queue_id = f"row_{row_idx + 1}"

            # Clean queue_id for use in source_record_id
            queue_id_clean = str(queue_id).strip().replace(" ", "_").replace("/", "-")
            source_record_id = f"iso_{name}_{queue_id_clean}"

            # Project name
            project_name = safe_str(get_column_value(row, column_map, "project_name"))
            if not project_name:
                project_name = f"{name.upper()} Queue {queue_id}"

            # Developer / Owner
            developer_name = safe_str(get_column_value(row, column_map, "developer"))

            # Location
            state = safe_str(get_column_value(row, column_map, "state"))
            if not state and default_state:
                state = default_state
            if state:
                state = state.strip()[:2].upper()

            county = safe_str(get_column_value(row, column_map, "county"))

            # Dates
            queue_date = parse_date(get_column_value(row, column_map, "queue_date"))
            proposed_cod = parse_date(get_column_value(row, column_map, "proposed_cod"))
            actual_cod = parse_date(get_column_value(row, column_map, "actual_cod"))
            interconnection_date_val = parse_date(
                get_column_value(row, column_map, "interconnection_date")
            )

            # Status
            site_status = map_site_status(row, column_map, sheet_status)

            # Point of Interconnection
            poi = safe_str(get_column_value(row, column_map, "poi"))

            # Build installation record
            inst_id = str(uuid.uuid4())
            installation = {
                "id": inst_id,
                "source_record_id": source_record_id,
                "data_source_id": data_source_id,
                "site_name": project_name[:255] if project_name else None,
                "state": state,
                "county": county,
                "capacity_mw": round(capacity_mw, 3),
                "capacity_dc_kw": round(capacity_mw * 1000, 3),
                "site_type": "utility",
                "site_status": site_status,
                "developer_name": developer_name[:255] if developer_name else None,
                "interconnection_date": actual_cod or interconnection_date_val,
                "install_date": actual_cod,
            }

            # Add POI as address if available
            if poi:
                installation["address"] = poi[:255]

            # Add proposed COD as install_date if no actual COD
            if not actual_cod and proposed_cod:
                installation["install_date"] = proposed_cod

            inst_batch.append(installation)

            # Flush batch
            if len(inst_batch) >= BATCH_SIZE:
                res = supabase_request("POST", "solar_installations", inst_batch)
                if res is not None:
                    sheet_created += len(inst_batch)
                else:
                    sheet_errors += len(inst_batch)
                inst_batch = []

                if (sheet_created + sheet_errors) % 100 == 0:
                    print(f"    Progress: {sheet_created} created, {sheet_errors} errors, "
                          f"{sheet_skipped} skipped ({sheet_solar} solar found)")

        # Flush remaining
        if inst_batch:
            res = supabase_request("POST", "solar_installations", inst_batch)
            if res is not None:
                sheet_created += len(inst_batch)
            else:
                sheet_errors += len(inst_batch)

        print(f"\n  Sheet '{sheet_label}' results:")
        print(f"    Solar projects found: {sheet_solar}")
        print(f"    Created: {sheet_created}")
        print(f"    Skipped (withdrawn/small): {sheet_skipped}")
        print(f"    Errors: {sheet_errors}")

        total_created += sheet_created
        total_skipped += sheet_skipped
        total_errors += sheet_errors
        total_solar += sheet_solar

    return {
        "created": total_created,
        "skipped": total_skipped,
        "errors": total_errors,
        "solar_total": total_solar,
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Ingest ISO interconnection queue data")
    parser.add_argument(
        "--iso",
        nargs="+",
        choices=list(ISO_CONFIGS.keys()),
        default=AUTO_DOWNLOAD_ISOS,
        help=f"ISOs to process (default: auto-downloadable: {', '.join(AUTO_DOWNLOAD_ISOS)})",
    )
    parser.add_argument(
        "--redownload",
        action="store_true",
        help="Force re-download of Excel files",
    )
    parser.add_argument(
        "--all",
        action="store_true",
        help="Process all ISOs (including those requiring manual download)",
    )
    args = parser.parse_args()

    isos_to_process = list(ISO_CONFIGS.keys()) if args.all else args.iso

    print("ISO Interconnection Queue Ingestion Script")
    print("=" * 60)
    print(f"ISOs to process: {', '.join(isos_to_process)}")
    print(f"Minimum capacity: {MIN_CAPACITY_MW} MW (utility-scale)")
    print(f"Batch size: {BATCH_SIZE}")
    print(f"Data directory: {DATA_DIR}")

    # Get/create data source
    data_source_id = get_or_create_data_source()
    print(f"Data source ID: {data_source_id}")

    # Process each ISO
    results = {}
    for iso_name in isos_to_process:
        config = ISO_CONFIGS[iso_name]
        stats = process_iso(config, data_source_id, redownload=args.redownload)
        results[iso_name] = stats

    # Update data source record count
    total_created = sum(r["created"] for r in results.values())
    supabase_request(
        "PATCH",
        "solar_data_sources",
        {"record_count": total_created},
        params={"name": "eq.iso_queues"},
    )

    # Print summary
    print("\n" + "=" * 60)
    print("ISO Queue Ingestion Summary")
    print("=" * 60)
    print(f"{'ISO':<10} {'Solar Found':>12} {'Created':>10} {'Skipped':>10} {'Errors':>10}")
    print("-" * 52)
    for iso_name, stats in results.items():
        print(f"{iso_name:<10} {stats['solar_total']:>12} {stats['created']:>10} "
              f"{stats['skipped']:>10} {stats['errors']:>10}")
    print("-" * 52)
    totals = {
        "solar_total": sum(r["solar_total"] for r in results.values()),
        "created": total_created,
        "skipped": sum(r["skipped"] for r in results.values()),
        "errors": sum(r["errors"] for r in results.values()),
    }
    print(f"{'TOTAL':<10} {totals['solar_total']:>12} {totals['created']:>10} "
          f"{totals['skipped']:>10} {totals['errors']:>10}")

    print(f"\nISO queue ingestion complete!")

    # Note about manual ISOs
    manual_isos = [n for n in isos_to_process if not ISO_CONFIGS[n].get("url")]
    if manual_isos:
        skipped_manual = [n for n in manual_isos
                          if results[n]["solar_total"] == 0 and results[n]["created"] == 0]
        if skipped_manual:
            print(f"\nNote: {len(skipped_manual)} ISO(s) skipped (manual download required):")
            for iso_name in skipped_manual:
                instructions = ISO_CONFIGS[iso_name].get("manual_download_instructions", "")
                print(f"  - {ISO_CONFIGS[iso_name]['label']}")
                if instructions:
                    for line in instructions.split("\n"):
                        print(f"      {line}")


if __name__ == "__main__":
    main()
