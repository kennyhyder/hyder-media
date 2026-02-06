#!/usr/bin/env python3
"""
EIA-860M (Monthly) Solar Data Ingestion Script

Downloads EIA Form 860M (Preliminary Monthly Electric Generator Inventory)
and imports solar generator data including:
- Operating, Planned, Retired, and Canceled solar generators
- Plant location (state, county, lat/lon)
- Nameplate and DC capacity
- Owner/operator (entity) information
- Status codes (operating, under construction, proposed, etc.)
- Expected online dates for planned generators

This supplements the annual EIA-860 with monthly pipeline updates.
Focus: Proposed and under-construction solar projects not yet in annual data.

Source: https://www.eia.gov/electricity/data/eia860m/
"""

import os
import sys
import json
import uuid
import urllib.request
import urllib.parse
from pathlib import Path
from datetime import datetime

# openpyxl for reading Excel files
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

from dotenv import load_dotenv

# Load env vars
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "data" / "eia860m"
BATCH_SIZE = 50

# EIA-860M URL pattern: https://www.eia.gov/electricity/data/eia860m/xls/{month}_generator{year}.xlsx
# The "latest" is typically the most recent month available on the main page.
# We try the current month first, then work backwards to find the latest available file.
EIA_860M_BASE_URL = "https://www.eia.gov/electricity/data/eia860m/xls"
EIA_860M_ARCHIVE_URL = "https://www.eia.gov/electricity/data/eia860m/archive/xls"

MONTH_NAMES = [
    "january", "february", "march", "april", "may", "june",
    "july", "august", "september", "october", "november", "december"
]

# Sheets to process and their status mapping
# Each sheet maps to a site_status value for our database
SHEET_CONFIG = {
    "Operating": "active",
    "Planned": "proposed",       # Will be refined based on Status column
    "Retired": "decommissioned",
    "Canceled or Postponed": "canceled",
}

# Status code mappings for the Planned sheet
# These refine "proposed" based on the actual EIA status code
PLANNED_STATUS_MAP = {
    "(TS)": "active",              # Construction complete, not yet commercial
    "(V)": "under_construction",   # Under construction, >50% complete
    "(U)": "under_construction",   # Under construction, <=50% complete
    "(T)": "proposed",             # Regulatory approvals received, not under construction
    "(P)": "proposed",             # Planned, no regulatory approvals
    "(L)": "proposed",             # Regulatory approvals pending
    "(OT)": "proposed",           # Other
}

# Operating status refinements
OPERATING_STATUS_MAP = {
    "(OP)": "active",              # Operating
    "(OA)": "active",              # Out of service, expected to return
    "(OS)": "inactive",            # Out of service, NOT expected to return
    "(SB)": "active",              # Standby/backup
}


def supabase_request(method, table, data=None, params=None, headers_extra=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(f"{k}={urllib.parse.quote(str(v), safe='.*,')}" for k, v in params.items())

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates",
    }
    if headers_extra:
        headers.update(headers_extra)

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else []
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()[:200]
        print(f"  Supabase error ({e.code}): {error_body}")
        return None


def safe_float(val):
    """Convert value to float, handling None, empty strings, and spaces."""
    if val is None or val == "" or val == " ":
        return None
    try:
        return float(str(val).replace(",", "").replace("$", "").strip())
    except (ValueError, TypeError):
        return None


def safe_str(val):
    """Convert value to string, handling None and empty."""
    if val is None or val == "" or val == "N/A" or val == "n/a":
        return None
    s = str(val).strip()
    return s if s else None


def safe_int(val):
    """Safely convert to int."""
    if val is None or val == "" or val == " ":
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def parse_date(val):
    """Parse date string to ISO format."""
    if not val:
        return None
    val = str(val).strip()
    for fmt in ["%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%Y-%m-%dT%H:%M:%S"]:
        try:
            dt = datetime.strptime(val.split(" ")[0], fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def build_date(month, year):
    """Build date string from month/year values."""
    y = safe_int(year)
    m = safe_int(month)
    if y and m:
        return f"{y}-{m:02d}-01"
    elif y:
        return f"{y}-01-01"
    return None


def get_or_create_data_source():
    """Get or create the EIA-860M data source record."""
    params = {"name": "eq.eia860m", "select": "id"}
    existing = supabase_request("GET", "solar_data_sources", params=params)
    if existing:
        return existing[0]["id"]

    ds_id = str(uuid.uuid4())
    supabase_request("POST", "solar_data_sources", {
        "id": ds_id,
        "name": "eia860m",
        "description": "EIA Form 860M - Preliminary Monthly Electric Generator Inventory (Solar)",
        "url": "https://www.eia.gov/electricity/data/eia860m/",
        "record_count": 0,
    })
    return ds_id


def find_latest_file_url():
    """Find the URL of the latest available EIA-860M Excel file.

    Tries the current year/month and works backwards. The EIA page
    typically has the latest file at the main xls/ path, and older
    files under archive/xls/.
    """
    now = datetime.now()
    year = now.year
    month = now.month

    # Try current year, working backwards from current month
    # Then try previous year if needed
    attempts = []
    for y in [year, year - 1]:
        start_month = month if y == year else 12
        for m in range(start_month, 0, -1):
            name = f"{MONTH_NAMES[m - 1]}_generator{y}.xlsx"
            # Try main URL first (latest month), then archive
            attempts.append((f"{EIA_860M_BASE_URL}/{name}", name))
            attempts.append((f"{EIA_860M_ARCHIVE_URL}/{name}", name))

    for url, name in attempts:
        try:
            req = urllib.request.Request(url, method="HEAD")
            with urllib.request.urlopen(req) as resp:
                content_type = resp.headers.get("Content-Type", "")
                # EIA returns HTML for missing files, so check content type
                if "spreadsheet" in content_type or "excel" in content_type or "octet-stream" in content_type:
                    print(f"  Found latest file: {name}")
                    return url, name
        except urllib.error.HTTPError:
            continue
        except Exception:
            continue

    # Fallback: try downloading and checking the file type
    print("  HEAD check inconclusive, trying direct download...")
    for url, name in attempts[:6]:  # Only try first 6 (3 months)
        try:
            import tempfile
            tmp = Path(tempfile.mktemp(suffix=".xlsx"))
            urllib.request.urlretrieve(url, tmp)
            # Check if it's actually an Excel file (not HTML error page)
            with open(tmp, "rb") as f:
                magic = f.read(4)
            tmp.unlink()
            if magic[:2] == b"PK":  # XLSX files are ZIP archives
                print(f"  Found latest file: {name}")
                return url, name
        except Exception:
            if tmp.exists():
                tmp.unlink()
            continue

    print("ERROR: Could not find any available EIA-860M file!")
    sys.exit(1)


def download_data():
    """Download the latest EIA-860M Excel file."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Check for existing files
    existing = sorted(DATA_DIR.glob("*_generator*.xlsx"))
    if existing:
        latest = existing[-1]
        size_mb = latest.stat().st_size / 1024 / 1024
        print(f"  Found existing file: {latest.name} ({size_mb:.1f} MB)")
        print(f"  Delete it to force re-download")
        return latest

    url, filename = find_latest_file_url()
    filepath = DATA_DIR / filename

    print(f"  Downloading from {url}...")
    urllib.request.urlretrieve(url, filepath)
    size_mb = filepath.stat().st_size / 1024 / 1024
    print(f"  Downloaded {size_mb:.1f} MB → {filepath.name}")
    return filepath


def read_sheet(filepath, sheet_name, header_row=3):
    """Read an Excel sheet, returning list of dicts keyed by header names.

    EIA-860M format:
    - Row 1: Title (e.g., "Inventory of Operating Generators as of December 2025")
    - Row 2: Blank
    - Row 3: Column headers
    - Row 4+: Data
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  WARNING: Sheet '{sheet_name}' not found in workbook")
        wb.close()
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < header_row:
        return []

    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_row - 1])]
    data = []
    for row in rows[header_row:]:
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val
        data.append(record)
    return data


def determine_site_status(sheet_name, row):
    """Determine site_status based on sheet and Status column value."""
    status_raw = safe_str(row.get("Status"))

    if sheet_name == "Operating":
        if status_raw:
            prefix = status_raw.split(")")[0] + ")" if ")" in status_raw else ""
            return OPERATING_STATUS_MAP.get(prefix, "active")
        return "active"

    elif sheet_name == "Planned":
        if status_raw:
            prefix = status_raw.split(")")[0] + ")" if ")" in status_raw else ""
            return PLANNED_STATUS_MAP.get(prefix, "proposed")
        return "proposed"

    elif sheet_name == "Retired":
        return "decommissioned"

    elif sheet_name == "Canceled or Postponed":
        return "canceled"

    return "unknown"


def determine_site_type(capacity_mw):
    """Determine site type based on capacity."""
    if capacity_mw is None:
        return "utility"
    if capacity_mw >= 1.0:
        return "utility"
    return "commercial"


def process_sheet(filepath, sheet_name, data_source_id):
    """Process one sheet of the EIA-860M workbook, returning solar records."""
    print(f"\n  Processing sheet: {sheet_name}...")
    rows = read_sheet(filepath, sheet_name)

    if not rows:
        print(f"    No data found")
        return []

    print(f"    {len(rows)} total rows")

    records = []
    solar_count = 0
    skipped_small = 0

    for row in rows:
        # Filter: Solar only (Energy Source Code = SUN)
        energy_source = safe_str(row.get("Energy Source Code"))
        if energy_source != "SUN":
            continue

        solar_count += 1

        # Get capacity
        capacity_mw = safe_float(row.get("Nameplate Capacity (MW)"))
        capacity_dc_mw = safe_float(row.get("DC Net Capacity (MW)"))

        # Use nameplate as primary, DC as fallback
        primary_mw = capacity_mw if capacity_mw is not None else capacity_dc_mw

        # Filter: >= 1 MW (EIA-860M only reports plants >= 1 MW)
        # But some entries may be < 1 MW generators within a larger plant
        # Keep all solar entries since EIA already filters to >= 1 MW plants
        if primary_mw is not None and primary_mw < 0.025:
            # Skip very small entries (< 25 kW) that may be data errors
            skipped_small += 1
            continue

        plant_id = safe_str(row.get("Plant ID"))
        gen_id = safe_str(row.get("Generator ID"))

        if not plant_id:
            continue

        source_record_id = f"eia860m_{plant_id}_{gen_id}" if gen_id else f"eia860m_{plant_id}"

        # Location
        lat = safe_float(row.get("Latitude"))
        lon = safe_float(row.get("Longitude"))
        state = safe_str(row.get("Plant State"))
        county = safe_str(row.get("County"))

        # Entity (owner/operator)
        entity_name = safe_str(row.get("Entity Name"))
        plant_name = safe_str(row.get("Plant Name"))

        # Status
        site_status = determine_site_status(sheet_name, row)
        site_type = determine_site_type(primary_mw)

        # Dates - differ by sheet
        if sheet_name in ("Operating", "Retired"):
            install_date = build_date(
                row.get("Operating Month"),
                row.get("Operating Year")
            )
        elif sheet_name == "Planned":
            # For planned: use expected operation date as install_date
            install_date = build_date(
                row.get("Planned Operation Month"),
                row.get("Planned Operation Year")
            )
        else:
            install_date = None

        # Build record
        record = {
            "id": str(uuid.uuid4()),
            "source_record_id": source_record_id,
            "data_source_id": data_source_id,
            "site_name": plant_name,
            "state": state,
            "county": county,
            "latitude": lat,
            "longitude": lon,
            "capacity_mw": round(primary_mw, 6) if primary_mw is not None else None,
            "capacity_ac_kw": round(capacity_mw * 1000, 2) if capacity_mw is not None else None,
            "capacity_dc_kw": round(capacity_dc_mw * 1000, 2) if capacity_dc_mw is not None else None,
            "site_type": site_type,
            "site_status": site_status,
            "owner_name": entity_name,
            "operator_name": entity_name,
            "install_date": install_date,
        }

        records.append(record)

    print(f"    {solar_count} solar (SUN) rows found")
    if skipped_small > 0:
        print(f"    {skipped_small} skipped (< 25 kW)")
    print(f"    {len(records)} records to insert")

    return records


def insert_records(records, data_source_id):
    """Insert records into Supabase in batches."""
    if not records:
        return 0, 0

    created = 0
    errors = 0

    for i in range(0, len(records), BATCH_SIZE):
        batch = records[i:i + BATCH_SIZE]
        res = supabase_request("POST", "solar_installations", batch)
        if res is not None:
            created += len(batch)
        else:
            errors += len(batch)

        if (created + errors) % 100 == 0 or (i + BATCH_SIZE) >= len(records):
            print(f"    Progress: {created} created, {errors} errors ({i + len(batch)}/{len(records)})")

    return created, errors


def main():
    print("EIA-860M Monthly Solar Data Ingestion Script")
    print("=" * 60)

    data_source_id = get_or_create_data_source()
    print(f"Data source ID: {data_source_id}")

    # Download data
    print("\nDownloading EIA-860M data...")
    filepath = download_data()

    # Process each sheet
    all_records = []
    sheet_stats = {}

    for sheet_name in SHEET_CONFIG:
        records = process_sheet(filepath, sheet_name, data_source_id)
        sheet_stats[sheet_name] = len(records)
        all_records.extend(records)

    print(f"\n  Total solar records to insert: {len(all_records)}")

    # Also process Puerto Rico sheets if they exist
    for pr_sheet in ["Operating_PR", "Planned_PR", "Retired_PR"]:
        base_name = pr_sheet.replace("_PR", "")
        if base_name in SHEET_CONFIG:
            records = process_sheet(filepath, pr_sheet, data_source_id)
            if records:
                sheet_stats[pr_sheet] = len(records)
                all_records.extend(records)
                print(f"    +{len(records)} from {pr_sheet}")

    # Insert all records
    print(f"\nInserting {len(all_records)} records into solar_installations...")
    created, errors = insert_records(all_records, data_source_id)

    # Update data source record count
    supabase_request(
        "PATCH",
        "solar_data_sources",
        {"record_count": created},
        params={"name": "eq.eia860m"},
    )

    # Summary
    print("\n" + "=" * 60)
    print("EIA-860M ingestion complete!")
    print(f"\n  Records by sheet:")
    for sheet, count in sheet_stats.items():
        print(f"    {sheet}: {count}")
    print(f"\n  Total created: {created}")
    print(f"  Total errors: {errors}")
    print(f"  (Duplicates are silently ignored via source_record_id UNIQUE constraint)")


if __name__ == "__main__":
    main()
