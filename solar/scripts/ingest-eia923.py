#!/usr/bin/env python3
"""
Ingest EIA-923 Generation Data — Monthly/annual solar plant generation.

Downloads EIA Form 923 data and cross-references with existing installations
by EIA Plant ID. Calculates capacity factors and identifies underperforming sites.

Adds annual_generation_mwh and capacity_factor fields to solar_installations.

Data source: https://www.eia.gov/electricity/data/eia923/

Usage:
  python3 -u scripts/ingest-eia923.py                    # Download + enrich
  python3 -u scripts/ingest-eia923.py --dry-run           # Preview
  python3 -u scripts/ingest-eia923.py --skip-download     # Use existing file
  python3 -u scripts/ingest-eia923.py --year 2023         # Specific year
"""

import os
import sys
import json
import time
import argparse
import urllib.request
import urllib.parse
import zipfile
import io
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

WORKERS = 20
DATA_DIR = Path(__file__).parent.parent / "data" / "eia923"


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def supabase_get(table, params, retries=3):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode())
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as e:
            if attempt < retries - 1:
                time.sleep(2 ** (attempt + 1))
            else:
                raise


def supabase_patch(table, data, params):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    body = json.dumps(data).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="PATCH")
    try:
        with urllib.request.urlopen(req) as resp:
            return True
    except urllib.error.HTTPError:
        return False


def _do_patch(args):
    inst_id, patch = args
    return supabase_patch("solar_installations", patch, {"id": f"eq.{inst_id}"})


# ---------------------------------------------------------------------------
# Download EIA-923
# ---------------------------------------------------------------------------

def download_eia923(year=2024):
    """Download EIA-923 Excel file."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    # Try the final release first, then early release
    urls = [
        f"https://www.eia.gov/electricity/data/eia923/xls/EIA923_Schedules_2_3_4_5_M_12_{year}_Final_Revision.xlsx",
        f"https://www.eia.gov/electricity/data/eia923/xls/EIA923_Schedules_2_3_4_5_M_12_{year}_Final.xlsx",
        f"https://www.eia.gov/electricity/data/eia923/xls/EIA923_Schedules_2_3_4_5_M_12_{year}_Early_Release.xlsx",
        f"https://www.eia.gov/electricity/data/eia923/xls/EIA923_Schedules_2_3_4_5_M_12_{year}.xlsx",
    ]

    for url in urls:
        fname = url.split("/")[-1]
        fpath = DATA_DIR / fname
        if fpath.exists():
            print(f"  Using existing file: {fpath}")
            return fpath

    # Try downloading (validate file is actually XLSX, not HTML error page)
    for url in urls:
        fname = url.split("/")[-1]
        fpath = DATA_DIR / fname
        print(f"  Trying: {url}")
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "SolarTrack/1.0"})
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = resp.read()
            # Validate: XLSX files start with PK (ZIP magic bytes), not HTML
            if data[:2] == b'PK':
                with open(fpath, "wb") as f:
                    f.write(data)
                print(f"  Downloaded: {fpath} ({len(data)/1024/1024:.1f} MB)")
                return fpath
            else:
                print(f"    Not an Excel file (got HTML/other), skipping")
                continue
        except urllib.error.HTTPError as e:
            print(f"    {e.code}: {e.reason}")
            continue
        except Exception as e:
            print(f"    Error: {e}")
            continue

    # Try ZIP format (archive format is the primary download since ~2025)
    zip_urls = [
        f"https://www.eia.gov/electricity/data/eia923/archive/xls/f923_{year}.zip",
        f"https://www.eia.gov/electricity/data/eia923/xls/EIA923_Schedules_2_3_4_5_M_12_{year}.zip",
        f"https://www.eia.gov/electricity/data/eia923/xls/f923_{year}.zip",
    ]
    for url in zip_urls:
        print(f"  Trying ZIP: {url}")
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "SolarTrack/1.0"})
            with urllib.request.urlopen(req, timeout=120) as resp:
                zip_data = resp.read()
            with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
                # Prefer "Schedules_2_3_4_5" file (generation data), not Schedule_8 (environmental)
                xlsx_files = [n for n in zf.namelist() if n.endswith(".xlsx")]
                target = None
                for name in xlsx_files:
                    if "Schedules_2_3_4_5" in name or "Schedule_2" in name:
                        target = name
                        break
                if not target and xlsx_files:
                    # Fallback: pick largest xlsx (generation file is ~20MB vs ~2MB)
                    target = max(xlsx_files, key=lambda n: zf.getinfo(n).file_size)
                if target:
                    fpath = DATA_DIR / target
                    with open(fpath, "wb") as f:
                        f.write(zf.read(target))
                    print(f"  Extracted: {fpath} ({zf.getinfo(target).file_size / 1024 / 1024:.1f} MB)")
                    return fpath
        except Exception as e:
            print(f"    Error: {e}")
            continue

    return None


# ---------------------------------------------------------------------------
# Parse EIA-923
# ---------------------------------------------------------------------------

def load_eia923_solar(file_path):
    """Load solar generation data from EIA-923 Excel."""
    import openpyxl

    print(f"  Loading EIA-923 from {file_path}...")
    wb = openpyxl.load_workbook(str(file_path), read_only=True)

    # Find the generation sheet (Schedule 2-3-4-5 or Page 1)
    target_sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if "page 1" in name_lower or "generation" in name_lower or "schedule" in name_lower:
            target_sheet = name
            break

    if not target_sheet:
        # Fall back to first sheet
        target_sheet = wb.sheetnames[0]

    print(f"  Using sheet: {target_sheet}")
    ws = wb[target_sheet]

    # Find header row (look for "Plant Id" or "Plant Code")
    header = None
    header_row_num = 0
    for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and any(str(c or "").strip().lower() in ("plant id", "plant code", "plant_id") for c in row):
            header = [str(c or "").strip() for c in row]
            header_row_num = row_num
            break

    if not header:
        print("  ERROR: Could not find header row")
        wb.close()
        return {}

    print(f"  Header found at row {header_row_num}: {len(header)} columns")

    # Find column indices
    def find_col(names):
        for i, h in enumerate(header):
            if h.lower().strip() in [n.lower() for n in names]:
                return i
        return None

    plant_id_col = find_col(["Plant Id", "Plant Code", "Plant_Id", "PLANT ID"])
    plant_name_col = find_col(["Plant Name", "PLANT NAME"])

    # Monthly generation columns (Jan-Dec)
    month_cols = []
    month_names = ["January", "February", "March", "April", "May", "June",
                   "July", "August", "September", "October", "November", "December"]
    # Also look for "Netgen January", "Net Generation January", etc.
    for i, h in enumerate(header):
        h_lower = h.lower().strip()
        for month_name in month_names:
            if month_name.lower() in h_lower and ("net" in h_lower or "generation" in h_lower or month_name.lower() == h_lower):
                month_cols.append(i)
                break

    # Look for annual total column
    annual_col = find_col(["YEAR", "Annual", "Net Generation\n(Megawatthours)", "Netgen"])

    # If we can't find monthly columns, look for a different pattern
    if not month_cols:
        # Try to find columns that look like generation data
        for i, h in enumerate(header):
            h_lower = h.lower().strip()
            if any(m.lower() in h_lower for m in month_names):
                month_cols.append(i)

    # Also look for "Reported Prime Mover" or "AER Fuel Type Code" to filter solar
    fuel_col = find_col(["AER\nFuel Type Code", "AER Fuel Type Code", "Reported\nFuel Type Code",
                          "Reported Fuel Type Code", "FUEL TYPE", "Fuel Type"])
    prime_mover_col = find_col(["Reported\nPrime Mover", "Reported Prime Mover", "Prime Mover",
                                 "PRIME MOVER"])

    print(f"  Plant ID col: {plant_id_col}, Fuel col: {fuel_col}, PM col: {prime_mover_col}")
    print(f"  Monthly cols found: {len(month_cols)}, Annual col: {annual_col}")

    # Parse data
    plants = {}  # plant_id -> {plant_name, annual_mwh, monthly: [12 values]}
    row_count = 0

    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        row_count += 1
        if not row:
            continue

        # Check if solar
        is_solar = False
        if fuel_col is not None and len(row) > fuel_col:
            fuel = str(row[fuel_col] or "").strip().upper()
            if fuel in ("SUN", "SOL", "SOLAR"):
                is_solar = True
        if prime_mover_col is not None and len(row) > prime_mover_col:
            pm = str(row[prime_mover_col] or "").strip().upper()
            if pm in ("PV", "CP"):  # PV = Photovoltaic, CP = Concentrated Solar
                is_solar = True
        if not is_solar:
            continue

        if plant_id_col is None or len(row) <= plant_id_col:
            continue
        try:
            pid = int(row[plant_id_col])
        except (ValueError, TypeError):
            continue

        plant_name = str(row[plant_name_col] or "").strip() if plant_name_col and len(row) > plant_name_col else None

        # Sum monthly generation
        monthly_mwh = []
        for mc in month_cols:
            if mc < len(row):
                try:
                    val = float(row[mc]) if row[mc] is not None else 0
                except (ValueError, TypeError):
                    val = 0
                monthly_mwh.append(val)
            else:
                monthly_mwh.append(0)

        annual_mwh = sum(monthly_mwh) if monthly_mwh else 0

        # Try annual column as fallback
        if annual_mwh == 0 and annual_col and len(row) > annual_col:
            try:
                annual_mwh = float(row[annual_col]) if row[annual_col] else 0
            except (ValueError, TypeError):
                pass

        if annual_mwh <= 0:
            continue

        # Aggregate by plant (multiple generators per plant)
        if pid in plants:
            plants[pid]["annual_mwh"] += annual_mwh
        else:
            plants[pid] = {
                "plant_name": plant_name,
                "annual_mwh": annual_mwh,
            }

    wb.close()
    print(f"  Parsed {row_count} rows, found {len(plants)} solar plants with generation data")
    return plants


# ---------------------------------------------------------------------------
# Load installations and match
# ---------------------------------------------------------------------------

def load_installations_by_eia_id():
    """Load installations that have EIA-based source_record_ids."""
    print("  Loading installations with EIA Plant IDs...")
    records = []
    offset = 0
    while True:
        batch = supabase_get("solar_installations", {
            "select": "id,source_record_id,capacity_mw,state",
            "source_record_id": "like.eia860_*",
            "limit": "1000",
            "offset": str(offset),
            "order": "id",
        })
        if not batch:
            break
        records.extend(batch)
        if len(batch) < 1000:
            break
        offset += len(batch)

    # Also load EIA-860M records
    offset = 0
    while True:
        batch = supabase_get("solar_installations", {
            "select": "id,source_record_id,capacity_mw,state",
            "source_record_id": "like.eia860m_*",
            "limit": "1000",
            "offset": str(offset),
            "order": "id",
        })
        if not batch:
            break
        records.extend(batch)
        if len(batch) < 1000:
            break
        offset += len(batch)

    # Build plant_id -> installations map
    by_eia = {}
    for r in records:
        sid = r.get("source_record_id", "")
        parts = sid.split("_")
        if len(parts) >= 2:
            try:
                pid = int(parts[1])
                by_eia.setdefault(pid, []).append(r)
            except ValueError:
                pass

    print(f"  Loaded {len(records)} EIA installations ({len(by_eia)} unique plant IDs)")
    return by_eia


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Ingest EIA-923 solar generation data")
    parser.add_argument("--dry-run", action="store_true", help="Preview without patching")
    parser.add_argument("--skip-download", action="store_true", help="Use existing file")
    parser.add_argument("--year", type=int, default=2024, help="Data year (default: 2024)")
    args = parser.parse_args()

    print("EIA-923 Solar Generation Ingestion")
    print("=" * 60)
    print(f"  Year: {args.year}")
    print(f"  Dry run: {args.dry_run}")

    # Step 1: Get EIA-923 file
    if args.skip_download:
        # Find existing file
        existing = list(DATA_DIR.glob("EIA923*.xlsx"))
        if not existing:
            print("  ERROR: No existing EIA-923 file found")
            sys.exit(1)
        file_path = existing[0]
        print(f"  Using existing file: {file_path}")
    else:
        try:
            import openpyxl
        except ImportError:
            print("  ERROR: openpyxl required. Run: pip3 install openpyxl")
            sys.exit(1)
        file_path = download_eia923(args.year)
        if not file_path:
            print("  ERROR: Could not download EIA-923 data")
            print("  Try manually downloading from: https://www.eia.gov/electricity/data/eia923/")
            print(f"  Place the .xlsx file in: {DATA_DIR}/")
            sys.exit(1)

    # Step 2: Parse generation data
    plants = load_eia923_solar(file_path)
    if not plants:
        print("  No solar generation data found in file")
        return

    # Step 3: Load our installations
    by_eia = load_installations_by_eia_id()

    # Step 4: Match and build patches
    print(f"\nMatching EIA-923 plants to installations...")
    patches = []
    matched_plants = 0
    matched_installations = 0

    hours_per_year = 8760  # Average hours in a year

    for pid, gen_data in plants.items():
        insts = by_eia.get(pid, [])
        if not insts:
            continue

        matched_plants += 1
        annual_mwh = gen_data["annual_mwh"]

        for inst in insts:
            matched_installations += 1
            patch = {"annual_generation_mwh": round(annual_mwh, 1)}

            # Calculate capacity factor if we have capacity
            cap_mw = inst.get("capacity_mw")
            if cap_mw and cap_mw > 0:
                max_mwh = cap_mw * hours_per_year
                cf = annual_mwh / max_mwh if max_mwh > 0 else None
                if cf is not None:
                    # Sanity check: CF should be 0-1 (solar rarely exceeds 0.35)
                    if 0 < cf <= 1.0:
                        patch["capacity_factor"] = round(cf, 4)
                    elif cf > 1.0:
                        # Multiple generators per plant — divide by # installations
                        n_insts = len(insts)
                        per_inst_mwh = annual_mwh / n_insts
                        patch["annual_generation_mwh"] = round(per_inst_mwh, 1)
                        cf_adj = per_inst_mwh / max_mwh
                        if 0 < cf_adj <= 1.0:
                            patch["capacity_factor"] = round(cf_adj, 4)

            patches.append((inst["id"], patch))

    print(f"  Matched plants: {matched_plants}")
    print(f"  Matched installations: {matched_installations}")
    print(f"  Unmatched EIA-923 plants: {len(plants) - matched_plants}")

    # Stats
    cf_values = [p["capacity_factor"] for _, p in patches if "capacity_factor" in p]
    if cf_values:
        avg_cf = sum(cf_values) / len(cf_values)
        min_cf = min(cf_values)
        max_cf = max(cf_values)
        print(f"\n  Capacity factor stats:")
        print(f"    Average: {avg_cf:.1%}")
        print(f"    Min: {min_cf:.1%}")
        print(f"    Max: {max_cf:.1%}")
        print(f"    Plants with CF: {len(cf_values)}")

        # Flag underperformers (< 10% CF — likely degraded or issues)
        underperformers = [(pid, p) for pid, p in patches if p.get("capacity_factor", 1) < 0.10]
        if underperformers:
            print(f"\n  Underperforming sites (CF < 10%): {len(underperformers)}")
            for inst_id, p in underperformers[:5]:
                print(f"    {inst_id}: CF={p.get('capacity_factor', 0):.1%}, {p.get('annual_generation_mwh', 0):.0f} MWh")

    if args.dry_run:
        print(f"\n  [DRY RUN] Would patch {len(patches)} installations")
        for inst_id, patch in patches[:10]:
            print(f"    {inst_id}: {patch}")
        return

    if not patches:
        print("\n  No patches to apply.")
        return

    # Apply patches
    print(f"\nApplying {len(patches)} patches ({WORKERS} workers)...")
    applied = 0
    errors = 0

    with ThreadPoolExecutor(max_workers=WORKERS) as executor:
        futures = {executor.submit(_do_patch, item): item for item in patches}
        for future in as_completed(futures):
            if future.result():
                applied += 1
            else:
                errors += 1
            if (applied + errors) % 500 == 0:
                print(f"    Progress: {applied} applied, {errors} errors")

    print(f"\n  Applied: {applied}")
    print(f"  Errors: {errors}")
    print("\nDone!")


if __name__ == "__main__":
    main()
