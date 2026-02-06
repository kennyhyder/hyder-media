#!/usr/bin/env python3
"""
EIA-860 Data Enrichment Script

1. Extract retirement events from the "Retired and Canceled" sheet → solar_site_events
2. Import owner data from Schedule 4 "Ownership" sheet → update solar_installations.owner_name
3. Cross-reference EIA-860 plant codes to propagate owner/operator to USPVDB records
"""

import os
import sys
import json
import uuid
import urllib.request
import urllib.parse
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "data" / "eia860_2024"
SOLAR_FILE = DATA_DIR / "3_3_Solar_Y2024.xlsx"
OWNER_FILE = DATA_DIR / "4___Owner_Y2024.xlsx"


def supabase_request(method, table, data=None, params=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items())

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates",
    }
    if headers_extra:
        headers.update(headers_extra)

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else []
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()[:200]
        print(f"  Supabase error ({e.code}): {error_body}")
        return None


def get_data_source_id():
    """Get the EIA-860 data source ID."""
    params = {"name": "eq.eia860", "select": "id"}
    result = supabase_request("GET", "solar_data_sources", params=params)
    if result:
        return result[0]["id"]
    return None


def extract_retirement_events(data_source_id):
    """Parse the Retired and Canceled sheet and create site_events."""
    print("\n" + "=" * 60)
    print("1. EIA-860 Retirement Events")
    print("=" * 60)

    wb = openpyxl.load_workbook(SOLAR_FILE, read_only=True)
    ws = wb['Retired and Canceled']

    # Row 2 = headers, data starts row 3
    # Col 3=Plant Code, 7=Generator ID, 8=Status, 4=Plant Name, 5=State
    # 13=Nameplate MW, 16=Operating Month, 17=Operating Year
    # 18=Retirement Month, 19=Retirement Year

    events = []
    matched = 0
    not_found = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        plant_code = row[2]  # Col 3
        gen_id = row[6]      # Col 7
        status = row[7]      # Col 8
        plant_name = row[3]  # Col 4
        state = row[4]       # Col 5
        capacity_mw = row[12]  # Col 13
        ret_month = row[17]  # Col 18
        ret_year = row[18]   # Col 19

        if not plant_code or not gen_id:
            continue

        # Build source_record_id to match our DB
        source_record_id = f"eia860_{plant_code}_{gen_id}"

        # Find installation in DB
        params = {
            "source_record_id": f"eq.{source_record_id}",
            "select": "id,site_name",
        }
        result = supabase_request("GET", "solar_installations", params=params)

        if not result:
            not_found += 1
            continue

        inst_id = result[0]["id"]
        matched += 1

        # Build event date (handle whitespace/empty values)
        event_date = None
        ret_year_str = str(ret_year).strip() if ret_year else ''
        ret_month_str = str(ret_month).strip() if ret_month else ''
        if ret_year_str and ret_year_str.replace('.', '', 1).isdigit():
            if ret_month_str and ret_month_str.replace('.', '', 1).isdigit():
                event_date = f"{int(float(ret_year_str))}-{int(float(ret_month_str)):02d}-01"
            else:
                event_date = f"{int(float(ret_year_str))}-01-01"

        # Event type based on status
        event_type = "retirement" if status == "RE" else "cancellation"

        event = {
            "id": str(uuid.uuid4()),
            "installation_id": inst_id,
            "event_type": event_type,
            "event_date": event_date,
            "description": f"Generator {gen_id} at {plant_name} ({state}) - {event_type} per EIA-860 2024",
            "old_capacity_kw": round(float(capacity_mw) * 1000, 1) if capacity_mw else None,
            "data_source_id": data_source_id,
        }
        events.append(event)

        # Also update the installation's decommission_date and site_status
        update_data = {}
        if event_date and event_type == "retirement":
            update_data["decommission_date"] = event_date
            update_data["site_status"] = "retired"
        elif event_type == "cancellation":
            update_data["site_status"] = "canceled"

        if update_data:
            supabase_request(
                "PATCH",
                "solar_installations",
                update_data,
                params={"id": f"eq.{inst_id}"},
            )

    wb.close()

    # Insert events in batches
    created = 0
    for i in range(0, len(events), 50):
        batch = events[i:i + 50]
        res = supabase_request("POST", "solar_site_events", batch)
        if res is not None:
            created += len(batch)

    print(f"  Retired/Canceled records in EIA-860: {matched + not_found}")
    print(f"  Matched to DB: {matched}")
    print(f"  Not found in DB: {not_found}")
    print(f"  Events created: {created}")


def import_owner_data(data_source_id):
    """Parse Schedule 4 Ownership and update owner_name on installations."""
    print("\n" + "=" * 60)
    print("2. EIA-860 Owner Data Import")
    print("=" * 60)

    wb = openpyxl.load_workbook(OWNER_FILE, read_only=True)
    ws = wb['Ownership']

    # Row 2 = headers, data starts row 3
    # Col 1=Utility ID, 3=Plant Code, 6=Generator ID, 7=Status
    # Col 8=Owner Name, 9=Owner Street, 10=Owner City, 11=Owner State, 12=Owner Zip
    # Col 14=Percent Owned

    # Build plant_code → owner mapping (take highest % owner)
    plant_owners = {}  # plant_code -> {owner_name, percent}

    for row in ws.iter_rows(min_row=3, values_only=True):
        plant_code = row[2]   # Col 3
        gen_id = row[5]       # Col 6
        owner_name = row[7]   # Col 8
        percent = row[13]     # Col 14

        if not plant_code or not owner_name:
            continue

        key = f"{plant_code}_{gen_id}"
        try:
            pct = float(percent) if percent and str(percent).strip() else 0
        except (ValueError, TypeError):
            pct = 0

        if key not in plant_owners or pct > plant_owners[key]["percent"]:
            plant_owners[key] = {
                "owner_name": str(owner_name).strip(),
                "percent": pct,
            }

    wb.close()

    print(f"  Found {len(plant_owners)} generator ownership records")

    # Update installations
    updated = 0
    not_found = 0

    for key, info in plant_owners.items():
        source_record_id = f"eia860_{key}"

        # Find installation
        params = {
            "source_record_id": f"eq.{source_record_id}",
            "select": "id,owner_name",
        }
        result = supabase_request("GET", "solar_installations", params=params)

        if not result:
            not_found += 1
            continue

        inst = result[0]
        # Only update if owner_name is currently empty
        if not inst.get("owner_name"):
            res = supabase_request(
                "PATCH",
                "solar_installations",
                {"owner_name": info["owner_name"]},
                params={"id": f"eq.{inst['id']}"},
            )
            if res is not None:
                updated += 1

    print(f"  Matched to DB: {len(plant_owners) - not_found}")
    print(f"  Not found: {not_found}")
    print(f"  Owner names updated: {updated}")


def cross_reference_uspvdb():
    """Try to propagate EIA-860 owner/operator data to matching USPVDB records."""
    print("\n" + "=" * 60)
    print("3. Cross-Reference EIA-860 ↔ USPVDB")
    print("=" * 60)

    # Get EIA-860 records with owner or operator data
    params = {
        "source_record_id": "like.eia860_*",
        "or": "(owner_name.not.is.null,operator_name.not.is.null)",
        "select": "site_name,state,capacity_mw,owner_name,operator_name,latitude,longitude",
        "limit": "5000",
    }
    eia_records = supabase_request("GET", "solar_installations", params=params)

    if not eia_records:
        print("  No EIA-860 records with owner/operator found")
        return

    print(f"  Found {len(eia_records)} EIA-860 records with owner/operator data")

    # Get USPVDB records without owner
    params2 = {
        "source_record_id": "like.uspvdb_*",
        "owner_name": "is.null",
        "select": "id,site_name,state,capacity_mw,latitude,longitude",
        "limit": "5000",
    }
    uspvdb_records = supabase_request("GET", "solar_installations", params=params2)

    if not uspvdb_records:
        print("  No USPVDB records without owner found")
        return

    print(f"  Found {len(uspvdb_records)} USPVDB records without owner data")

    # Try to match by site_name + state
    eia_by_name = {}
    for rec in eia_records:
        if rec.get("site_name"):
            key = (rec["site_name"].lower().strip(), rec.get("state", ""))
            eia_by_name[key] = rec

    matched = 0
    for uspvdb in uspvdb_records:
        name = uspvdb.get("site_name", "")
        if not name:
            continue

        key = (name.lower().strip(), uspvdb.get("state", ""))
        if key in eia_by_name:
            eia = eia_by_name[key]
            update = {}
            if eia.get("owner_name"):
                update["owner_name"] = eia["owner_name"]
            if eia.get("operator_name"):
                update["operator_name"] = eia["operator_name"]

            if update:
                res = supabase_request(
                    "PATCH",
                    "solar_installations",
                    update,
                    params={"id": f"eq.{uspvdb['id']}"},
                )
                if res is not None:
                    matched += 1

    print(f"  USPVDB records enriched with EIA-860 owner/operator: {matched}")


def main():
    print("EIA-860 Data Enrichment Script")
    print("=" * 60)

    data_source_id = get_data_source_id()
    if not data_source_id:
        print("Error: eia860 data source not found")
        sys.exit(1)
    print(f"Data source ID: {data_source_id}")

    # 1. Retirement events (already completed - 471 events created)
    # extract_retirement_events(data_source_id)

    # 2. Owner data
    import_owner_data(data_source_id)

    # 3. Cross-reference
    cross_reference_uspvdb()

    print("\n" + "=" * 60)
    print("EIA-860 enrichment complete!")


if __name__ == "__main__":
    main()
