#!/usr/bin/env python3
"""
EIA-860 Plant & Generator Enrichment Script

1. Extract operator names and exact coordinates from Plant schedule (Schedule 2)
2. Extract uprate/derate/repower events from Generator schedule (Schedule 3)
3. Extract planned retirement and repower events from Generator "Operable" sheet
"""

import os
import sys
import json
import uuid
import urllib.request
import urllib.parse
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "data" / "eia860_2024"


def safe_float(val):
    """Safely convert a value to float, returning None for empty/invalid values."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None
PLANT_FILE = DATA_DIR / "2___Plant_Y2024.xlsx"
GENERATOR_FILE = DATA_DIR / "3_1_Generator_Y2024.xlsx"


def supabase_request(method, table, data=None, params=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items())

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates",
    }
    if headers_extra:
        headers.update(headers_extra)

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else []
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()[:200]
        print(f"  Supabase error ({e.code}): {error_body}")
        return None


def get_data_source_id():
    """Get the EIA-860 data source ID."""
    params = {"name": "eq.eia860", "select": "id"}
    result = supabase_request("GET", "solar_data_sources", params=params)
    if result:
        return result[0]["id"]
    return None


def enrich_from_plant_schedule():
    """Extract operator names, coordinates, and addresses from Plant schedule."""
    print("\n" + "=" * 60)
    print("1. Plant Schedule - Operator Names & Coordinates")
    print("=" * 60)

    wb = openpyxl.load_workbook(PLANT_FILE, read_only=True)
    ws = wb['Plant']

    # Build plant_code -> plant data mapping
    # Row 2 = headers, data starts row 3
    # Col 0=Utility ID, 1=Utility Name, 2=Plant Code, 3=Plant Name
    # Col 4=Street Address, 5=City, 6=State, 7=Zip, 8=County
    # Col 9=Latitude, 10=Longitude

    plant_data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        plant_code = row[2]
        if not plant_code:
            continue

        utility_name = str(row[1]).strip() if row[1] else None
        lat = row[9]
        lon = row[10]
        street = str(row[4]).strip() if row[4] else None
        city = str(row[5]).strip() if row[5] else None
        state = str(row[6]).strip() if row[6] else None
        zip_code = str(row[7]).strip() if row[7] else None
        county = str(row[8]).strip() if row[8] else None

        try:
            lat_f = float(lat) if lat and str(lat).strip() else None
        except (ValueError, TypeError):
            lat_f = None
        try:
            lon_f = float(lon) if lon and str(lon).strip() else None
        except (ValueError, TypeError):
            lon_f = None

        plant_data[str(plant_code)] = {
            "operator_name": utility_name,
            "latitude": lat_f,
            "longitude": lon_f,
            "street": street,
            "city": city,
            "state": state,
            "zip_code": zip_code,
            "county": county,
        }

    wb.close()
    print(f"  Loaded {len(plant_data)} plant records from Plant schedule")

    # Fetch all EIA-860 installations from DB in pages
    print("  Fetching EIA-860 installations from DB...")
    all_installations = []
    offset = 0
    while True:
        params = {
            "source_record_id": "like.eia860_*",
            "select": "id,source_record_id,operator_name,latitude,longitude",
            "limit": "1000",
            "offset": str(offset),
            "order": "id",
        }
        result = supabase_request("GET", "solar_installations", params=params)
        if not result:
            break
        all_installations.extend(result)
        if len(result) < 1000:
            break
        offset += 1000

    print(f"  Found {len(all_installations)} EIA-860 installations in DB")

    # Match and update
    updated_operator = 0
    updated_coords = 0
    not_found = 0

    for inst in all_installations:
        # Extract plant_code from source_record_id: "eia860_{plant_code}_{gen_id}"
        parts = inst["source_record_id"].split("_")
        if len(parts) < 3:
            continue
        plant_code = parts[1]

        if plant_code not in plant_data:
            not_found += 1
            continue

        plant = plant_data[plant_code]
        update = {}

        # Update operator name if missing
        if not inst.get("operator_name") and plant["operator_name"]:
            update["operator_name"] = plant["operator_name"]

        # Update coordinates if missing
        if not inst.get("latitude") and plant["latitude"]:
            update["latitude"] = plant["latitude"]
            update["longitude"] = plant["longitude"]

        if update:
            res = supabase_request(
                "PATCH",
                "solar_installations",
                update,
                params={"id": f"eq.{inst['id']}"},
            )
            if res is not None:
                if "operator_name" in update:
                    updated_operator += 1
                if "latitude" in update:
                    updated_coords += 1

    print(f"  Plant codes not found: {not_found}")
    print(f"  Operator names updated: {updated_operator}")
    print(f"  Coordinates updated: {updated_coords}")


def extract_generator_events(data_source_id):
    """Extract uprate/derate/repower events from Generator Operable sheet."""
    print("\n" + "=" * 60)
    print("2. Generator Events - Uprates, Derates, Repowers")
    print("=" * 60)

    wb = openpyxl.load_workbook(GENERATOR_FILE, read_only=True)
    ws = wb['Operable']

    # Col 2=Plant Code, 6=Generator ID, 7=Technology, 8=Prime Mover
    # Col 15=Nameplate MW
    # Col 20=Uprate or Derate Completed During Year
    # Col 21=Month Uprate/Derate Completed, Col 22=Year Uprate/Derate Completed
    # Col 27=Planned Retirement Month, Col 28=Planned Retirement Year
    # Col 33=Energy Source 1
    # Col 54=Planned Net Summer Capacity Uprate (MW)
    # Col 56=Planned Uprate Month, Col 57=Planned Uprate Year
    # Col 58=Planned Net Summer Capacity Derate (MW)
    # Col 60=Planned Derate Month, Col 61=Planned Derate Year
    # Col 65=Planned Repower Month, Col 66=Planned Repower Year
    # Col 67=Other Planned Modifications?
    # Col 68=Other Modifications Month, Col 69=Other Modifications Year

    events = []
    solar_rows = 0
    matched = 0

    for row in ws.iter_rows(min_row=3, values_only=True):
        energy_source = str(row[33]).strip().upper() if row[33] else ""
        # Only process solar generators (SUN = solar)
        if energy_source != "SUN":
            continue

        solar_rows += 1
        plant_code = row[2]
        gen_id = row[6]

        if not plant_code or not gen_id:
            continue

        source_record_id = f"eia860_{plant_code}_{gen_id}"
        nameplate_mw = row[15]

        # Check if installation exists in DB
        params = {
            "source_record_id": f"eq.{source_record_id}",
            "select": "id",
        }
        result = supabase_request("GET", "solar_installations", params=params)
        if not result:
            continue

        inst_id = result[0]["id"]
        matched += 1

        # 1. Check for completed uprate/derate
        uprate_derate = str(row[20]).strip() if row[20] else ""
        if uprate_derate and uprate_derate not in ("", "None", "N"):
            month_str = str(row[21]).strip() if row[21] else ""
            year_str = str(row[22]).strip() if row[22] else ""
            event_date = None
            if year_str and year_str.replace('.', '', 1).isdigit():
                yr = int(float(year_str))
                if month_str and month_str.replace('.', '', 1).isdigit():
                    mo = int(float(month_str))
                    event_date = f"{yr}-{mo:02d}-01"
                else:
                    event_date = f"{yr}-01-01"

            event_type = "uprate" if uprate_derate in ("1", "U") else "derate" if uprate_derate in ("2", "D") else "modification"
            events.append({
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "event_type": event_type,
                "event_date": event_date,
                "description": f"Generator {gen_id} - {event_type} completed per EIA-860 2024",
                "new_capacity_kw": None,
                "data_source_id": data_source_id,
            })

        # 2. Check for planned retirement
        ret_month_str = str(row[27]).strip() if row[27] else ""
        ret_year_str = str(row[28]).strip() if row[28] else ""
        if ret_year_str and ret_year_str.replace('.', '', 1).isdigit():
            yr = int(float(ret_year_str))
            event_date = f"{yr}-01-01"
            if ret_month_str and ret_month_str.replace('.', '', 1).isdigit():
                mo = int(float(ret_month_str))
                event_date = f"{yr}-{mo:02d}-01"
            events.append({
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "event_type": "planned_retirement",
                "event_date": event_date,
                "description": f"Generator {gen_id} - planned retirement per EIA-860 2024",
                "new_capacity_kw": None,
                "data_source_id": data_source_id,
            })

        # 3. Check for planned uprate
        planned_uprate_mw = row[54]
        if planned_uprate_mw and str(planned_uprate_mw).strip():
            up_month_str = str(row[56]).strip() if row[56] else ""
            up_year_str = str(row[57]).strip() if row[57] else ""
            event_date = None
            if up_year_str and up_year_str.replace('.', '', 1).isdigit():
                yr = int(float(up_year_str))
                event_date = f"{yr}-01-01"
                if up_month_str and up_month_str.replace('.', '', 1).isdigit():
                    mo = int(float(up_month_str))
                    event_date = f"{yr}-{mo:02d}-01"

            events.append({
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "event_type": "planned_uprate",
                "event_date": event_date,
                "description": f"Generator {gen_id} - planned uprate of {planned_uprate_mw} MW per EIA-860 2024",
                "new_capacity_kw": round(float(str(planned_uprate_mw).strip()) * 1000, 1) if str(planned_uprate_mw).strip().replace('.', '', 1).isdigit() else None,
                "data_source_id": data_source_id,
            })

        # 4. Check for planned derate
        planned_derate_mw = row[58]
        if planned_derate_mw and str(planned_derate_mw).strip():
            de_month_str = str(row[60]).strip() if row[60] else ""
            de_year_str = str(row[61]).strip() if row[61] else ""
            event_date = None
            if de_year_str and de_year_str.replace('.', '', 1).isdigit():
                yr = int(float(de_year_str))
                event_date = f"{yr}-01-01"
                if de_month_str and de_month_str.replace('.', '', 1).isdigit():
                    mo = int(float(de_month_str))
                    event_date = f"{yr}-{mo:02d}-01"

            events.append({
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "event_type": "planned_derate",
                "event_date": event_date,
                "description": f"Generator {gen_id} - planned derate of {planned_derate_mw} MW per EIA-860 2024",
                "new_capacity_kw": None,
                "data_source_id": data_source_id,
            })

        # 5. Check for planned repower
        repower_month_str = str(row[65]).strip() if row[65] else ""
        repower_year_str = str(row[66]).strip() if row[66] else ""
        if repower_year_str and repower_year_str.replace('.', '', 1).isdigit():
            yr = int(float(repower_year_str))
            event_date = f"{yr}-01-01"
            if repower_month_str and repower_month_str.replace('.', '', 1).isdigit():
                mo = int(float(repower_month_str))
                event_date = f"{yr}-{mo:02d}-01"
            events.append({
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "event_type": "planned_repower",
                "event_date": event_date,
                "description": f"Generator {gen_id} - planned repower per EIA-860 2024",
                "new_capacity_kw": None,
                "data_source_id": data_source_id,
            })

        # 6. Check for other planned modifications
        other_mods = str(row[67]).strip() if row[67] else ""
        if other_mods and other_mods not in ("", "None", "N"):
            mod_month_str = str(row[68]).strip() if row[68] else ""
            mod_year_str = str(row[69]).strip() if row[69] else ""
            event_date = None
            if mod_year_str and mod_year_str.replace('.', '', 1).isdigit():
                yr = int(float(mod_year_str))
                event_date = f"{yr}-01-01"
                if mod_month_str and mod_month_str.replace('.', '', 1).isdigit():
                    mo = int(float(mod_month_str))
                    event_date = f"{yr}-{mo:02d}-01"

            events.append({
                "id": str(uuid.uuid4()),
                "installation_id": inst_id,
                "event_type": "modification",
                "event_date": event_date,
                "description": f"Generator {gen_id} - other modifications planned per EIA-860 2024",
                "new_capacity_kw": None,
                "data_source_id": data_source_id,
            })

    wb.close()

    print(f"  Solar generators in Operable sheet: {solar_rows}")
    print(f"  Matched to DB: {matched}")
    print(f"  Events found: {len(events)}")

    # Insert events in batches
    created = 0
    for i in range(0, len(events), 50):
        batch = events[i:i + 50]
        res = supabase_request("POST", "solar_site_events", batch)
        if res is not None:
            created += len(batch)

    print(f"  Events created: {created}")


def main():
    print("EIA-860 Plant & Generator Enrichment")
    print("=" * 60)

    data_source_id = get_data_source_id()
    if not data_source_id:
        print("Error: eia860 data source not found")
        sys.exit(1)
    print(f"Data source ID: {data_source_id}")

    # 1. Plant schedule - operator names & coordinates
    enrich_from_plant_schedule()

    # 2. Generator events - uprates, derates, repowers
    extract_generator_events(data_source_id)

    print("\n" + "=" * 60)
    print("Plant & Generator enrichment complete!")


if __name__ == "__main__":
    main()
