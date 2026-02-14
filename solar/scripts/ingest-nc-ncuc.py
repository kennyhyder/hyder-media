#!/usr/bin/env python3
"""
North Carolina NCUC Renewable Energy Facility Registration Ingestion

Downloads and ingests NC Utilities Commission renewable energy facility
registrations into solar_installations table.

Source: https://www.ncuc.gov/Reps/RegistrationSpreadsheetPresent.xlsx

Usage:
  python3 -u scripts/ingest-nc-ncuc.py              # Full ingestion
  python3 -u scripts/ingest-nc-ncuc.py --dry-run     # Count without ingesting
"""

import os
import sys
import json
import re
import uuid
import argparse
import urllib.request
import urllib.parse
import time
from pathlib import Path
from openpyxl import load_workbook

from dotenv import load_dotenv

# Load env vars
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

BATCH_SIZE = 50
DATA_FILE = Path(__file__).parent.parent / "data" / "nc_ncuc" / "ncuc_registrations.xlsx"
DATA_SOURCE_NAME = "nc_ncuc"


def supabase_get(table, params, retries=3):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=headers)
            with urllib.request.urlopen(req, timeout=60) as resp:
                return json.loads(resp.read().decode())
        except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError) as e:
            if attempt < retries - 1:
                wait = 2 ** (attempt + 1)
                print(f"    Retry {attempt+1}/{retries} after {e} (waiting {wait}s)")
                time.sleep(wait)
            else:
                raise


def supabase_post(table, records):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates,return=minimal",
    }
    body = json.dumps(records).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            return True
    except urllib.error.HTTPError as e:
        err = e.read().decode()[:300]
        print(f"  POST error ({e.code}): {err}")
        return False


def get_data_source_id(name):
    rows = supabase_get("solar_data_sources", {
        "select": "id",
        "name": f"eq.{name}",
    })
    if rows:
        return rows[0]["id"]
    supabase_post("solar_data_sources", [{
        "id": str(uuid.uuid4()),
        "name": name,
        "url": "https://www.ncuc.gov/Reps/reps.html",
        "description": "North Carolina Utilities Commission renewable energy facility registrations",
    }])
    rows = supabase_get("solar_data_sources", {
        "select": "id",
        "name": f"eq.{name}",
    })
    return rows[0]["id"] if rows else None


def get_existing_source_ids(prefix):
    existing = set()
    offset = 0
    while True:
        rows = supabase_get("solar_installations", {
            "select": "source_record_id",
            "source_record_id": f"like.{prefix}*",
            "limit": "1000",
            "offset": str(offset),
            "order": "source_record_id",
        })
        if not rows:
            break
        for r in rows:
            existing.add(r["source_record_id"])
        if len(rows) < 1000:
            break
        offset += 1000
    return existing


def safe_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f != f:
            return None
        return f
    except (ValueError, TypeError):
        return None


def main():
    parser = argparse.ArgumentParser(description="Ingest NC NCUC facility registrations")
    parser.add_argument("--dry-run", action="store_true", help="Count without ingesting")
    args = parser.parse_args()

    print("NC NCUC Renewable Energy Facility Registration Ingestion")
    print("=" * 60)

    if not DATA_FILE.exists():
        print(f"Error: Data file not found: {DATA_FILE}")
        print("Download from: https://www.ncuc.gov/Reps/RegistrationSpreadsheetPresent.xlsx")
        sys.exit(1)

    print(f"  Loading: {DATA_FILE}")
    wb = load_workbook(DATA_FILE, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    records = []
    skipped_fuel = 0
    skipped_small = 0

    # Process sheets — "New REF - All" has active registrations, "Rev|Can" has revoked/canceled
    for sheet_name in wb.sheetnames:
        # Only process "All" sheets (not yearly subsets)
        if "2024" in sheet_name and "All" not in sheet_name:
            continue
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            continue

        is_revoked = "rev" in sheet_name.lower() or "can" in sheet_name.lower()

        # Header is always at row 6 (0-indexed) — rows 0-5 are title + blank + total count
        header_idx = None
        for i, row in enumerate(rows[:10]):
            if row and str(row[0]).strip().startswith("Docket"):
                header_idx = i
                break
        if header_idx is None:
            print(f"\n  Sheet: {sheet_name} — no header row found, skipping")
            continue

        headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(rows[header_idx])]
        print(f"\n  Sheet: {sheet_name} ({len(rows) - header_idx - 1} rows)")
        print(f"  Headers: {headers}")
        if is_revoked:
            print(f"  (Revoked/Canceled — will mark as canceled)")

        # Fixed column mapping based on known schema
        col_docket = 0   # "Docket #"
        col_sub = 1      # "Sub"
        col_company = 2  # "Company"
        col_facility = 3 # "Facility"
        col_state = 4    # "State"
        col_fuel = 5     # "Primary Fuel Type"
        col_capacity = 6 # "Capacity (kW)"

        for row in rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            fuel = str(row[col_fuel]).strip() if col_fuel is not None and row[col_fuel] else ""
            if "solar" not in fuel.lower() and "photovoltaic" not in fuel.lower():
                skipped_fuel += 1
                continue

            capacity_kw = safe_float(row[col_capacity] if col_capacity is not None else None)
            if capacity_kw is not None and capacity_kw < 25:
                skipped_small += 1
                continue

            docket = str(row[col_docket]).strip() if col_docket is not None and row[col_docket] else ""
            sub = str(row[col_sub]).strip() if col_sub is not None and row[col_sub] else ""
            company = str(row[col_company]).strip() if col_company is not None and row[col_company] else ""
            facility = str(row[col_facility]).strip() if col_facility is not None and row[col_facility] else ""
            state = str(row[col_state]).strip() if col_state is not None and row[col_state] else "NC"

            # Build unique source ID from docket + sub
            docket_key = re.sub(r'[^a-z0-9]', '_', docket.lower())
            sub_key = re.sub(r'[^a-z0-9]', '', str(sub).lower()) if sub else "0"
            source_id = f"ncncuc_{docket_key}_{sub_key}"

            # Determine site type
            site_type = "commercial"
            capacity_mw = round(capacity_kw / 1000, 3) if capacity_kw else None
            if capacity_mw and capacity_mw >= 1:
                site_type = "utility"

            # Company is likely owner/developer
            owner_name = company if company and company.lower() not in ("none", "nan", "") else None

            site_status = "canceled" if is_revoked else "active"

            record = {
                "source_record_id": source_id,
                "site_name": facility if facility and facility.lower() not in ("none", "nan", "") else None,
                "site_type": site_type,
                "address": None,
                "city": None,
                "state": state if state else "NC",
                "zip_code": None,
                "county": None,
                "latitude": None,
                "longitude": None,
                "capacity_dc_kw": capacity_kw,
                "capacity_mw": capacity_mw,
                "install_date": None,
                "site_status": site_status,
                "installer_name": None,
                "owner_name": owner_name,
                "developer_name": None,
                "operator_name": None,
                "total_cost": None,
                "data_source_id": None,
                "has_battery_storage": False,
            }
            records.append(record)

    print(f"\n  Total solar records >= 25 kW: {len(records)}")
    print(f"  Skipped (non-solar): {skipped_fuel}")
    print(f"  Skipped (< 25 kW): {skipped_small}")

    # Owner breakdown
    owners = {}
    for r in records:
        o = r.get("owner_name") or "Unknown"
        owners[o] = owners.get(o, 0) + 1
    print(f"\n  Top owners/companies:")
    for o, c in sorted(owners.items(), key=lambda x: -x[1])[:10]:
        print(f"    {o}: {c}")

    # Capacity distribution
    utility = sum(1 for r in records if r["site_type"] == "utility")
    commercial = sum(1 for r in records if r["site_type"] == "commercial")
    print(f"\n  Utility-scale (>= 1 MW): {utility}")
    print(f"  Commercial (25 kW - 1 MW): {commercial}")

    if args.dry_run:
        print(f"\n  [DRY RUN] No records created.")
        return

    # Get data source ID
    data_source_id = get_data_source_id(DATA_SOURCE_NAME)
    for r in records:
        r["data_source_id"] = data_source_id

    # Check existing
    existing = get_existing_source_ids("ncncuc_")
    print(f"\n  Existing records: {len(existing)}")
    new_records = [r for r in records if r["source_record_id"] not in existing]
    print(f"  New records to create: {len(new_records)}")

    if not new_records:
        print("  All records already exist!")
        return

    # Insert
    created = 0
    errors = 0
    for i in range(0, len(new_records), BATCH_SIZE):
        batch = new_records[i:i + BATCH_SIZE]
        ok = supabase_post("solar_installations", batch)
        if ok:
            created += len(batch)
        else:
            errors += len(batch)
        if (i // BATCH_SIZE) % 20 == 0:
            print(f"  Progress: {created + errors}/{len(new_records)}")

    print(f"\n  Created: {created}")
    print(f"  Errors: {errors}")
    print("Done!")


if __name__ == "__main__":
    main()
