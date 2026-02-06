#!/usr/bin/env python3
"""
LBNL "Queued Up" Interconnection Queue Enrichment Script

Reads the LBNL Queued Up dataset (interconnection queue data from 50+ grid
operators) and enriches existing solar installations with developer_name by
cross-referencing via EIA Plant ID, location+capacity, or state+capacity matching.

Also ingests new solar projects not already in the database.

Data source: https://emp.lbl.gov/queues
File: Download the Excel data file from the publications page.
Save to: solar/data/lbnl_queues/

Usage:
  python3 -u scripts/enrich-lbnl-queues.py                    # Full run
  python3 -u scripts/enrich-lbnl-queues.py --dry-run           # Report only
  python3 -u scripts/enrich-lbnl-queues.py --ingest-new        # Also ingest unmatched projects
"""

import os
import sys
import json
import re
import uuid
import argparse
import math
import urllib.request
import urllib.parse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
from dotenv import load_dotenv

# Load env vars
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "data" / "lbnl_queued_up"
WORKERS = 20
BATCH_SIZE = 50


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def supabase_get(table, params):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
    }
    req = urllib.request.Request(url, headers=headers)
    with urllib.request.urlopen(req) as resp:
        return json.loads(resp.read().decode())


def supabase_patch(table, data, params):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    body = json.dumps(data).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="PATCH")
    try:
        with urllib.request.urlopen(req) as resp:
            return True
    except urllib.error.HTTPError as e:
        print(f"  PATCH error ({e.code}): {e.read().decode()[:200]}")
        return False


def supabase_post(table, data):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=ignore-duplicates",
    }
    body = json.dumps(data).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req) as resp:
            return True
    except urllib.error.HTTPError as e:
        print(f"  POST error ({e.code}): {e.read().decode()[:200]}")
        return False


def _do_patch(args):
    inst_id, patch = args
    return supabase_patch(
        "solar_installations",
        patch,
        {"id": f"eq.{inst_id}"},
    )


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("n/a", "na", "none", "nan", "", "unknown", "tbd", "masked"):
        return None
    return s


def safe_float(val):
    if val is None:
        return None
    try:
        v = float(str(val).replace(",", "").replace("$", ""))
        if math.isnan(v) or math.isinf(v):
            return None
        return v
    except (ValueError, TypeError):
        return None


def normalize_name(name):
    """Normalize a name for comparison."""
    if not name:
        return ""
    return re.sub(r'[^a-z0-9]', '', name.lower())


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(dlam / 2) ** 2
    return R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


# ---------------------------------------------------------------------------
# Load LBNL Queued Up data
# ---------------------------------------------------------------------------

def find_data_file():
    """Find the LBNL Queued Up Excel file in the data directory."""
    if not DATA_DIR.exists():
        return None
    for f in DATA_DIR.iterdir():
        if f.suffix in (".xlsx", ".xls") and f.stat().st_size > 10000:
            return f
    return None


def load_queued_up(filepath):
    """Load solar projects from LBNL Queued Up Excel."""
    print(f"Loading LBNL Queued Up data from {filepath.name}...")
    wb = openpyxl.load_workbook(str(filepath), read_only=True)

    print(f"  Sheets: {wb.sheetnames}")

    # Find the project-level data sheet
    # Prioritize "Complete Queue Data", then other queue/project/data sheets
    target_sheet = None
    for name in wb.sheetnames:
        lower = name.lower()
        if "complete queue" in lower or "complete data" in lower:
            target_sheet = name
            break
    if not target_sheet:
        for name in wb.sheetnames:
            lower = name.lower()
            if "queue data" in lower or "project data" in lower or "all request" in lower:
                target_sheet = name
                break

    if not target_sheet:
        # Use the sheet with the most rows (likely the data sheet)
        max_rows = 0
        for name in wb.sheetnames:
            ws = wb[name]
            count = 0
            for _ in ws.iter_rows(max_row=100, values_only=True):
                count += 1
            if count > max_rows:
                max_rows = count
                target_sheet = name
        print(f"  Auto-selected sheet: '{target_sheet}' ({max_rows} sample rows)")
    else:
        print(f"  Using sheet: '{target_sheet}'")

    ws = wb[target_sheet]

    # Read headers - may be on row 1 or row 2 (LBNL v2 has row 1 as nav text)
    headers = None
    projects = []
    row_num = 0

    for row in ws.iter_rows(values_only=True):
        row_num += 1

        # Detect header row: look for known column names
        if not headers:
            cells = [str(h).strip() if h else "" for h in row]
            cells_lower = [c.lower() for c in cells]
            if any(kw in cells_lower for kw in ["developer", "state", "type_clean",
                                                  "capacity_mw", "fuel", "mw1"]):
                headers = [c if c else f"col_{i}" for i, c in enumerate(cells)]
                print(f"  Header row: {row_num}")
                print(f"  {len(headers)} columns")
                print(f"  Headers: {headers[:20]}...")
            continue

        record = dict(zip(headers, row))

        # Try to find solar projects - look for fuel/type column
        fuel = None
        for key in ["type", "fuel", "generation_type", "Generation Type", "Fuel",
                     "Type", "resource", "Resource", "technology", "Technology",
                     "type_clean", "fuel_type"]:
            fuel = safe_str(record.get(key))
            if fuel:
                break

        if not fuel:
            # Check all columns for solar keyword
            for k, v in record.items():
                if v and "solar" in str(v).lower():
                    fuel = str(v)
                    break

        if not fuel or not any(kw in fuel.lower() for kw in ["solar", "photovoltaic", "pv"]):
            continue

        # Extract developer name - the key field we want
        developer = None
        for key in ["developer", "Developer", "developer_name", "Developer Name",
                     "developer_clean", "entity_name", "Entity Name",
                     "interconnecting_entity", "Interconnecting Entity",
                     "applicant", "Applicant"]:
            developer = safe_str(record.get(key))
            if developer:
                break

        # Extract capacity
        capacity_mw = None
        for key in ["capacity_mw", "Capacity (MW)", "MW", "mw", "mw1",
                     "nameplate_capacity_mw", "Nameplate Capacity (MW)",
                     "capacity", "Capacity", "mw_1", "net_mw"]:
            capacity_mw = safe_float(record.get(key))
            if capacity_mw and capacity_mw > 0:
                break

        if not capacity_mw or capacity_mw <= 0:
            continue

        # Extract state
        state = None
        for key in ["state", "State", "state_clean"]:
            state = safe_str(record.get(key))
            if state:
                state = state[:2].upper()
                break

        # Extract county
        county = None
        for key in ["county", "County", "county_clean"]:
            county = safe_str(record.get(key))
            if county:
                break

        # Extract project name
        name = None
        for key in ["project_name", "Project Name", "name", "Name", "project"]:
            name = safe_str(record.get(key))
            if name:
                break

        # Extract status
        status = None
        for key in ["q_status", "status", "Status", "queue_status", "Queue Status"]:
            status = safe_str(record.get(key))
            if status:
                break

        # Extract queue/ISO
        iso = None
        for key in ["region", "Region", "iso", "ISO", "rto", "RTO",
                     "queue_name", "Queue Name", "entity"]:
            iso = safe_str(record.get(key))
            if iso:
                break

        # Extract dates
        queue_date = None
        for key in ["queue_date", "Queue Date", "date", "Date",
                     "queue_year", "year"]:
            val = record.get(key)
            if val:
                try:
                    from datetime import datetime
                    if isinstance(val, datetime):
                        queue_date = val.strftime("%Y-%m-%d")
                    elif isinstance(val, (int, float)):
                        queue_date = f"{int(val)}-01-01"
                    else:
                        queue_date = str(val).strip()[:10]
                except Exception:
                    pass
            if queue_date:
                break

        # Extract EIA plant ID if available
        eia_id = None
        for key in ["eia_id", "EIA ID", "plant_code", "Plant Code", "ORISPL",
                     "eia_plant_code", "EIA Plant Code"]:
            eia_id = safe_str(record.get(key))
            if eia_id:
                try:
                    eia_id = int(float(eia_id))
                except (ValueError, TypeError):
                    eia_id = None
            if eia_id:
                break

        projects.append({
            "developer": developer,
            "name": name,
            "capacity_mw": capacity_mw,
            "state": state,
            "county": county,
            "status": status,
            "iso": iso,
            "queue_date": queue_date,
            "eia_id": eia_id,
        })

    wb.close()
    print(f"  Loaded {len(projects)} solar projects")
    with_developer = sum(1 for p in projects if p["developer"])
    print(f"  With developer name: {with_developer}")
    return projects


# ---------------------------------------------------------------------------
# Load installations
# ---------------------------------------------------------------------------

def load_installations():
    """Load all installations from database."""
    print("Loading installations from database...")
    all_records = []
    offset = 0
    limit = 1000

    while True:
        params = {
            "select": "id,source_record_id,developer_name,owner_name,operator_name,"
                      "state,county,capacity_mw,latitude,longitude,site_name",
            "limit": str(limit),
            "offset": str(offset),
            "order": "id",
        }
        batch = supabase_get("solar_installations", params)
        if not batch:
            break
        all_records.extend(batch)
        if len(all_records) % 10000 == 0:
            print(f"  Fetched {len(all_records)} records...")
        if len(batch) < limit:
            break
        offset += limit

    print(f"  Total: {len(all_records)} installations loaded")
    return all_records


# ---------------------------------------------------------------------------
# Matching
# ---------------------------------------------------------------------------

def extract_eia_plant_code(source_record_id):
    if not source_record_id:
        return None
    for prefix in ["eia860_", "eia860m_", "lbnl_"]:
        if source_record_id.startswith(prefix):
            code = source_record_id[len(prefix):].split("_")[0]
            try:
                return int(code)
            except ValueError:
                return None
    return None


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Enrich with LBNL Queued Up developer names")
    parser.add_argument("--dry-run", action="store_true", help="Report without patching")
    parser.add_argument("--ingest-new", action="store_true",
                        help="Also ingest unmatched projects as new installations")
    args = parser.parse_args()

    filepath = find_data_file()
    if not filepath:
        print(f"Error: No Excel file found in {DATA_DIR}/")
        print(f"Download from: https://emp.lbl.gov/queues")
        print(f"Save the Excel data file to: {DATA_DIR}/")
        sys.exit(1)

    # Load data
    projects = load_queued_up(filepath)
    if not projects:
        print("No solar projects found in the data file.")
        print("Check the sheet names and column mappings.")
        sys.exit(1)

    installations = load_installations()

    # Build indexes
    # 1. EIA plant code -> installations
    inst_by_eia = {}
    for inst in installations:
        code = extract_eia_plant_code(inst.get("source_record_id"))
        if code:
            inst_by_eia.setdefault(code, []).append(inst)

    # 2. State+capacity index for approximate matching
    inst_by_state = {}
    for inst in installations:
        state = inst.get("state")
        if state:
            inst_by_state.setdefault(state, []).append(inst)

    print(f"\nInstallations by EIA code: {len(inst_by_eia)} unique codes")
    print(f"Installations by state: {len(inst_by_state)} states")

    # Phase 1: Match by EIA Plant ID
    print(f"\n{'='*60}")
    print("Phase 1: EIA Plant ID matching")
    print(f"{'='*60}")

    patches = []
    matched_project_idxs = set()
    phase1_matches = 0

    for idx, proj in enumerate(projects):
        if not proj["developer"]:
            continue
        if not proj["eia_id"]:
            continue

        insts = inst_by_eia.get(proj["eia_id"], [])
        if not insts:
            continue

        for inst in insts:
            if inst.get("developer_name"):
                continue  # Already has developer
            patch = {"developer_name": proj["developer"][:255]}
            patches.append((inst["id"], patch))
            matched_project_idxs.add(idx)
            phase1_matches += 1

    print(f"  Phase 1 matches: {phase1_matches}")

    # Phase 2: State + capacity matching (within 25%)
    print(f"\n{'='*60}")
    print("Phase 2: State + capacity matching")
    print(f"{'='*60}")

    phase2_matches = 0
    matched_inst_ids = {inst_id for inst_id, _ in patches}

    for idx, proj in enumerate(projects):
        if idx in matched_project_idxs:
            continue
        if not proj["developer"]:
            continue
        if not proj["state"] or not proj["capacity_mw"]:
            continue

        candidates = inst_by_state.get(proj["state"], [])
        best = None
        best_score = 0

        for inst in candidates:
            if inst["id"] in matched_inst_ids:
                continue
            if inst.get("developer_name"):
                continue

            cap = inst.get("capacity_mw")
            if not cap or cap <= 0:
                continue

            # Capacity within 25%
            ratio = proj["capacity_mw"] / cap if cap > 0 else 999
            if ratio < 0.75 or ratio > 1.25:
                continue

            # Name similarity bonus
            score = 1.0
            if proj["name"] and inst.get("site_name"):
                proj_norm = normalize_name(proj["name"])
                inst_norm = normalize_name(inst["site_name"])
                if proj_norm and inst_norm:
                    # Check for word overlap
                    proj_words = set(re.findall(r'[a-z]+', proj_norm))
                    inst_words = set(re.findall(r'[a-z]+', inst_norm))
                    overlap = len(proj_words & inst_words)
                    total = len(proj_words | inst_words)
                    if total > 0:
                        score = overlap / total
                        if score > 0.5:
                            score += 10  # Strong name match bonus

            # County match bonus
            if proj["county"] and inst.get("county"):
                if normalize_name(proj["county"]) == normalize_name(inst["county"]):
                    score += 5

            if score > best_score:
                best = inst
                best_score = score

        # Only accept high-confidence matches (name or county match)
        if best and best_score >= 5:
            patch = {"developer_name": proj["developer"][:255]}
            patches.append((best["id"], patch))
            matched_inst_ids.add(best["id"])
            matched_project_idxs.add(idx)
            phase2_matches += 1

    print(f"  Phase 2 matches: {phase2_matches}")

    # Summary
    print(f"\n{'='*60}")
    print("LBNL Queued Up Enrichment Summary")
    print(f"{'='*60}")
    print(f"  Total projects loaded: {len(projects)}")
    print(f"  Projects with developer: {sum(1 for p in projects if p['developer'])}")
    print(f"  Total patches: {len(patches)}")
    print(f"  Phase 1 (EIA ID): {phase1_matches}")
    print(f"  Phase 2 (state+cap): {phase2_matches}")
    print(f"  Unmatched projects: {len(projects) - len(matched_project_idxs)}")

    if args.dry_run:
        print("\n  [DRY RUN] No patches applied.")
        for inst_id, patch in patches[:10]:
            print(f"    {inst_id}: {patch}")
        return

    # Apply patches
    if patches:
        print(f"\nApplying {len(patches)} patches ({WORKERS} workers)...")
        applied = 0
        errors = 0

        with ThreadPoolExecutor(max_workers=WORKERS) as executor:
            futures = {executor.submit(_do_patch, item): item for item in patches}
            for future in as_completed(futures):
                if future.result():
                    applied += 1
                else:
                    errors += 1
                if (applied + errors) % 500 == 0:
                    print(f"  Progress: {applied} applied, {errors} errors")

        print(f"\n  Applied: {applied}")
        print(f"  Errors: {errors}")

    # Optionally ingest unmatched projects as new installations
    if args.ingest_new:
        unmatched = [projects[i] for i in range(len(projects))
                     if i not in matched_project_idxs
                     and projects[i]["capacity_mw"] >= 1  # Utility-scale only
                     and projects[i]["state"]]

        if unmatched:
            print(f"\nIngesting {len(unmatched)} new projects...")
            # Get data source ID
            ds = supabase_get("solar_data_sources", {"name": "eq.lbnl_queues", "select": "id"})
            if not ds:
                ds_id = str(uuid.uuid4())
                supabase_post("solar_data_sources", {
                    "id": ds_id,
                    "name": "lbnl_queues",
                    "description": "LBNL Queued Up - Interconnection queue data from 50+ grid operators",
                    "url": "https://emp.lbl.gov/queues",
                    "record_count": 0,
                })
            else:
                ds_id = ds[0]["id"]

            batch = []
            created = 0
            for proj in unmatched:
                inst_id = str(uuid.uuid4())
                name_slug = re.sub(r'[^a-z0-9]+', '_', (proj["name"] or "unknown").lower())[:40]
                record = {
                    "id": inst_id,
                    "source_record_id": f"lbnlq_{name_slug}_{proj['state'] or 'XX'}_{int(proj['capacity_mw'])}",
                    "data_source_id": ds_id,
                    "site_name": (proj["name"] or "")[:255] or None,
                    "state": proj["state"],
                    "county": proj["county"],
                    "capacity_mw": round(proj["capacity_mw"], 3),
                    "capacity_dc_kw": round(proj["capacity_mw"] * 1000, 3),
                    "site_type": "utility" if proj["capacity_mw"] >= 1 else "commercial",
                    "site_status": "proposed",
                    "developer_name": (proj["developer"] or "")[:255] or None,
                }
                batch.append(record)

                if len(batch) >= BATCH_SIZE:
                    if supabase_post("solar_installations", batch):
                        created += len(batch)
                    batch = []

            if batch:
                if supabase_post("solar_installations", batch):
                    created += len(batch)

            print(f"  Created: {created} new installations")

    print("\nDone!")


if __name__ == "__main__":
    main()
