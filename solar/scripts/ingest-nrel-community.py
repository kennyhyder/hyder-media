#!/usr/bin/env python3
"""
NREL Sharing the Sun Community Solar Project Database Ingestion Script

Downloads and imports community solar projects from NREL's "Sharing the Sun"
dataset. Covers ~4,000 projects across the U.S. with:
- Project Name, City, State
- Developer/Subscription Management/Contractor Name
- Utility (operator)
- System Size (kW-AC, kW-DC, MW-AC, MW-DC)
- Year of Interconnection

Only ingests projects >= 25 kW-AC (commercial threshold).
Skips aggregated data entries (roll-ups, not individual projects).

Source: https://data.nrel.gov/submissions/244
Download: https://data.nrel.gov/system/files/244/1763653621-Sharing%20the%20Sun%20Community%20Solar%20Project%20Data%20%28June%202025%29_Nov.xlsx

Usage:
  python3 -u scripts/ingest-nrel-community.py              # Full ingestion
  python3 -u scripts/ingest-nrel-community.py --dry-run     # Report without ingesting
"""

import os
import sys
import json
import re
import argparse
import urllib.request
import urllib.parse
import urllib.error
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install with: pip3 install openpyxl")
    sys.exit(1)

from dotenv import load_dotenv

# Load env
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

DOWNLOAD_URL = "https://data.nrel.gov/system/files/244/1763653621-Sharing%20the%20Sun%20Community%20Solar%20Project%20Data%20%28June%202025%29_Nov.xlsx"
DATA_DIR = Path(__file__).parent.parent / "data" / "nrel_community_solar"
XLSX_FILENAME = "sharing_the_sun_june2025.xlsx"

BATCH_SIZE = 50
MIN_CAPACITY_KW_AC = 25  # 25 kW-AC commercial threshold

# Excel structure: "Project List" sheet, header at row 1, data from row 2
SHEET_NAME = "Project List"

# Column indices (0-based) confirmed from inspecting the Excel file
COL_UTILITY_ID = 0          # Utility ID
COL_PROJECT_NAME = 1        # Project Name
COL_CITY = 2                # City
COL_STATE = 3               # State
COL_UTILITY = 4             # Utility
COL_UTILITY_TYPE = 5        # Utility Type
COL_SUB_MARKETER = 6        # Subscription Marketer
COL_PROGRAM_NAME = 7        # Program Name
COL_DEVELOPER = 8           # Developer, Subscription Management, or Contractor Name
COL_SIZE_MW_AC = 9          # System Size (MW-AC)
COL_SIZE_KW_AC = 10         # System Size (kW-AC)
COL_SIZE_MW_DC = 11         # System Size (MW-DC)
COL_SIZE_KW_DC = 12         # System Size (kW-DC)
COL_YEAR = 13               # Year of Interconnection
COL_LMI_REQ = 14            # Does this Project have LMI Portion Requirement?
COL_LI_LMI_SUBS = 15        # LI/LMI Subscribers (Count)
COL_TOTAL_SUBS = 16         # Total Subscribers (Count)
COL_LI_LMI_KWH = 17         # LI/LMI Subscriber (kWh)
COL_TOTAL_KWH = 18          # Total Subscribed (kWh)
COL_LI_LMI_PORTION = 19     # LI/LMI Portion
COL_CALC_REG = 20           # Is the LI/LMI Portion Calculated or Regulated?
COL_LI_LMI_MW = 21          # LI/LMI System Size (MW-AC)
COL_AGGREGATED = 22         # Aggregated Data Entry


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def supabase_request(method, table, data=None, params=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else []
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()[:300]
        print(f"  Supabase error ({e.code}): {error_body}")
        return None


def supabase_post_batch(table, records):
    """POST a batch of records. On duplicate error, insert one by one."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    body = json.dumps(records).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        urllib.request.urlopen(req)
        return len(records), 0
    except urllib.error.HTTPError as e:
        err = e.read().decode()[:300]
        if "duplicate" in err.lower() or "unique" in err.lower():
            # Insert one by one to skip duplicates
            created = 0
            errors = 0
            for rec in records:
                try:
                    body = json.dumps([rec]).encode()
                    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
                    urllib.request.urlopen(req)
                    created += 1
                except urllib.error.HTTPError:
                    errors += 1
            return created, errors
        print(f"  POST error ({e.code}): {err}")
        return 0, len(records)


def get_existing_source_ids(prefix):
    """Query DB for existing source_record_ids with given prefix."""
    existing = set()
    offset = 0
    while True:
        params = {
            "source_record_id": f"like.{prefix}*",
            "select": "source_record_id",
            "limit": "1000",
            "offset": str(offset),
        }
        rows = supabase_request("GET", "solar_installations", params=params)
        if not rows:
            break
        for r in rows:
            existing.add(r["source_record_id"])
        if len(rows) < 1000:
            break
        offset += 1000
    return existing


def get_or_create_data_source():
    """Get or create the NREL Community Solar data source record."""
    params = {"name": "eq.nrel_community_solar", "select": "id"}
    existing = supabase_request("GET", "solar_data_sources", params=params)
    if existing:
        return existing[0]["id"]

    result = supabase_request("POST", "solar_data_sources", {
        "name": "nrel_community_solar",
        "description": "NREL Sharing the Sun Community Solar Project Database (June 2025) - Community solar projects across the U.S.",
        "url": "https://data.nrel.gov/submissions/244",
        "record_count": 0,
    })
    # Re-fetch to get auto-generated ID
    existing = supabase_request("GET", "solar_data_sources", params={"name": "eq.nrel_community_solar", "select": "id"})
    if existing:
        return existing[0]["id"]
    return None


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def safe_str(val):
    """Convert value to cleaned string, returning None for empty/placeholder values."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("n/a", "nan", "none", "na", "null", "-", ".", "unknown", "0"):
        return None
    return s


def safe_float(val):
    """Convert value to float, handling None, empty, and placeholder values."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("-", "n/a", "unknown", "none", "."):
        return None
    try:
        v = float(s.replace(",", ""))
        if v <= 0:
            return None
        return v
    except (ValueError, TypeError):
        return None


def parse_year(val):
    """Parse year of interconnection to install_date (YYYY-01-01 format)."""
    if not val:
        return None

    # Handle datetime objects from openpyxl
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    s = str(val).strip()
    if not s or s.lower() in ("-", "n/a", "unknown", "none"):
        return None

    # Extract first 4-digit year
    m = re.match(r'(\d{4})', s)
    if m:
        year = int(m.group(1))
        if 2000 <= year <= 2030:
            return f"{year}-01-01"
    return None


def sanitize_id(name):
    """Sanitize a string for use in source_record_id."""
    if not name:
        return ""
    clean = re.sub(r'[^a-z0-9]', '_', name.lower())
    clean = re.sub(r'_+', '_', clean).strip('_')
    return clean[:40]


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_data():
    """Download the NREL Sharing the Sun Excel file if not already present."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = DATA_DIR / XLSX_FILENAME

    if xlsx_path.exists():
        size_kb = xlsx_path.stat().st_size / 1024
        print(f"  Found existing file ({size_kb:.0f} KB), skipping download")
        return xlsx_path

    print(f"  Downloading from NREL...")
    print(f"  URL: {DOWNLOAD_URL}")
    try:
        req = urllib.request.Request(DOWNLOAD_URL, headers={
            "User-Agent": "Mozilla/5.0 (SolarTrack Data Ingestion)"
        })
        with urllib.request.urlopen(req, timeout=120) as resp:
            with open(xlsx_path, "wb") as f:
                f.write(resp.read())
        size_kb = xlsx_path.stat().st_size / 1024
        print(f"  Downloaded {size_kb:.0f} KB")
    except urllib.error.HTTPError as e:
        print(f"  Download failed ({e.code}): {e.reason}")
        print(f"\n  Please manually download from:")
        print(f"  https://data.nrel.gov/submissions/244")
        print(f"  Save to: {xlsx_path}")
        sys.exit(1)

    return xlsx_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Ingest NREL Community Solar projects")
    parser.add_argument("--dry-run", action="store_true", help="Report without ingesting")
    args = parser.parse_args()

    print("NREL Sharing the Sun Community Solar - Ingestion Script")
    print("=" * 60)
    print(f"Source: https://data.nrel.gov/submissions/244")
    print(f"Filter: Community solar projects >= {MIN_CAPACITY_KW_AC} kW-AC")
    if args.dry_run:
        print("MODE: DRY RUN (no database changes)")
    print()

    # Download data
    print("Step 1: Downloading NREL Community Solar data...")
    xlsx_path = download_data()

    # Read Excel
    print(f"\nStep 2: Reading {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        print(f"  ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        wb.close()
        sys.exit(1)

    ws = wb[SHEET_NAME]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    headers = all_rows[0]
    data_rows = all_rows[1:]
    print(f"  Headers: {[str(h)[:40] for h in headers if h][:10]}")
    print(f"  Total data rows: {len(data_rows)}")

    # Get or create data source
    print("\nStep 3: Setting up data source...")
    if args.dry_run:
        data_source_id = "DRY_RUN"
        print(f"  [DRY RUN] Skipping data source creation")
    else:
        data_source_id = get_or_create_data_source()
        print(f"  Data source ID: {data_source_id}")

    # Pre-check existing records
    print("\nStep 4: Checking for existing records...")
    if args.dry_run:
        existing_ids = set()
    else:
        existing_ids = get_existing_source_ids("nrelcs_")
    print(f"  Existing in DB: {len(existing_ids)}")

    # Build installation records
    print("\nStep 5: Processing project rows...")
    installations = []
    skipped_aggregated = 0
    skipped_no_state = 0
    skipped_small = 0
    skipped_no_cap = 0
    skipped_existing = 0
    skipped_empty = 0

    # Track seen source_record_ids within this run to handle name+state duplicates
    seen_ids = {}

    for row_idx, row in enumerate(data_rows):
        # Skip empty rows
        if not any(c for c in row if c):
            skipped_empty += 1
            continue

        # Skip aggregated entries (roll-ups, not individual projects)
        aggregated = safe_str(row[COL_AGGREGATED]) if len(row) > COL_AGGREGATED else None
        if aggregated and aggregated.lower() == "yes":
            skipped_aggregated += 1
            continue

        # State is required
        state = safe_str(row[COL_STATE]) if len(row) > COL_STATE else None
        if not state:
            skipped_no_state += 1
            continue
        state = state.strip().upper()[:2]

        # Capacity filter: prefer kW-AC, fallback to MW-AC converted
        cap_kw_ac = safe_float(row[COL_SIZE_KW_AC]) if len(row) > COL_SIZE_KW_AC else None
        cap_mw_ac = safe_float(row[COL_SIZE_MW_AC]) if len(row) > COL_SIZE_MW_AC else None
        cap_kw_dc = safe_float(row[COL_SIZE_KW_DC]) if len(row) > COL_SIZE_KW_DC else None
        cap_mw_dc = safe_float(row[COL_SIZE_MW_DC]) if len(row) > COL_SIZE_MW_DC else None

        # Use kW-AC for filtering
        if cap_kw_ac is None and cap_mw_ac is not None:
            cap_kw_ac = round(cap_mw_ac * 1000, 3)

        if cap_kw_ac is None:
            skipped_no_cap += 1
            continue
        if cap_kw_ac < MIN_CAPACITY_KW_AC:
            skipped_small += 1
            continue

        # Compute capacity_mw for DB
        if cap_mw_ac:
            capacity_mw = round(cap_mw_ac, 6)
        else:
            capacity_mw = round(cap_kw_ac / 1000, 6)

        # Parse fields
        project_name = safe_str(row[COL_PROJECT_NAME]) if len(row) > COL_PROJECT_NAME else None
        utility_id = safe_str(row[COL_UTILITY_ID]) if len(row) > COL_UTILITY_ID else None
        city = safe_str(row[COL_CITY]) if len(row) > COL_CITY else None
        utility = safe_str(row[COL_UTILITY]) if len(row) > COL_UTILITY else None
        developer = safe_str(row[COL_DEVELOPER]) if len(row) > COL_DEVELOPER else None
        year = row[COL_YEAR] if len(row) > COL_YEAR else None

        # Build source_record_id: nrelcs_{state}_{sanitized_name}
        # For duplicate name+state combos, append _2, _3, etc.
        name_part = sanitize_id(project_name) if project_name else f"row_{row_idx}"
        base_source_id = f"nrelcs_{state.lower()}_{name_part}"

        # Handle duplicates by appending counter
        if base_source_id in seen_ids:
            seen_ids[base_source_id] += 1
            source_record_id = f"{base_source_id}_{seen_ids[base_source_id]}"
        else:
            seen_ids[base_source_id] = 1
            source_record_id = base_source_id

        if source_record_id in existing_ids:
            skipped_existing += 1
            continue

        # Parse install date
        install_date = parse_year(year)

        # Location precision: city if we have city, otherwise state
        if city:
            loc_precision = "city"
        elif state:
            loc_precision = "state"
        else:
            loc_precision = None

        # CRITICAL: All records in a batch MUST have identical keys (PostgREST requirement)
        # Never strip None values from records!
        installation = {
            "source_record_id": source_record_id,
            "data_source_id": data_source_id,
            "site_name": project_name[:255] if project_name else None,
            "site_type": "community",
            "site_status": "active",
            "state": state,
            "county": None,
            "city": city[:255] if city else None,
            "zip_code": None,
            "address": None,
            "latitude": None,
            "longitude": None,
            "capacity_mw": capacity_mw,
            "capacity_dc_kw": round(cap_kw_dc, 1) if cap_kw_dc else None,
            "capacity_ac_kw": round(cap_kw_ac, 1),
            "install_date": install_date,
            "owner_name": None,
            "developer_name": developer[:255] if developer else None,
            "operator_name": utility[:255] if utility else None,
            "installer_name": None,
            "mount_type": None,
            "has_battery_storage": None,
            "location_precision": loc_precision,
        }

        installations.append(installation)

    print(f"  New records to ingest: {len(installations)}")
    print(f"  Skipped (aggregated): {skipped_aggregated}")
    print(f"  Skipped (no state): {skipped_no_state}")
    print(f"  Skipped (no capacity): {skipped_no_cap}")
    print(f"  Skipped (< {MIN_CAPACITY_KW_AC} kW-AC): {skipped_small}")
    print(f"  Skipped (already exist): {skipped_existing}")

    # Show stats
    if installations:
        from collections import Counter
        states = Counter(r["state"] for r in installations)
        with_dev = sum(1 for r in installations if r["developer_name"])
        with_utility = sum(1 for r in installations if r["operator_name"])
        with_date = sum(1 for r in installations if r["install_date"])
        with_city = sum(1 for r in installations if r["city"])

        print(f"\n  Field coverage:")
        print(f"    With developer: {with_dev}/{len(installations)} ({with_dev/len(installations)*100:.0f}%)")
        print(f"    With utility: {with_utility}/{len(installations)} ({with_utility/len(installations)*100:.0f}%)")
        print(f"    With install date: {with_date}/{len(installations)} ({with_date/len(installations)*100:.0f}%)")
        print(f"    With city: {with_city}/{len(installations)} ({with_city/len(installations)*100:.0f}%)")
        print(f"    Top 10 states: {states.most_common(10)}")

        print(f"\n  Sample records:")
        for r in installations[:5]:
            print(f"    {r['source_record_id']}: {r['site_name']} ({r['state']}) "
                  f"- {r['capacity_mw']} MW, dev={r['developer_name']}, util={r['operator_name']}")

    if args.dry_run:
        print(f"\n  [DRY RUN] Would create {len(installations)} records")
        return

    if not installations:
        print("\n  No new records to create.")
        return

    # Insert in batches
    print(f"\nStep 6: Inserting {len(installations)} records...")
    total_created = 0
    total_errors = 0

    for i in range(0, len(installations), BATCH_SIZE):
        batch = installations[i:i + BATCH_SIZE]
        created, errors = supabase_post_batch("solar_installations", batch)
        total_created += created
        total_errors += errors
        if (i + BATCH_SIZE) % 200 == 0 or i + BATCH_SIZE >= len(installations):
            print(f"    Progress: {min(i + BATCH_SIZE, len(installations))}/{len(installations)} "
                  f"(created: {total_created}, errors: {total_errors})")

    # Update data source record count and last_import
    supabase_request(
        "PATCH",
        "solar_data_sources",
        {"record_count": total_created, "last_import": datetime.utcnow().isoformat()},
        params={"name": "eq.nrel_community_solar"},
    )

    print(f"\n{'=' * 60}")
    print("NREL Community Solar Ingestion Complete")
    print(f"{'=' * 60}")
    print(f"  Total projects in file: {len(data_rows)}")
    print(f"  Already existed: {skipped_existing}")
    print(f"  Created: {total_created}")
    print(f"  Errors: {total_errors}")


if __name__ == "__main__":
    main()
