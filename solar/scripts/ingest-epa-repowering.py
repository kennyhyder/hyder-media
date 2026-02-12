#!/usr/bin/env python3
"""
EPA RE-Powering America's Land Tracking Matrix Ingestion Script

Downloads and imports solar installations from the EPA RE-Powering initiative,
which tracks renewable energy projects on contaminated lands, landfills, and
mine sites across the U.S.

Data includes:
- Site/Project Name, State, City
- Site Owner, Primary RE Developer Name
- Project Capacity (MW)
- Completion Date (year)
- Type of Site (Superfund, Landfill, Brownfields, etc.)
- RE Type (Solar PV, Solar PV with Battery)

Only ingests solar projects >= 25 kW (0.025 MW) commercial threshold.
All RE-Powering sites are ground-mounted (landfills, brownfields, mine sites).

Source: https://www.epa.gov/re-powering/re-powering-tracking-matrix
Download: https://www.epa.gov/system/files/documents/2024-12/repowering_tracking_matrix_sites_december_2024_508.xlsx

Usage:
  python3 -u scripts/ingest-epa-repowering.py              # Full ingestion
  python3 -u scripts/ingest-epa-repowering.py --dry-run     # Report without ingesting
"""

import os
import sys
import json
import re
import argparse
import urllib.request
import urllib.parse
import urllib.error
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl required. Install with: pip3 install openpyxl")
    sys.exit(1)

from dotenv import load_dotenv

# Load env
env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)

DOWNLOAD_URL = "https://www.epa.gov/system/files/documents/2024-12/repowering_tracking_matrix_sites_december_2024_508.xlsx"
DATA_DIR = Path(__file__).parent.parent / "data" / "epa_repowering"
XLSX_FILENAME = "repowering_tracking_matrix.xlsx"

BATCH_SIZE = 50
MIN_CAPACITY_MW = 0.025  # 25 kW commercial threshold

# Excel structure: header at row 9 (1-indexed), data starts at row 10
# Row 8 (1-indexed) has category headers ("Site Description", "Renewable Energy Info", etc.)
HEADER_ROW_1IDX = 9   # 1-indexed for openpyxl
DATA_START_1IDX = 10   # 1-indexed for openpyxl
SHEET_NAME = "Tracking Matrix"

# Column indices (0-based) confirmed from inspecting the Excel file
COL_SITE_NAME = 0           # Site/Project Name
COL_EPA_REGION = 1          # EPA Region
COL_STATE = 2               # State
COL_CITY = 3                # City
COL_SITE_TYPE = 4           # Type of Site (Superfund, Landfill, Brownfields, etc.)
COL_SITE_OWNER = 5          # Site Owner
COL_OWNERSHIP_TYPE = 6      # Site Ownership Type
COL_PROPERTY_ACREAGE = 7    # Property Acreage
COL_FORMER_USE = 8          # Former Use Description
COL_RE_TYPE = 9             # RE Type (Solar PV, Solar PV with Battery, etc.)
COL_CAPACITY_MW = 10        # Project Capacity (MW)
COL_PROJECT_ACREAGE = 11    # Project Acreage
COL_DEVELOPER = 12          # Primary RE Developer Name
COL_COMPLETION_DATE = 13    # Completion Date (year)
COL_PROJECT_TYPE = 14       # Project Type (Wholesale Electricity, etc.)
COL_FEDERAL_ASSET = 15      # Current/Former Federal Asset
COL_LANDFILL_APP = 16       # Landfill Application (Cross Program)


# ---------------------------------------------------------------------------
# Supabase helpers
# ---------------------------------------------------------------------------

def supabase_request(method, table, data=None, params=None):
    """Make a request to Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(
            f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items()
        )

    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else []
    except urllib.error.HTTPError as e:
        error_body = e.read().decode()[:300]
        print(f"  Supabase error ({e.code}): {error_body}")
        return None


def supabase_post_batch(table, records):
    """POST a batch of records. On duplicate error, insert one by one."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    body = json.dumps(records).encode()
    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
    try:
        urllib.request.urlopen(req)
        return len(records), 0
    except urllib.error.HTTPError as e:
        err = e.read().decode()[:300]
        if "duplicate" in err.lower() or "unique" in err.lower():
            # Insert one by one to skip duplicates
            created = 0
            errors = 0
            for rec in records:
                try:
                    body = json.dumps([rec]).encode()
                    req = urllib.request.Request(url, data=body, headers=headers, method="POST")
                    urllib.request.urlopen(req)
                    created += 1
                except urllib.error.HTTPError:
                    errors += 1
            return created, errors
        print(f"  POST error ({e.code}): {err}")
        return 0, len(records)


def get_existing_source_ids(prefix):
    """Query DB for existing source_record_ids with given prefix."""
    existing = set()
    offset = 0
    while True:
        params = {
            "source_record_id": f"like.{prefix}*",
            "select": "source_record_id",
            "limit": "1000",
            "offset": str(offset),
        }
        rows = supabase_request("GET", "solar_installations", params=params)
        if not rows:
            break
        for r in rows:
            existing.add(r["source_record_id"])
        if len(rows) < 1000:
            break
        offset += 1000
    return existing


def get_or_create_data_source():
    """Get or create the EPA RE-Powering data source record."""
    params = {"name": "eq.epa_repowering", "select": "id"}
    existing = supabase_request("GET", "solar_data_sources", params=params)
    if existing:
        return existing[0]["id"]

    result = supabase_request("POST", "solar_data_sources", {
        "name": "epa_repowering",
        "description": "EPA RE-Powering America's Land Tracking Matrix - Solar installations on contaminated lands, landfills, and mine sites",
        "url": "https://www.epa.gov/re-powering/re-powering-tracking-matrix",
        "record_count": 0,
    })
    # Re-fetch to get auto-generated ID
    existing = supabase_request("GET", "solar_data_sources", params={"name": "eq.epa_repowering", "select": "id"})
    if existing:
        return existing[0]["id"]
    return None


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def safe_str(val):
    """Convert value to cleaned string, returning None for empty/placeholder values."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("n/a", "nan", "none", "na", "null", "-", "\u2014", "unknown", "tbd", "0"):
        return None
    return s


def safe_float(val):
    """Convert value to float, handling None, empty, and placeholder values."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s in ("-", "\u2014", "N/A", "TBD", "Unknown"):
        return None
    try:
        return float(s.replace(",", ""))
    except (ValueError, TypeError):
        return None


def parse_year(val):
    """Parse completion date to install_date. EPA data uses year only (e.g. 2017).
    Also handles multi-year like '2017/2018' or '2020 (Phase 1)'."""
    if not val:
        return None

    # Handle datetime objects from openpyxl
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")

    s = str(val).strip()
    if not s or s in ("-", "\u2014", "N/A", "TBD", "Unknown"):
        return None

    # Extract first 4-digit year
    m = re.match(r'(\d{4})', s)
    if m:
        year = int(m.group(1))
        if 1990 <= year <= 2030:
            return f"{year}-01-01"
    return None


def sanitize_id(name):
    """Sanitize a string for use in source_record_id."""
    if not name:
        return ""
    clean = re.sub(r'[^a-z0-9]', '_', name.lower())
    clean = re.sub(r'_+', '_', clean).strip('_')
    return clean[:50]


# ---------------------------------------------------------------------------
# Download
# ---------------------------------------------------------------------------

def download_data():
    """Download the EPA RE-Powering Excel file if not already present."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = DATA_DIR / XLSX_FILENAME

    if xlsx_path.exists():
        size_kb = xlsx_path.stat().st_size / 1024
        print(f"  Found existing file ({size_kb:.0f} KB), skipping download")
        return xlsx_path

    print(f"  Downloading from EPA...")
    print(f"  URL: {DOWNLOAD_URL}")
    try:
        req = urllib.request.Request(DOWNLOAD_URL, headers={
            "User-Agent": "Mozilla/5.0 (SolarTrack Data Ingestion)"
        })
        with urllib.request.urlopen(req, timeout=60) as resp:
            with open(xlsx_path, "wb") as f:
                f.write(resp.read())
        size_kb = xlsx_path.stat().st_size / 1024
        print(f"  Downloaded {size_kb:.0f} KB")
    except urllib.error.HTTPError as e:
        print(f"  Download failed ({e.code}): {e.reason}")
        print(f"\n  Please manually download from:")
        print(f"  {DOWNLOAD_URL}")
        print(f"  Save to: {xlsx_path}")
        sys.exit(1)

    return xlsx_path


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Ingest EPA RE-Powering solar projects")
    parser.add_argument("--dry-run", action="store_true", help="Report without ingesting")
    args = parser.parse_args()

    print("EPA RE-Powering America's Land - Solar Ingestion Script")
    print("=" * 60)
    print(f"Source: https://www.epa.gov/re-powering")
    print(f"Filter: Solar PV projects >= {MIN_CAPACITY_MW} MW ({MIN_CAPACITY_MW * 1000:.0f} kW)")
    if args.dry_run:
        print("MODE: DRY RUN (no database changes)")
    print()

    # Download data
    print("Step 1: Downloading EPA RE-Powering data...")
    xlsx_path = download_data()

    # Read Excel
    print(f"\nStep 2: Reading {xlsx_path.name}...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find sheet
    if SHEET_NAME in wb.sheetnames:
        ws = wb[SHEET_NAME]
    else:
        print(f"  WARNING: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        ws = wb[wb.sheetnames[0]]
        print(f"  Using first sheet: '{wb.sheetnames[0]}'")

    # Read all rows
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Header at row 9 (1-indexed) = index 8 (0-indexed)
    header_idx = HEADER_ROW_1IDX - 1
    headers = all_rows[header_idx]
    data_rows = all_rows[header_idx + 1:]
    print(f"  Headers (row {HEADER_ROW_1IDX}): {[str(h)[:30] for h in headers if h]}")
    print(f"  Total data rows: {len(data_rows)}")

    # Filter solar PV rows
    solar_rows = []
    for row in data_rows:
        if not any(c for c in row if c):
            continue
        re_type = safe_str(row[COL_RE_TYPE]) if len(row) > COL_RE_TYPE else None
        if re_type and "solar" in re_type.lower():
            solar_rows.append(row)
    print(f"  Solar PV projects: {len(solar_rows)}")

    # Get or create data source
    print("\nStep 3: Setting up data source...")
    if args.dry_run:
        data_source_id = "DRY_RUN"
        print(f"  [DRY RUN] Skipping data source creation")
    else:
        data_source_id = get_or_create_data_source()
        print(f"  Data source ID: {data_source_id}")

    # Pre-check existing records
    print("\nStep 4: Checking for existing records...")
    if args.dry_run:
        existing_ids = set()
    else:
        existing_ids = get_existing_source_ids("epa_repower_")
    print(f"  Existing in DB: {len(existing_ids)}")

    # Build installation records
    print("\nStep 5: Processing solar rows...")
    installations = []
    skipped_no_cap = 0
    skipped_small = 0
    skipped_existing = 0

    for row in solar_rows:
        # Parse capacity
        capacity_mw = safe_float(row[COL_CAPACITY_MW]) if len(row) > COL_CAPACITY_MW else None
        if capacity_mw is None:
            skipped_no_cap += 1
            continue
        if capacity_mw < MIN_CAPACITY_MW:
            skipped_small += 1
            continue

        # Parse fields
        site_name = safe_str(row[COL_SITE_NAME]) if len(row) > COL_SITE_NAME else None
        state = safe_str(row[COL_STATE]) if len(row) > COL_STATE else None
        city = safe_str(row[COL_CITY]) if len(row) > COL_CITY else None
        site_owner = safe_str(row[COL_SITE_OWNER]) if len(row) > COL_SITE_OWNER else None
        developer = safe_str(row[COL_DEVELOPER]) if len(row) > COL_DEVELOPER else None
        completion = row[COL_COMPLETION_DATE] if len(row) > COL_COMPLETION_DATE else None
        re_type = safe_str(row[COL_RE_TYPE]) if len(row) > COL_RE_TYPE else None
        site_type_raw = safe_str(row[COL_SITE_TYPE]) if len(row) > COL_SITE_TYPE else None

        # Normalize state to 2-letter code
        if state and len(state) > 2:
            state = state[:2].upper()
        elif state:
            state = state.upper()

        # Build source_record_id from state + sanitized project name
        name_part = sanitize_id(site_name) if site_name else "unknown"
        state_part = state.lower() if state else "xx"
        source_record_id = f"epa_repower_{state_part}_{name_part}"

        if source_record_id in existing_ids:
            skipped_existing += 1
            continue

        # Classify site type by capacity
        site_type = "utility" if capacity_mw >= 1.0 else "commercial"

        # Battery detection from RE Type
        has_battery = bool(re_type and "battery" in re_type.lower())

        # Parse install date
        install_date = parse_year(completion)

        # Location precision: city if we have city, otherwise state
        if city:
            loc_precision = "city"
        elif state:
            loc_precision = "state"
        else:
            loc_precision = None

        # CRITICAL: All records in a batch MUST have identical keys (PostgREST requirement)
        # Never strip None values from records!
        installation = {
            "source_record_id": source_record_id,
            "data_source_id": data_source_id,
            "site_name": site_name[:255] if site_name else None,
            "site_type": site_type,
            "site_status": "active",
            "state": state,
            "county": None,
            "city": city[:255] if city else None,
            "zip_code": None,
            "address": None,
            "latitude": None,
            "longitude": None,
            "capacity_mw": round(capacity_mw, 6),
            "capacity_dc_kw": round(capacity_mw * 1000, 1),
            "capacity_ac_kw": None,
            "install_date": install_date,
            "owner_name": site_owner[:255] if site_owner else None,
            "developer_name": developer[:255] if developer else None,
            "operator_name": None,
            "installer_name": None,
            "mount_type": "ground",
            "has_battery_storage": has_battery,
            "location_precision": loc_precision,
        }

        installations.append(installation)

    print(f"  New records to ingest: {len(installations)}")
    print(f"  Skipped (no capacity): {skipped_no_cap}")
    print(f"  Skipped (< {MIN_CAPACITY_MW} MW): {skipped_small}")
    print(f"  Skipped (already exist): {skipped_existing}")

    # Show stats
    if installations:
        from collections import Counter
        states = Counter(r["state"] for r in installations)
        with_owner = sum(1 for r in installations if r["owner_name"])
        with_dev = sum(1 for r in installations if r["developer_name"])
        with_date = sum(1 for r in installations if r["install_date"])
        with_battery = sum(1 for r in installations if r["has_battery_storage"])
        utility_count = sum(1 for r in installations if r["site_type"] == "utility")
        commercial_count = sum(1 for r in installations if r["site_type"] == "commercial")

        print(f"\n  Field coverage:")
        print(f"    With owner: {with_owner}/{len(installations)} ({with_owner/len(installations)*100:.0f}%)")
        print(f"    With developer: {with_dev}/{len(installations)} ({with_dev/len(installations)*100:.0f}%)")
        print(f"    With install date: {with_date}/{len(installations)} ({with_date/len(installations)*100:.0f}%)")
        print(f"    With battery: {with_battery}")
        print(f"    Utility-scale (>= 1 MW): {utility_count}")
        print(f"    Commercial (< 1 MW): {commercial_count}")
        print(f"    Top 10 states: {states.most_common(10)}")

        print(f"\n  Sample records:")
        for r in installations[:5]:
            print(f"    {r['source_record_id']}: {r['site_name']} ({r['state']}) "
                  f"- {r['capacity_mw']} MW, owner={r['owner_name']}, dev={r['developer_name']}")

    if args.dry_run:
        print(f"\n  [DRY RUN] Would create {len(installations)} records")
        return

    if not installations:
        print("\n  No new records to create.")
        return

    # Insert in batches
    print(f"\nStep 6: Inserting {len(installations)} records...")
    total_created = 0
    total_errors = 0

    for i in range(0, len(installations), BATCH_SIZE):
        batch = installations[i:i + BATCH_SIZE]
        created, errors = supabase_post_batch("solar_installations", batch)
        total_created += created
        total_errors += errors
        if (i + BATCH_SIZE) % 200 == 0 or i + BATCH_SIZE >= len(installations):
            print(f"    Progress: {min(i + BATCH_SIZE, len(installations))}/{len(installations)} "
                  f"(created: {total_created}, errors: {total_errors})")

    # Update data source record count and last_import
    supabase_request(
        "PATCH",
        "solar_data_sources",
        {"record_count": total_created, "last_import": datetime.utcnow().isoformat()},
        params={"name": "eq.epa_repowering"},
    )

    print(f"\n{'=' * 60}")
    print("EPA RE-Powering Ingestion Complete")
    print(f"{'=' * 60}")
    print(f"  Solar projects found: {len(solar_rows)}")
    print(f"  Already existed: {skipped_existing}")
    print(f"  Created: {total_created}")
    print(f"  Errors: {total_errors}")


if __name__ == "__main__":
    main()
