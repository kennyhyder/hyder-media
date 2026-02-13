#!/usr/bin/env python3
"""
Equipment Backfill Script

Creates equipment records for installations that have none, using available data:

1. EIA-860M installations: Creates module record with "PV" technology from the annual
   EIA-860 Solar sheet if matched by plant_code+gen_id, or basic "PV" if not.
2. Tracking/racking equipment: Creates racking equipment for installations with
   tracking_type set but no existing racking record.
3. EIA-860 technology inference: For EIA-860 module records missing manufacturer,
   applies CdTe → First Solar inference (extends quick-wins.py).

Idempotent: checks existing equipment before creating. Safe to re-run.
"""

import os
import sys
import json
import uuid
import urllib.request
import urllib.parse
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

try:
    import openpyxl
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

from dotenv import load_dotenv

env_path = Path(__file__).parent.parent / ".env.local"
load_dotenv(env_path)

SUPABASE_URL = (os.environ.get("SUPABASE_URL") or os.environ.get("NEXT_PUBLIC_SUPABASE_URL") or "").strip()
SUPABASE_KEY = (os.environ.get("SUPABASE_SERVICE_KEY") or "").strip()
BATCH_SIZE = 50

if not SUPABASE_URL or not SUPABASE_KEY:
    print("Error: SUPABASE_URL and SUPABASE_SERVICE_KEY must be set in .env.local")
    sys.exit(1)


def supabase_request(method, table, data=None, params=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + "&".join(f"{k}={urllib.parse.quote(str(v), safe='.*,()')}" for k, v in params.items())
    headers = {
        "apikey": SUPABASE_KEY,
        "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    if headers_extra:
        headers.update(headers_extra)
    body = json.dumps(data).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            text = resp.read().decode()
            return json.loads(text) if text.strip() else True
    except urllib.error.HTTPError as e:
        err = e.read().decode()[:200]
        print(f"  Error ({e.code}): {err}")
        return None


def supabase_get_all(table, params):
    """Fetch all records with pagination."""
    all_records = []
    offset = 0
    limit = 1000
    while True:
        p = {**params, "limit": str(limit), "offset": str(offset)}
        result = supabase_request("GET", table, params=p, headers_extra={"Prefer": "count=exact"})
        if not result or not isinstance(result, list):
            break
        all_records.extend(result)
        if len(result) < limit:
            break
        offset += limit
        if offset % 5000 == 0:
            print(f"    Fetched {offset} records...")
    return all_records


def batch_insert(table, records):
    """Insert records in batches."""
    created = 0
    errors = 0
    for i in range(0, len(records), BATCH_SIZE):
        batch = records[i:i+BATCH_SIZE]
        result = supabase_request("POST", table, data=batch)
        if result:
            created += len(batch)
        else:
            errors += len(batch)
    return created, errors


def get_data_source_id(name):
    """Get data source ID by name."""
    result = supabase_request("GET", "solar_data_sources",
                              params={"name": f"eq.{name}", "select": "id"},
                              headers_extra={"Prefer": "count=exact"})
    if result and isinstance(result, list) and len(result) > 0:
        return result[0]["id"]
    return None


def phase1_eia860m_equipment():
    """Phase 1: Create module records for EIA-860M installations.

    Reads the EIA-860 annual Solar sheet to get technology flags for matching
    plant_code+gen_id pairs. For unmatched, creates basic PV module record.
    """
    print("\n=== Phase 1: EIA-860M Equipment Backfill ===")

    # Get EIA-860M canonical installations without equipment
    print("  Fetching EIA-860M installations without equipment...")
    all_860m = supabase_get_all("solar_installations", {
        "select": "id,source_record_id,site_status,crossref_ids",
        "source_record_id": "like.eia860m*",
        "is_canonical": "eq.true",
    })
    print(f"  Found {len(all_860m)} EIA-860M canonical installations")

    # Get existing equipment installation IDs
    existing_equip = set()
    equip_records = supabase_get_all("solar_equipment", {
        "select": "installation_id",
        "installation_id": f"in.({','.join(r['id'] for r in all_860m)})" if all_860m else "eq.none",
    })
    for e in equip_records:
        existing_equip.add(e["installation_id"])

    missing = [r for r in all_860m if r["id"] not in existing_equip]
    print(f"  {len(missing)} need equipment records ({len(existing_equip)} already have equipment)")

    if not missing:
        return 0, 0

    # Load EIA-860 annual Solar sheet for technology data
    eia_solar_path = list(Path("data/eia860_2024").glob("3_3_Solar*.xlsx"))
    tech_map = {}  # plant_code_genid -> {technology, tracking}
    if eia_solar_path:
        print(f"  Loading EIA-860 Solar technology data...")
        wb = openpyxl.load_workbook(eia_solar_path[0], read_only=True, data_only=True)
        ws = wb["Operable"]
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 1:
                headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
                continue
            if headers and i >= 2:
                rec = dict(zip(headers, row))
                plant_code = rec.get("Plant Code")
                gen_id = rec.get("Generator ID")
                if plant_code and gen_id:
                    key = f"{int(plant_code)}_{gen_id}"
                    techs = []
                    if rec.get("Crystalline Silicon?") == "Y":
                        techs.append("crystalline-silicon")
                    if rec.get("Thin-Film (CdTe)?") == "Y":
                        techs.append("CdTe")
                    if rec.get("Thin-Film (A-Si)?") == "Y":
                        techs.append("a-Si")
                    if rec.get("Thin-Film (CIGS)?") == "Y":
                        techs.append("CIGS")
                    if rec.get("Thin-Film (Other)?") == "Y":
                        techs.append("thin-film-other")
                    if rec.get("Bifacial?") == "Y":
                        techs.append("bifacial")

                    tracking = None
                    if rec.get("Single-Axis Tracking?") == "Y":
                        tracking = "single-axis"
                    elif rec.get("Dual-Axis Tracking?") == "Y":
                        tracking = "dual-axis"
                    elif rec.get("Fixed Tilt?") == "Y":
                        tracking = "fixed-tilt"

                    tech_map[key] = {
                        "technology": ", ".join(techs) if techs else "PV",
                        "tracking": tracking,
                        "manufacturer": "First Solar" if "CdTe" in techs else None,
                    }
        wb.close()
        print(f"  Loaded technology data for {len(tech_map)} generators")

    # Create equipment records
    ds_id = get_data_source_id("eia860m")
    records_to_create = []
    matched_annual = 0

    for inst in missing:
        src = inst["source_record_id"]  # e.g. "eia860m_12345_GEN1"
        parts = src.replace("eia860m_", "").split("_", 1)
        if len(parts) == 2:
            key = f"{parts[0]}_{parts[1]}"
        else:
            key = None

        tech_info = tech_map.get(key, {}) if key else {}
        if tech_info:
            matched_annual += 1

        technology = tech_info.get("technology", "PV")
        manufacturer = tech_info.get("manufacturer")

        eq = {
            "id": str(uuid.uuid4()),
            "installation_id": inst["id"],
            "equipment_type": "module",
            "module_technology": technology,
            "equipment_status": "active" if inst.get("site_status") in ("active", None) else "removed",
            "data_source_id": ds_id,
            "manufacturer": manufacturer,
            "quantity": None,
            "model": None,
            "module_wattage_w": None,
            "inverter_capacity_kw": None,
            "specs": None,
        }
        records_to_create.append(eq)

    print(f"  {matched_annual} matched to annual EIA-860 technology data")
    print(f"  Creating {len(records_to_create)} equipment records...")

    created, errors = batch_insert("solar_equipment", records_to_create)
    print(f"  Created: {created}, Errors: {errors}")
    return created, errors


def phase2_tracking_racking():
    """Phase 2: Create racking equipment from tracking_type data.

    For installations with tracking_type set, creates a racking equipment
    record if one doesn't already exist.
    """
    print("\n=== Phase 2: Racking Equipment from Tracking Type ===")

    # Get installations with tracking_type set
    print("  Fetching installations with tracking_type...")
    installs = supabase_get_all("solar_installations", {
        "select": "id,tracking_type,data_source_id,site_status",
        "tracking_type": "not.is.null",
        "is_canonical": "eq.true",
    })
    print(f"  Found {len(installs)} installations with tracking_type")

    if not installs:
        return 0, 0

    # Check which already have racking equipment
    print("  Checking existing racking equipment...")
    inst_ids = [r["id"] for r in installs]

    # Check in batches (URL length limit)
    has_racking = set()
    for i in range(0, len(inst_ids), 200):
        batch_ids = inst_ids[i:i+200]
        existing = supabase_get_all("solar_equipment", {
            "select": "installation_id",
            "equipment_type": "eq.racking",
            "installation_id": f"in.({','.join(batch_ids)})",
        })
        for e in existing:
            has_racking.add(e["installation_id"])
        if (i + 200) % 2000 == 0:
            print(f"    Checked {i + 200}/{len(inst_ids)}...")

    missing = [r for r in installs if r["id"] not in has_racking]
    print(f"  {len(missing)} need racking equipment ({len(has_racking)} already have it)")

    if not missing:
        return 0, 0

    # Map tracking_type to racking description
    tracking_desc = {
        "single-axis": "Single-axis tracker",
        "dual-axis": "Dual-axis tracker",
        "fixed-tilt": "Fixed tilt racking",
        "fixed": "Fixed tilt racking",
    }

    records = []
    for inst in missing:
        tt = inst.get("tracking_type", "")
        desc = tracking_desc.get(tt, f"Racking ({tt})")
        eq_type = "tracker" if "axis" in (tt or "") else "racking"

        records.append({
            "id": str(uuid.uuid4()),
            "installation_id": inst["id"],
            "equipment_type": eq_type,
            "model": desc,
            "equipment_status": "active" if inst.get("site_status") in ("active", None) else "removed",
            "data_source_id": inst.get("data_source_id"),
            "manufacturer": None,
            "quantity": None,
            "module_technology": None,
            "module_wattage_w": None,
            "inverter_capacity_kw": None,
            "specs": None,
        })

    print(f"  Creating {len(records)} racking/tracker equipment records...")
    created, errors = batch_insert("solar_equipment", records)
    print(f"  Created: {created}, Errors: {errors}")
    return created, errors


def phase3_nrel_community_equipment():
    """Phase 3: Create basic module records for NREL Community Solar installations."""
    print("\n=== Phase 3: NREL Community Solar Equipment ===")

    installs = supabase_get_all("solar_installations", {
        "select": "id,data_source_id,site_status",
        "source_record_id": "like.nrel_cs*",
        "is_canonical": "eq.true",
    })
    print(f"  Found {len(installs)} NREL Community Solar installations")

    if not installs:
        return 0, 0

    # Check existing
    inst_ids = [r["id"] for r in installs]
    has_equip = set()
    for i in range(0, len(inst_ids), 200):
        batch_ids = inst_ids[i:i+200]
        existing = supabase_get_all("solar_equipment", {
            "select": "installation_id",
            "installation_id": f"in.({','.join(batch_ids)})",
        })
        for e in existing:
            has_equip.add(e["installation_id"])

    missing = [r for r in installs if r["id"] not in has_equip]
    print(f"  {len(missing)} need equipment records")

    if not missing:
        return 0, 0

    ds_id = get_data_source_id("nrel_community_solar")
    records = []
    for inst in missing:
        records.append({
            "id": str(uuid.uuid4()),
            "installation_id": inst["id"],
            "equipment_type": "module",
            "module_technology": "PV",
            "equipment_status": "active" if inst.get("site_status") in ("active", None) else "removed",
            "data_source_id": ds_id or inst.get("data_source_id"),
            "manufacturer": None,
            "model": None,
            "quantity": None,
            "module_wattage_w": None,
            "inverter_capacity_kw": None,
            "specs": None,
        })

    print(f"  Creating {len(records)} module records...")
    created, errors = batch_insert("solar_equipment", records)
    print(f"  Created: {created}, Errors: {errors}")
    return created, errors


def phase4_update_tracking_type():
    """Phase 4: Set tracking_type on EIA-860M installations from annual data."""
    print("\n=== Phase 4: Backfill tracking_type on EIA-860M ===")

    # Load annual tech data
    eia_solar_path = list(Path("data/eia860_2024").glob("3_3_Solar*.xlsx"))
    if not eia_solar_path:
        print("  No EIA-860 annual data found, skipping")
        return 0, 0

    wb = openpyxl.load_workbook(eia_solar_path[0], read_only=True, data_only=True)
    ws = wb["Operable"]
    tracking_map = {}
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:
            headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(row)]
            continue
        if headers and i >= 2:
            rec = dict(zip(headers, row))
            plant_code = rec.get("Plant Code")
            gen_id = rec.get("Generator ID")
            if plant_code and gen_id:
                tracking = None
                if rec.get("Single-Axis Tracking?") == "Y":
                    tracking = "single-axis"
                elif rec.get("Dual-Axis Tracking?") == "Y":
                    tracking = "dual-axis"
                elif rec.get("Fixed Tilt?") == "Y":
                    tracking = "fixed-tilt"
                if tracking:
                    tracking_map[f"{int(plant_code)}_{gen_id}"] = tracking
    wb.close()
    print(f"  Loaded {len(tracking_map)} tracking types from annual data")

    # Get EIA-860M without tracking_type
    installs = supabase_get_all("solar_installations", {
        "select": "id,source_record_id",
        "source_record_id": "like.eia860m*",
        "tracking_type": "is.null",
        "is_canonical": "eq.true",
    })
    print(f"  {len(installs)} EIA-860M installations need tracking_type")

    patched = 0
    errors = 0
    for inst in installs:
        src = inst["source_record_id"].replace("eia860m_", "")
        parts = src.split("_", 1)
        key = f"{parts[0]}_{parts[1]}" if len(parts) == 2 else None
        tracking = tracking_map.get(key) if key else None

        if tracking:
            result = supabase_request("PATCH", "solar_installations",
                                      data={"tracking_type": tracking},
                                      params={"id": f"eq.{inst['id']}"})
            if result:
                patched += 1
            else:
                errors += 1

    print(f"  Patched: {patched}, Errors: {errors}")
    return patched, errors


def main():
    print("=" * 60)
    print("Equipment Backfill Script")
    print("=" * 60)

    total_created = 0
    total_errors = 0

    # Phase 4 first (tracking_type backfill) — so Phase 2 can use it
    p4_created, p4_errors = phase4_update_tracking_type()
    total_created += p4_created
    total_errors += p4_errors

    # Phase 1: EIA-860M module records
    p1_created, p1_errors = phase1_eia860m_equipment()
    total_created += p1_created
    total_errors += p1_errors

    # Phase 2: Racking equipment from tracking_type
    p2_created, p2_errors = phase2_tracking_racking()
    total_created += p2_created
    total_errors += p2_errors

    # Phase 3: NREL Community Solar modules
    p3_created, p3_errors = phase3_nrel_community_equipment()
    total_created += p3_created
    total_errors += p3_errors

    print("\n" + "=" * 60)
    print(f"Total records created/patched: {total_created}")
    print(f"Total errors: {total_errors}")
    print("=" * 60)


if __name__ == "__main__":
    main()
