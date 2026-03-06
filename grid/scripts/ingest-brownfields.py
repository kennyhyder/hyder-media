#!/usr/bin/env python3
"""
Ingest brownfield/retired plant sites from two sources:
1. EIA-860 retired generators — coal/gas/oil/nuclear plants with grid connections
2. EPA RE-Powering Tracking Matrix — brownfield/landfill sites suitable for energy development

Target: grid_brownfield_sites table

These sites are prime datacenter conversion candidates because they already have:
- Grid connections (substations, transformers, switchyards)
- Industrial zoning
- Road access
- Water access (for cooling)

Usage:
  python3 -u scripts/ingest-brownfields.py              # Full ingestion (both sources)
  python3 -u scripts/ingest-brownfields.py --dry-run     # Preview without inserting
  python3 -u scripts/ingest-brownfields.py --eia-only    # EIA-860 retired plants only
  python3 -u scripts/ingest-brownfields.py --epa-only    # EPA RE-Powering only
  python3 -u scripts/ingest-brownfields.py --skip-download  # Use cached files
"""

import os
import sys
import json
import csv
import time
import math
import io
import urllib.request
import urllib.parse
import urllib.error
import zipfile
from datetime import datetime, timezone
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env.local'))

SUPABASE_URL = os.environ.get('SUPABASE_URL') or os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

# EIA-860 2024 data
EIA_860_URL = "https://www.eia.gov/electricity/data/eia860/xls/eia8602024.zip"

# EPA RE-Powering Tracking Matrix
EPA_REPOWER_URL = "https://www.epa.gov/sites/default/files/2021-04/re-powering_tracking_matrix_through_december_2020_508.xlsx"

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'brownfields')
BATCH_SIZE = 50

# Energy source code -> former_use mapping
ENERGY_SOURCE_MAP = {
    # Coal
    'SUB': 'coal', 'BIT': 'coal', 'LIG': 'coal', 'RC': 'coal',
    'ANT': 'coal', 'WC': 'coal', 'SC': 'coal', 'COL': 'coal',
    # Natural Gas
    'NG': 'gas', 'LFG': 'gas', 'OG': 'gas', 'BFG': 'gas',
    # Oil / Petroleum
    'DFO': 'oil', 'RFO': 'oil', 'JF': 'oil', 'KER': 'oil',
    'PC': 'oil', 'WO': 'oil', 'PET': 'oil',
    # Nuclear
    'NUC': 'nuclear', 'UR': 'nuclear',
}

# Technology types that indicate thermal/conventional plants (brownfield candidates)
THERMAL_TECHNOLOGIES = [
    'Steam Turbine',
    'Combustion Turbine',
    'Combined Cycle',
    'Nuclear',
    'Conventional Steam Coal',
    'Natural Gas Steam Turbine',
    'Natural Gas Fired Combined Cycle',
    'Natural Gas Fired Combustion Turbine',
    'Conventional Hydroelectric',  # Include hydro for completeness
    'Nuclear Steam',
    'Petroleum Liquids',
    'Coal Integrated Gasification Combined Cycle',
    'Other Gases',
    'Fluidized Bed Combustion',
    'Landfill Gas',
    'Municipal Solid Waste',
    'Wood/Wood Waste Biomass',
    'Geothermal',
]

# Technologies to EXCLUDE (not brownfield conversion candidates)
EXCLUDE_TECHNOLOGIES = [
    'Solar Photovoltaic',
    'Solar Thermal',
    'Onshore Wind Turbine',
    'Offshore Wind Turbine',
    'Batteries',
    'Battery',
    'Flywheel',
    'Hydroelectric Pumped Storage',
    'All Other',  # Too vague
]


def supabase_request(method, path, data=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
    }
    if headers_extra:
        headers.update(headers_extra)
    body = json.dumps(data, allow_nan=False).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                text = resp.read().decode()
                return json.loads(text) if text else None
        except urllib.error.HTTPError as e:
            error_body = e.read().decode() if e.fp else ''
            if e.code in (500, 502, 503) and attempt < 2:
                print(f"  HTTP {e.code}, retrying in {2 ** attempt}s...")
                time.sleep(2 ** attempt)
                continue
            print(f"  HTTP {e.code}: {error_body[:500]}")
            raise
        except Exception as e:
            if attempt < 2:
                print(f"  Error: {e}, retrying in {2 ** attempt}s...")
                time.sleep(2 ** attempt)
                continue
            raise


def safe_float(val):
    if val is None or val == '' or val == ' ':
        return None
    try:
        f = float(val)
        return f if not math.isnan(f) and not math.isinf(f) else None
    except (ValueError, TypeError):
        return None


def safe_str(val, max_len=500):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('none', 'null', 'n/a', 'unknown', '', '.'):
        return None
    return s[:max_len] if len(s) > max_len else s


def safe_int(val):
    if val is None or val == '' or val == ' ':
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def download_eia860(skip_download=False):
    """Download and extract EIA-860 2024 ZIP."""
    os.makedirs(DATA_DIR, exist_ok=True)
    eia_dir = os.path.join(DATA_DIR, 'eia860_2024')
    os.makedirs(eia_dir, exist_ok=True)

    # Check if we already have the files we need
    generator_file = None
    plant_file = None
    for f in os.listdir(eia_dir) if os.path.exists(eia_dir) else []:
        if '3_1_Generator' in f and f.endswith('.xlsx'):
            generator_file = os.path.join(eia_dir, f)
        if '2___Plant' in f and f.endswith('.xlsx'):
            plant_file = os.path.join(eia_dir, f)

    if generator_file and plant_file:
        print(f"  Using cached EIA-860 files:")
        print(f"    Generator: {os.path.basename(generator_file)}")
        print(f"    Plant: {os.path.basename(plant_file)}")
        return generator_file, plant_file

    # Also check the solar project's EIA-860 data directory
    solar_eia_dir = os.path.join(os.path.dirname(__file__), '..', '..', 'solar', 'data', 'eia860_2024')
    if os.path.exists(solar_eia_dir):
        for f in os.listdir(solar_eia_dir):
            if '3_1_Generator' in f and f.endswith('.xlsx'):
                generator_file = os.path.join(solar_eia_dir, f)
            if '2___Plant' in f and f.endswith('.xlsx'):
                plant_file = os.path.join(solar_eia_dir, f)
        if generator_file and plant_file:
            print(f"  Using EIA-860 files from solar project:")
            print(f"    Generator: {generator_file}")
            print(f"    Plant: {plant_file}")
            return generator_file, plant_file

    if skip_download:
        print("ERROR: EIA-860 files not found and --skip-download specified")
        return None, None

    # Download the full ZIP
    zip_path = os.path.join(DATA_DIR, 'eia8602024.zip')
    if not os.path.exists(zip_path):
        print(f"  Downloading EIA-860 2024 ZIP...")
        req = urllib.request.Request(EIA_860_URL)
        req.add_header('User-Agent', 'GridScout/1.0')
        with urllib.request.urlopen(req, timeout=300) as resp:
            data = resp.read()
        with open(zip_path, 'wb') as f:
            f.write(data)
        print(f"  Downloaded {len(data) / (1024*1024):.1f} MB")

    # Extract needed files
    print(f"  Extracting EIA-860 ZIP...")
    with zipfile.ZipFile(zip_path, 'r') as zf:
        for name in zf.namelist():
            if '3_1_Generator' in name and name.endswith('.xlsx'):
                zf.extract(name, eia_dir)
                generator_file = os.path.join(eia_dir, name)
                print(f"    Extracted: {name}")
            elif '2___Plant' in name and name.endswith('.xlsx'):
                zf.extract(name, eia_dir)
                plant_file = os.path.join(eia_dir, name)
                print(f"    Extracted: {name}")

    if not generator_file or not plant_file:
        # Try alternate naming
        with zipfile.ZipFile(zip_path, 'r') as zf:
            all_names = zf.namelist()
            print(f"  ZIP contents ({len(all_names)} files):")
            for n in sorted(all_names):
                if n.endswith('.xlsx'):
                    print(f"    {n}")

        print("ERROR: Could not find required EIA-860 files in ZIP")
        return None, None

    return generator_file, plant_file


def parse_eia860_retired(generator_file, plant_file):
    """Parse EIA-860 for retired coal/gas/oil/nuclear generators.

    Returns list of records for grid_brownfield_sites.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install: pip3 install openpyxl")
        sys.exit(1)

    # Step 1: Load plant data for lat/lng, address, county
    print("  Loading Plant sheet for coordinates...")
    plant_wb = openpyxl.load_workbook(plant_file, read_only=True, data_only=True)

    # Find the right sheet
    plant_sheet = None
    for name in plant_wb.sheetnames:
        if 'plant' in name.lower():
            plant_sheet = plant_wb[name]
            break
    if not plant_sheet:
        plant_sheet = plant_wb[plant_wb.sheetnames[0]]

    # Find header row (look for 'Plant Code' or 'Plant Id')
    plant_data = {}
    header_row = None
    headers = []
    for row_idx, row in enumerate(plant_sheet.iter_rows(values_only=True), 1):
        if row is None:
            continue
        row_str = [str(c).strip() if c else '' for c in row]
        # Look for header row (only match the FIRST occurrence)
        if not header_row and any('plant' in c.lower() and ('code' in c.lower() or 'id' in c.lower()) for c in row_str):
            header_row = row_idx
            headers = row_str
            continue
        if header_row and row_idx > header_row:
            if not row[0]:
                continue
            record = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = val
            plant_id = safe_int(record.get('Plant Code') or record.get('Plant Id')
                                or record.get('Plant ID'))
            if plant_id:
                plant_data[plant_id] = {
                    'name': safe_str(record.get('Plant Name')),
                    'state': safe_str(record.get('State')),
                    'county': safe_str(record.get('County')),
                    'city': safe_str(record.get('City')),
                    'latitude': safe_float(record.get('Latitude')),
                    'longitude': safe_float(record.get('Longitude')),
                    'address': safe_str(record.get('Street Address')),
                    'zip': safe_str(record.get('Zip') or record.get('Zip Code')),
                }

    plant_wb.close()
    print(f"    {len(plant_data)} plants loaded with coordinates")

    # Step 2: Load generator data — find retired generators
    print("  Loading Generator sheet for retired units...")
    gen_wb = openpyxl.load_workbook(generator_file, read_only=True, data_only=True)

    # Try "Retired and Canceled" sheet first, then "Retired"
    gen_sheet = None
    for name in gen_wb.sheetnames:
        name_lower = name.lower()
        if 'retired' in name_lower and 'cancel' in name_lower:
            gen_sheet = gen_wb[name]
            print(f"    Using sheet: {name}")
            break
    if not gen_sheet:
        for name in gen_wb.sheetnames:
            if 'retired' in name.lower():
                gen_sheet = gen_wb[name]
                print(f"    Using sheet: {name}")
                break
    if not gen_sheet:
        # Fall back to main generator sheet and filter by status
        for name in gen_wb.sheetnames:
            if 'operable' not in name.lower() and 'generator' in name.lower():
                gen_sheet = gen_wb[name]
                print(f"    Using sheet: {name}")
                break
    if not gen_sheet:
        gen_sheet = gen_wb[gen_wb.sheetnames[0]]
        print(f"    Using first sheet: {gen_wb.sheetnames[0]}")

    # Parse header
    header_row = None
    headers = []
    retired_generators = []

    for row_idx, row in enumerate(gen_sheet.iter_rows(values_only=True), 1):
        if row is None:
            continue
        row_str = [str(c).strip() if c else '' for c in row]

        # Look for header row with generator-specific columns
        if not header_row:
            if any('plant' in c.lower() and ('code' in c.lower() or 'id' in c.lower()) for c in row_str):
                header_row = row_idx
                headers = row_str
                continue
            continue

        if not row[0]:
            continue

        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val

        plant_id = safe_int(record.get('Plant Code') or record.get('Plant Id')
                            or record.get('Plant ID'))
        if not plant_id:
            continue

        # Get technology and energy source
        technology = safe_str(record.get('Technology') or record.get('Prime Mover'))
        energy_source = safe_str(record.get('Energy Source 1') or record.get('Energy Source Code')
                                 or record.get('Energy Source'))
        status = safe_str(record.get('Status') or record.get('Operating Status')
                          or record.get('Plant Status'))
        nameplate = safe_float(record.get('Nameplate Capacity (MW)')
                               or record.get('Nameplate Capacity')
                               or record.get('Generator Nameplate Capacity (MW)'))
        gen_id = safe_str(record.get('Generator ID') or record.get('Generator Id'))

        # Get retirement date
        retire_year = safe_int(record.get('Retirement Year')
                               or record.get('Operating Year')
                               or record.get('Planned Retirement Year'))
        retire_month = safe_int(record.get('Retirement Month')
                                or record.get('Planned Retirement Month'))

        # Determine if this is a retired thermal plant
        is_retired = True  # On the "Retired and Canceled" sheet, all are retired
        if status:
            status_upper = status.upper()
            if 'CANCEL' in status_upper:
                continue  # Skip canceled, only want actually retired
            if 'RETIRED' not in status_upper and 'RE' not in status_upper:
                # If we're on the main generator sheet, filter by status
                if 'retired' not in (gen_sheet.title or '').lower():
                    continue

        # Determine former_use from energy source
        former_use = None
        if energy_source:
            former_use = ENERGY_SOURCE_MAP.get(energy_source.upper().strip())

        # Skip solar/wind/battery (not brownfield conversion candidates)
        if technology:
            tech_lower = technology.lower()
            if any(excl.lower() in tech_lower for excl in EXCLUDE_TECHNOLOGIES):
                continue
            # If no former_use from energy source, try technology
            if not former_use:
                if 'coal' in tech_lower:
                    former_use = 'coal'
                elif 'gas' in tech_lower or 'combustion' in tech_lower:
                    former_use = 'gas'
                elif 'nuclear' in tech_lower:
                    former_use = 'nuclear'
                elif 'petroleum' in tech_lower or 'oil' in tech_lower:
                    former_use = 'oil'
                elif 'hydro' in tech_lower:
                    former_use = 'hydro'
                elif 'biomass' in tech_lower or 'wood' in tech_lower:
                    former_use = 'biomass'
                elif 'waste' in tech_lower or 'landfill' in tech_lower:
                    former_use = 'waste'
                elif 'geothermal' in tech_lower:
                    former_use = 'geothermal'
                else:
                    former_use = 'other'

        if not former_use:
            former_use = 'other'

        # Skip if no capacity (can't be meaningful brownfield)
        if not nameplate or nameplate < 1.0:
            continue

        retired_generators.append({
            'plant_id': plant_id,
            'gen_id': gen_id,
            'technology': technology,
            'energy_source': energy_source,
            'former_use': former_use,
            'nameplate_mw': nameplate,
            'retire_year': retire_year,
            'retire_month': retire_month,
        })

    gen_wb.close()
    print(f"    {len(retired_generators)} retired thermal generators found (>= 1 MW)")

    # Step 3: Aggregate by plant (sum capacity, pick dominant fuel)
    plant_gens = {}
    for gen in retired_generators:
        pid = gen['plant_id']
        if pid not in plant_gens:
            plant_gens[pid] = []
        plant_gens[pid].append(gen)

    print(f"    {len(plant_gens)} unique plants")

    # Step 4: Build records
    records = []
    for plant_id, gens in plant_gens.items():
        plant_info = plant_data.get(plant_id, {})
        lat = plant_info.get('latitude')
        lng = plant_info.get('longitude')

        # Sum capacity across all retired generators at this plant
        total_capacity = sum(g['nameplate_mw'] for g in gens if g['nameplate_mw'])

        # Dominant fuel type (by capacity)
        fuel_capacity = {}
        for g in gens:
            fu = g['former_use']
            fuel_capacity[fu] = fuel_capacity.get(fu, 0) + (g['nameplate_mw'] or 0)
        dominant_fuel = max(fuel_capacity, key=fuel_capacity.get) if fuel_capacity else 'other'

        # Latest retirement date
        latest_year = max((g['retire_year'] for g in gens if g['retire_year']), default=None)
        latest_month = None
        if latest_year:
            year_gens = [g for g in gens if g['retire_year'] == latest_year and g['retire_month']]
            if year_gens:
                latest_month = max(g['retire_month'] for g in year_gens)

        retire_date = None
        if latest_year:
            month = latest_month or 1
            retire_date = f"{latest_year}-{month:02d}-01"

        # Get voltage from plant (if available — not always in EIA-860)
        # We don't have voltage in the standard Plant sheet but it's a schema field
        voltage_kv = None

        plant_name = plant_info.get('name')
        state = plant_info.get('state')

        source_id = f"eia_retired_{plant_id}"

        record = {
            'source_record_id': source_id,
            'name': plant_name,
            'site_type': 'retired_plant',
            'former_use': dominant_fuel,
            'state': state,
            'county': plant_info.get('county'),
            'city': plant_info.get('city'),
            'latitude': lat,
            'longitude': lng,
            'eia_plant_id': plant_id,
            'existing_capacity_mw': round(total_capacity, 2) if total_capacity else None,
            'retirement_date': retire_date,
            'grid_connection_voltage_kv': voltage_kv,
            'acreage': None,
            'epa_id': None,
            'cleanup_status': None,
            'contaminant_type': None,
        }
        records.append(record)

    # Sort by capacity descending
    records.sort(key=lambda r: r.get('existing_capacity_mw') or 0, reverse=True)

    return records


def download_epa_repowering(skip_download=False):
    """Download EPA RE-Powering Tracking Matrix or use cached copy."""
    os.makedirs(DATA_DIR, exist_ok=True)
    epa_path = os.path.join(DATA_DIR, 'repowering_tracking_matrix.xlsx')

    # Check local cache first
    if os.path.exists(epa_path):
        size_kb = os.path.getsize(epa_path) / 1024
        if size_kb > 10:
            print(f"  Using cached EPA file ({size_kb:.1f} KB)")
            return epa_path

    # Check solar project's copy
    solar_epa = os.path.join(os.path.dirname(__file__), '..', '..', 'solar', 'data',
                             'epa_repowering', 'repowering_tracking_matrix.xlsx')
    if os.path.exists(solar_epa):
        size_kb = os.path.getsize(solar_epa) / 1024
        if size_kb > 10:
            print(f"  Using EPA file from solar project ({size_kb:.1f} KB)")
            return solar_epa

    if skip_download:
        print("ERROR: EPA RE-Powering file not found and --skip-download specified")
        return None

    # Download
    print(f"  Downloading EPA RE-Powering Tracking Matrix...")
    urls_to_try = [
        EPA_REPOWER_URL,
        "https://www.epa.gov/system/files/documents/2024-01/re-powering_tracking_matrix_12-2023.xlsx",
        "https://www.epa.gov/re-powering/re-powering-tracking-matrix",
    ]
    for url in urls_to_try:
        try:
            req = urllib.request.Request(url)
            req.add_header('User-Agent', 'GridScout/1.0')
            with urllib.request.urlopen(req, timeout=120) as resp:
                data = resp.read()
                if len(data) > 10000:
                    with open(epa_path, 'wb') as f:
                        f.write(data)
                    print(f"    Downloaded {len(data) / 1024:.1f} KB from {url}")
                    return epa_path
        except Exception as e:
            print(f"    Failed {url}: {e}")

    return None


def parse_epa_repowering(epa_file):
    """Parse EPA RE-Powering Tracking Matrix for brownfield sites."""
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install: pip3 install openpyxl")
        sys.exit(1)

    print("  Loading EPA RE-Powering spreadsheet...")
    wb = openpyxl.load_workbook(epa_file, read_only=True, data_only=True)

    # Find the main data sheet
    sheet = None
    for name in wb.sheetnames:
        name_lower = name.lower()
        if 'data' in name_lower or 'tracking' in name_lower or 'matrix' in name_lower:
            sheet = wb[name]
            break
    if not sheet:
        sheet = wb[wb.sheetnames[0]]
    print(f"    Using sheet: {sheet.title}")

    # Parse headers — EPA RE-Powering has headers around row 9
    # Known columns: Site/Project Name, EPA Region, State, City, Type of Site,
    # Site Owner, Site Ownership Type, Property Acreage, Former Use Description,
    # RE Type, Project Capacity (MW), Project Acreage, Primary RE Developer Name,
    # Completion Date, Project Type
    header_row = None
    headers = []
    records = []

    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row is None:
            continue
        row_str = [str(c).strip().lower() if c else '' for c in row]

        # Find header row — look for "state" in column position (not in a description paragraph)
        if not header_row:
            # Header row has "State" as a short cell value (not inside a paragraph)
            has_state = False
            has_site = False
            for c in row_str:
                if c == 'state':
                    has_state = True
                if 'site' in c and ('name' in c or 'project' in c):
                    has_site = True
            if has_state and has_site:
                header_row = row_idx
                headers = [str(c).strip() if c else f'col_{i}' for i, c in enumerate(row)]
                print(f"    Header at row {row_idx}: {headers[:10]}...")
                continue
            continue

        # Parse data rows
        record = {}
        for i, val in enumerate(row):
            if i < len(headers):
                record[headers[i]] = val

        # Map known columns from EPA RE-Powering format
        name = safe_str(record.get('Site/Project Name'))
        state = safe_str(record.get('State'))
        city = safe_str(record.get('City'))
        site_type_raw = safe_str(record.get('Type of Site'))
        owner = safe_str(record.get('Site Owner'))
        acreage = safe_float(record.get('Property Acreage'))
        former_use_raw = safe_str(record.get('Former Use Description'))
        capacity_raw = safe_float(record.get('Project Capacity (MW)'))
        developer = safe_str(record.get('Primary RE Developer Name'))
        completion = safe_str(record.get('Completion Date'))
        project_type = safe_str(record.get('Project Type'))

        if not state:
            continue
        if state and len(state) > 2:
            state = STATE_ABBREVS.get(state.upper(), state)

        # Determine former_use from Type of Site + Former Use Description
        former_use = 'brownfield'
        if site_type_raw:
            st_lower = site_type_raw.lower()
            if 'landfill' in st_lower:
                former_use = 'landfill'
            elif 'mine' in st_lower:
                former_use = 'mine'
            elif 'superfund' in st_lower:
                former_use = 'superfund'
            elif 'rcra' in st_lower:
                former_use = 'rcra'
            elif 'brown' in st_lower:
                former_use = 'brownfield'
            elif 'contam' in st_lower:
                former_use = 'contaminated'

        if former_use == 'brownfield' and former_use_raw:
            fu_lower = former_use_raw.lower()
            if 'landfill' in fu_lower:
                former_use = 'landfill'
            elif 'mine' in fu_lower or 'mining' in fu_lower:
                former_use = 'mine'
            elif 'military' in fu_lower:
                former_use = 'military'
            elif 'industrial' in fu_lower or 'manufacturing' in fu_lower:
                former_use = 'manufacturing'
            elif 'power' in fu_lower or 'plant' in fu_lower or 'energy' in fu_lower:
                former_use = 'power_plant'

        # Create unique ID
        idx = len(records) + 1
        name_key = ''
        if name:
            name_key = name.lower().replace(' ', '_').replace('/', '_')[:30]
        source_id = f"epa_repower_{name_key}_{state}_{idx}"

        # Acreage from Property or Project Acreage
        if not acreage:
            acreage = safe_float(record.get('Project Acreage'))

        record_out = {
            'source_record_id': source_id,
            'name': name,
            'site_type': 'epa_brownfield',
            'former_use': former_use,
            'state': state,
            'county': None,
            'city': city,
            'latitude': None,  # EPA RE-Powering file has no lat/lng columns
            'longitude': None,
            'eia_plant_id': None,
            'existing_capacity_mw': capacity_raw,
            'retirement_date': None,
            'grid_connection_voltage_kv': None,
            'acreage': acreage,
            'epa_id': None,
            'cleanup_status': 'cleanup_complete',  # RE-Powering tracks completed installations
            'contaminant_type': None,
        }
        records.append(record_out)

    wb.close()
    return records


def get_existing_ids():
    existing = set()
    offset = 0
    while True:
        result = supabase_request(
            'GET',
            f'grid_brownfield_sites?select=source_record_id&limit=1000&offset={offset}'
        )
        if not result:
            break
        for r in result:
            existing.add(r['source_record_id'])
        if len(result) < 1000:
            break
        offset += 1000
    return existing


STATE_ABBREVS = {
    'ALABAMA': 'AL', 'ALASKA': 'AK', 'ARIZONA': 'AZ', 'ARKANSAS': 'AR',
    'CALIFORNIA': 'CA', 'COLORADO': 'CO', 'CONNECTICUT': 'CT', 'DELAWARE': 'DE',
    'DISTRICT OF COLUMBIA': 'DC', 'FLORIDA': 'FL', 'GEORGIA': 'GA', 'HAWAII': 'HI',
    'IDAHO': 'ID', 'ILLINOIS': 'IL', 'INDIANA': 'IN', 'IOWA': 'IA',
    'KANSAS': 'KS', 'KENTUCKY': 'KY', 'LOUISIANA': 'LA', 'MAINE': 'ME',
    'MARYLAND': 'MD', 'MASSACHUSETTS': 'MA', 'MICHIGAN': 'MI', 'MINNESOTA': 'MN',
    'MISSISSIPPI': 'MS', 'MISSOURI': 'MO', 'MONTANA': 'MT', 'NEBRASKA': 'NE',
    'NEVADA': 'NV', 'NEW HAMPSHIRE': 'NH', 'NEW JERSEY': 'NJ', 'NEW MEXICO': 'NM',
    'NEW YORK': 'NY', 'NORTH CAROLINA': 'NC', 'NORTH DAKOTA': 'ND', 'OHIO': 'OH',
    'OKLAHOMA': 'OK', 'OREGON': 'OR', 'PENNSYLVANIA': 'PA', 'RHODE ISLAND': 'RI',
    'SOUTH CAROLINA': 'SC', 'SOUTH DAKOTA': 'SD', 'TENNESSEE': 'TN', 'TEXAS': 'TX',
    'UTAH': 'UT', 'VERMONT': 'VT', 'VIRGINIA': 'VA', 'WASHINGTON': 'WA',
    'WEST VIRGINIA': 'WV', 'WISCONSIN': 'WI', 'WYOMING': 'WY',
    'PUERTO RICO': 'PR', 'GUAM': 'GU', 'VIRGIN ISLANDS': 'VI',
}


def main():
    print("=" * 60)
    print("GridScout Brownfield / Retired Plant Ingestion")
    print("=" * 60)

    dry_run = '--dry-run' in sys.argv
    eia_only = '--eia-only' in sys.argv
    epa_only = '--epa-only' in sys.argv
    skip_download = '--skip-download' in sys.argv

    all_records = []

    # ---- EIA-860 Retired Plants ----
    if not epa_only:
        print("\n--- EIA-860 Retired Power Plants ---")
        generator_file, plant_file = download_eia860(skip_download)
        if generator_file and plant_file:
            eia_records = parse_eia860_retired(generator_file, plant_file)
            print(f"\n  {len(eia_records)} retired plant sites parsed")

            # Summary by former_use
            use_counts = {}
            use_capacity = {}
            state_counts = {}
            for r in eia_records:
                fu = r.get('former_use', 'unknown')
                use_counts[fu] = use_counts.get(fu, 0) + 1
                use_capacity[fu] = use_capacity.get(fu, 0) + (r.get('existing_capacity_mw') or 0)
                st = r.get('state', '??')
                state_counts[st] = state_counts.get(st, 0) + 1

            print(f"\n  By former use:")
            for fu, cnt in sorted(use_counts.items(), key=lambda x: -x[1]):
                cap = use_capacity.get(fu, 0)
                print(f"    {fu:12s}: {cnt:5d} plants, {cap:10,.1f} MW total")

            print(f"\n  By state (top 15):")
            for st, cnt in sorted(state_counts.items(), key=lambda x: -x[1])[:15]:
                print(f"    {st}: {cnt}")

            has_coords = sum(1 for r in eia_records if r.get('latitude') and r.get('longitude'))
            print(f"\n  With coordinates: {has_coords}/{len(eia_records)} ({100*has_coords/len(eia_records):.1f}%)")

            # Top 10 by capacity
            print(f"\n  Top 10 by capacity:")
            for r in eia_records[:10]:
                name = r.get('name') or 'Unknown'
                state = r.get('state') or '??'
                fu = r.get('former_use') or '?'
                cap = r.get('existing_capacity_mw') or 0
                rd = r.get('retirement_date') or 'N/A'
                print(f"    {name:40s} {state:2s} {fu:8s} {cap:8.1f} MW retired {rd}")

            all_records.extend(eia_records)
        else:
            print("  SKIPPED: Could not find/download EIA-860 files")

    # ---- EPA RE-Powering Brownfields ----
    if not eia_only:
        print("\n--- EPA RE-Powering Brownfield Sites ---")
        epa_file = download_epa_repowering(skip_download)
        if epa_file:
            epa_records = parse_epa_repowering(epa_file)
            print(f"\n  {len(epa_records)} EPA brownfield sites parsed")

            # Summary
            use_counts = {}
            state_counts = {}
            for r in epa_records:
                fu = r.get('former_use', 'unknown')
                use_counts[fu] = use_counts.get(fu, 0) + 1
                st = r.get('state', '??')
                state_counts[st] = state_counts.get(st, 0) + 1

            print(f"\n  By former use:")
            for fu, cnt in sorted(use_counts.items(), key=lambda x: -x[1]):
                print(f"    {fu:15s}: {cnt}")

            has_coords = sum(1 for r in epa_records if r.get('latitude') and r.get('longitude'))
            print(f"\n  With coordinates: {has_coords}/{len(epa_records)} ({100*has_coords/max(len(epa_records),1):.1f}%)")

            print(f"\n  By state (top 10):")
            for st, cnt in sorted(state_counts.items(), key=lambda x: -x[1])[:10]:
                print(f"    {st}: {cnt}")

            all_records.extend(epa_records)
        else:
            print("  SKIPPED: Could not find/download EPA RE-Powering file")

    # ---- Summary ----
    print(f"\n{'=' * 60}")
    print(f"Total records: {len(all_records)}")
    type_counts = {}
    for r in all_records:
        t = r.get('site_type', '?')
        type_counts[t] = type_counts.get(t, 0) + 1
    for t, c in sorted(type_counts.items()):
        print(f"  {t}: {c}")

    if dry_run:
        print(f"\n[DRY RUN] Would insert {len(all_records)} brownfield records.")
        return

    if not all_records:
        print("\nNo records to insert.")
        return

    # ---- Insert into Supabase ----

    # Get/create data source IDs
    ds_eia = supabase_request('GET', 'grid_data_sources?name=eq.eia_retired_plants&select=id')
    ds_eia_id = ds_eia[0]['id'] if ds_eia else None
    if not ds_eia_id:
        print("  Creating eia_retired_plants data source...")
        result = supabase_request('POST', 'grid_data_sources', [{
            'name': 'eia_retired_plants',
            'url': 'https://www.eia.gov/electricity/data/eia860/',
            'description': 'EIA-860 Retired Power Plants',
        }], {'Prefer': 'return=representation'})
        if result:
            ds_eia_id = result[0]['id']

    ds_epa = supabase_request('GET', 'grid_data_sources?name=eq.epa_brownfields&select=id')
    ds_epa_id = ds_epa[0]['id'] if ds_epa else None
    if not ds_epa_id:
        print("  Creating epa_brownfields data source...")
        result = supabase_request('POST', 'grid_data_sources', [{
            'name': 'epa_brownfields',
            'url': 'https://www.epa.gov/re-powering',
            'description': 'EPA RE-Powering Brownfield Sites',
        }], {'Prefer': 'return=representation'})
        if result:
            ds_epa_id = result[0]['id']

    # Assign data_source_id based on site_type
    for r in all_records:
        if r['site_type'] == 'retired_plant':
            r['data_source_id'] = ds_eia_id
        else:
            r['data_source_id'] = ds_epa_id
        r['created_at'] = datetime.now(timezone.utc).isoformat()

    # Load existing records
    print("\nLoading existing records...")
    existing_ids = get_existing_ids()
    print(f"  {len(existing_ids)} existing records in DB")

    new_records = [r for r in all_records if r['source_record_id'] not in existing_ids]
    print(f"  {len(new_records)} new records to insert ({len(all_records) - len(new_records)} already exist)")

    # Insert in batches
    created = 0
    errors = 0
    eia_created = 0
    epa_created = 0

    for i in range(0, len(new_records), BATCH_SIZE):
        batch = new_records[i:i + BATCH_SIZE]
        try:
            supabase_request('POST', 'grid_brownfield_sites', batch, {'Prefer': 'return=minimal'})
            created += len(batch)
            for b in batch:
                if b['site_type'] == 'retired_plant':
                    eia_created += 1
                else:
                    epa_created += 1
        except Exception as e:
            print(f"  Batch error at {i}: {e}")
            for rec in batch:
                try:
                    supabase_request('POST', 'grid_brownfield_sites', [rec],
                                     {'Prefer': 'return=minimal'})
                    created += 1
                    if rec['site_type'] == 'retired_plant':
                        eia_created += 1
                    else:
                        epa_created += 1
                except Exception as e2:
                    errors += 1
                    if errors <= 10:
                        print(f"  Record error ({rec['source_record_id']}): {e2}")

        if (i // BATCH_SIZE) % 10 == 0 and i > 0:
            print(f"  Progress: {min(i + BATCH_SIZE, len(new_records))}/{len(new_records)} "
                  f"({created} created, {errors} errors)")

    # Update data source counts
    if ds_eia_id and eia_created > 0:
        eia_total = sum(1 for r in all_records if r['site_type'] == 'retired_plant')
        supabase_request('PATCH', f'grid_data_sources?id=eq.{ds_eia_id}', {
            'record_count': eia_total,
            'last_import': datetime.now(timezone.utc).isoformat()
        })

    if ds_epa_id and epa_created > 0:
        epa_total = sum(1 for r in all_records if r['site_type'] == 'epa_brownfield')
        supabase_request('PATCH', f'grid_data_sources?id=eq.{ds_epa_id}', {
            'record_count': epa_total,
            'last_import': datetime.now(timezone.utc).isoformat()
        })

    print(f"\n{'=' * 60}")
    print(f"Brownfield Ingestion Complete")
    print(f"  EIA-860 retired plants created: {eia_created}")
    print(f"  EPA brownfield sites created: {epa_created}")
    print(f"  Total created: {created}")
    print(f"  Skipped (existing): {len(all_records) - len(new_records)}")
    print(f"  Errors: {errors}")
    print(f"  Total in DB: {len(existing_ids) + created}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
