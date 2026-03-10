#!/usr/bin/env python3
"""
Ingest proposed generators from EIA-860M monthly data to identify where utilities
are adding power capacity (natural gas, battery storage, solar).

These planned generators indicate DC-ready power additions — sites where grid
capacity is being expanded, making them ideal for datacenter siting.

Downloads the latest EIA-860M Excel file, parses the "Proposed" sheet/tab, and
stores results in grid_proposed_generators table + enriches grid_county_data with
proposed_capacity_mw.

Usage:
  python3 -u scripts/ingest-eia860m-proposed.py              # Download + ingest
  python3 -u scripts/ingest-eia860m-proposed.py --dry-run    # Preview without inserting
  python3 -u scripts/ingest-eia860m-proposed.py --skip-download  # Use cached file
  python3 -u scripts/ingest-eia860m-proposed.py --file /path/to/file.xlsx  # Custom file
"""

import os
import sys
import json
import time
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime, timezone
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env.local'))

SUPABASE_URL = os.environ.get('SUPABASE_URL') or os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

BATCH_SIZE = 50
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'eia860m')

# EIA-860M download URL — updated monthly. Try february 2025 first, fall back.
EIA_860M_URLS = [
    "https://www.eia.gov/electricity/data/eia860m/xls/february_generator2025.xlsx",
    "https://www.eia.gov/electricity/data/eia860m/xls/january_generator2025.xlsx",
    "https://www.eia.gov/electricity/data/eia860m/xls/december_generator2024.xlsx",
]

# Technologies relevant to DC power capacity additions
RELEVANT_TECHNOLOGIES = {
    # Natural Gas — indicates grid power expansion
    'Natural Gas Fired Combined Cycle',
    'Natural Gas Fired Combustion Turbine',
    'Natural Gas Steam Turbine',
    'Natural Gas Internal Combustion Engine',
    'Natural Gas with Compressed Air Storage',
    'Other Natural Gas',
    # Battery Storage — increasingly co-located with DCs
    'Batteries',
    'Battery',
    'Battery Storage',
    # Solar — indicates renewable capacity additions
    'Solar Photovoltaic',
    'Solar Thermal with Energy Storage',
    'Solar Thermal without Energy Storage',
}

# Normalized technology categories
TECH_CATEGORY_MAP = {
    'Natural Gas Fired Combined Cycle': 'natural_gas',
    'Natural Gas Fired Combustion Turbine': 'natural_gas',
    'Natural Gas Steam Turbine': 'natural_gas',
    'Natural Gas Internal Combustion Engine': 'natural_gas',
    'Natural Gas with Compressed Air Storage': 'natural_gas',
    'Other Natural Gas': 'natural_gas',
    'Batteries': 'battery_storage',
    'Battery': 'battery_storage',
    'Battery Storage': 'battery_storage',
    'Solar Photovoltaic': 'solar',
    'Solar Thermal with Energy Storage': 'solar',
    'Solar Thermal without Energy Storage': 'solar',
}

# US state abbreviations for validation
US_STATES = {
    'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'DC', 'FL',
    'GA', 'HI', 'ID', 'IL', 'IN', 'IA', 'KS', 'KY', 'LA', 'ME',
    'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH',
    'NJ', 'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI',
    'SC', 'SD', 'TN', 'TX', 'UT', 'VT', 'VA', 'WA', 'WV', 'WI',
    'WY',
}

# State FIPS lookup
STATE_FIPS = {
    'AL': '01', 'AK': '02', 'AZ': '04', 'AR': '05', 'CA': '06', 'CO': '08',
    'CT': '09', 'DE': '10', 'DC': '11', 'FL': '12', 'GA': '13', 'HI': '15',
    'ID': '16', 'IL': '17', 'IN': '18', 'IA': '19', 'KS': '20', 'KY': '21',
    'LA': '22', 'ME': '23', 'MD': '24', 'MA': '25', 'MI': '26', 'MN': '27',
    'MS': '28', 'MO': '29', 'MT': '30', 'NE': '31', 'NV': '32', 'NH': '33',
    'NJ': '34', 'NM': '35', 'NY': '36', 'NC': '37', 'ND': '38', 'OH': '39',
    'OK': '40', 'OR': '41', 'PA': '42', 'RI': '44', 'SC': '45', 'SD': '46',
    'TN': '47', 'TX': '48', 'UT': '49', 'VT': '50', 'VA': '51', 'WA': '53',
    'WV': '54', 'WI': '55', 'WY': '56',
}


def supabase_request(method, path, data=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
    }
    if headers_extra:
        headers.update(headers_extra)
    body = json.dumps(data, allow_nan=False).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                text = resp.read().decode()
                return json.loads(text) if text else None
        except urllib.error.HTTPError as e:
            error_body = e.read().decode() if e.fp else ''
            if e.code in (500, 502, 503) and attempt < 2:
                time.sleep(2 ** attempt)
                continue
            print(f"  HTTP {e.code}: {error_body[:500]}")
            raise
        except Exception:
            if attempt < 2:
                time.sleep(2 ** attempt)
                continue
            raise


def safe_str(val, max_len=500):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('none', 'null', 'n/a', 'unknown', '', 'nan'):
        return None
    return s[:max_len] if len(s) > max_len else s


def safe_float(val):
    if val is None:
        return None
    try:
        f = float(val)
        if f != f:  # NaN check
            return None
        return f
    except (ValueError, TypeError):
        return None


def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def download_eia860m(skip_download=False, custom_file=None):
    """Download EIA-860M Excel file."""
    os.makedirs(DATA_DIR, exist_ok=True)

    if custom_file:
        if os.path.exists(custom_file):
            print(f"  Using custom file: {custom_file}")
            return custom_file
        else:
            print(f"  ERROR: File not found: {custom_file}")
            return None

    # Check for existing cached file
    for fname in os.listdir(DATA_DIR):
        if fname.endswith('.xlsx') and 'generator' in fname.lower():
            fpath = os.path.join(DATA_DIR, fname)
            age_days = (time.time() - os.path.getmtime(fpath)) / 86400
            if age_days < 30 or skip_download:
                print(f"  Using cached file: {fname} ({age_days:.1f} days old)")
                return fpath

    if skip_download:
        print("  ERROR: No cached file found and --skip-download specified")
        return None

    # Try each URL
    for url in EIA_860M_URLS:
        fname = url.split('/')[-1]
        fpath = os.path.join(DATA_DIR, fname)
        print(f"  Downloading {fname}...")

        req = urllib.request.Request(url)
        req.add_header('User-Agent', 'GridScout/1.0 (energy-research)')
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                with open(fpath, 'wb') as f:
                    f.write(resp.read())
                size_mb = os.path.getsize(fpath) / (1024 * 1024)
                print(f"  Downloaded {size_mb:.1f} MB")
                return fpath
        except urllib.error.HTTPError as e:
            print(f"  HTTP {e.code} for {url}")
            continue
        except Exception as e:
            print(f"  Error downloading: {e}")
            continue

    print("  ERROR: Could not download EIA-860M from any URL")
    return None


def find_col(headers, candidates):
    """Find column index by trying multiple candidate names (case-insensitive, exact)."""
    norm_headers = {h.strip().lower(): i for i, h in enumerate(headers) if h}
    for c in candidates:
        if c.lower() in norm_headers:
            return norm_headers[c.lower()]
    return None


def parse_proposed_generators(filepath):
    """Parse the Proposed sheet from EIA-860M Excel file."""
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl required. Install: pip3 install openpyxl")
        sys.exit(1)

    print(f"  Opening {os.path.basename(filepath)}...")
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    # Find the Proposed sheet
    proposed_sheet = None
    for sheet_name in wb.sheetnames:
        if 'proposed' in sheet_name.lower():
            proposed_sheet = sheet_name
            break

    if not proposed_sheet:
        print(f"  Available sheets: {wb.sheetnames}")
        print("  ERROR: No 'Proposed' sheet found")
        return []

    print(f"  Reading sheet: {proposed_sheet}")
    ws = wb[proposed_sheet]

    # Find header row (look for row containing "Plant Name" or "Entity Name")
    header_row = None
    headers = []
    for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), 1):
        row_strs = [str(c).strip() if c else '' for c in row]
        joined = ' '.join(row_strs).lower()
        if 'plant name' in joined or 'entity name' in joined:
            header_row = row_idx
            headers = row_strs
            break

    if not header_row:
        print("  ERROR: Could not find header row")
        return []

    print(f"  Header at row {header_row}: {len(headers)} columns")

    # Map columns
    col_plant_id = find_col(headers, ['Plant ID', 'EIA Plant Code', 'Plant Code'])
    col_plant_name = find_col(headers, ['Plant Name'])
    col_state = find_col(headers, ['Plant State', 'State'])
    col_county = find_col(headers, ['County', 'Plant County'])
    col_capacity = find_col(headers, ['Nameplate Capacity (MW)', 'Capacity (MW)', 'Nameplate Capacity(MW)'])
    col_technology = find_col(headers, ['Technology', 'Prime Mover', 'Technology Type'])
    col_status = find_col(headers, ['Status', 'Generator Status'])
    col_planned_date = find_col(headers, [
        'Planned Operation Month', 'Planned Operation Date',
        'Current Month of Scheduled Delivery', 'Planned Online Month'
    ])
    col_planned_year = find_col(headers, [
        'Planned Operation Year', 'Current Year of Scheduled Delivery',
        'Planned Online Year'
    ])
    col_entity = find_col(headers, ['Entity Name', 'Utility Name', 'Developer'])
    col_entity_id = find_col(headers, ['Entity ID', 'Utility ID'])
    col_gen_id = find_col(headers, ['Generator ID'])
    col_lat = find_col(headers, ['Latitude'])
    col_lon = find_col(headers, ['Longitude'])
    col_energy_source = find_col(headers, ['Energy Source Code', 'Energy Source 1', 'Energy Source'])

    # Debug column mapping
    print(f"  Column mapping:")
    for name, idx in [
        ('Plant ID', col_plant_id), ('Plant Name', col_plant_name),
        ('State', col_state), ('County', col_county),
        ('Capacity', col_capacity), ('Technology', col_technology),
        ('Status', col_status), ('Planned Date', col_planned_date),
        ('Planned Year', col_planned_year), ('Entity', col_entity),
        ('Generator ID', col_gen_id), ('Lat', col_lat), ('Lon', col_lon),
        ('Energy Source', col_energy_source),
    ]:
        print(f"    {name}: col {idx}")

    if col_plant_name is None and col_entity is None:
        print("  ERROR: Could not find plant name or entity columns")
        return []

    # Parse rows
    records = []
    skipped_tech = {}
    total_rows = 0

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        total_rows += 1
        cells = list(row)

        def get(idx):
            if idx is None or idx >= len(cells):
                return None
            return cells[idx]

        plant_name = safe_str(get(col_plant_name))
        state = safe_str(get(col_state))
        technology = safe_str(get(col_technology))
        energy_source = safe_str(get(col_energy_source))

        if not state or state not in US_STATES:
            continue

        # Determine technology category
        tech_category = None
        if technology:
            tech_category = TECH_CATEGORY_MAP.get(technology)
        if not tech_category and energy_source:
            es = energy_source.upper()
            if es in ('NG', 'LFG', 'OG', 'BFG'):
                tech_category = 'natural_gas'
            elif es in ('SUN',):
                tech_category = 'solar'
            elif es in ('MWH', 'WAT'):  # battery codes
                tech_category = 'battery_storage'

        if not tech_category:
            t = technology or energy_source or 'Unknown'
            skipped_tech[t] = skipped_tech.get(t, 0) + 1
            continue

        capacity_mw = safe_float(get(col_capacity))
        if not capacity_mw or capacity_mw <= 0:
            continue

        plant_id = safe_int(get(col_plant_id))
        gen_id = safe_str(get(col_gen_id))
        county = safe_str(get(col_county))
        entity = safe_str(get(col_entity))
        status = safe_str(get(col_status))
        lat = safe_float(get(col_lat))
        lon = safe_float(get(col_lon))

        # Build planned operation date
        planned_month = safe_int(get(col_planned_date))
        planned_year = safe_int(get(col_planned_year))
        planned_operation_date = None
        if planned_year:
            if planned_month and 1 <= planned_month <= 12:
                planned_operation_date = f"{planned_year}-{planned_month:02d}-01"
            else:
                planned_operation_date = f"{planned_year}-01-01"

        # Source record ID
        sid = f"eia860m_prop_{plant_id or 0}_{gen_id or total_rows}"

        # Build FIPS code from state + county
        fips_code = None
        if county and state in STATE_FIPS:
            # We'll match FIPS codes later via county name lookup
            pass

        records.append({
            'source_record_id': sid,
            'plant_name': plant_name,
            'plant_id': plant_id,
            'generator_id': gen_id,
            'state': state,
            'county': county,
            'latitude': lat,
            'longitude': lon,
            'capacity_mw': round(capacity_mw, 2),
            'technology': technology,
            'tech_category': tech_category,
            'energy_source': energy_source,
            'developer': entity,
            'entity_id': safe_int(get(col_entity_id)),
            'status': status,
            'planned_operation_date': planned_operation_date,
        })

    wb.close()

    print(f"  {total_rows} total rows, {len(records)} relevant proposed generators")
    if skipped_tech:
        print(f"  Skipped technologies:")
        for t, c in sorted(skipped_tech.items(), key=lambda x: -x[1])[:10]:
            print(f"    {t}: {c}")

    return records


def ensure_table_exists():
    """Create grid_proposed_generators table if it doesn't exist via a test query."""
    try:
        supabase_request('GET', 'grid_proposed_generators?select=id&limit=1')
        return True
    except urllib.error.HTTPError as e:
        if e.code == 404 or '42P01' in str(e.read().decode() if e.fp else ''):
            print("  WARNING: grid_proposed_generators table does not exist.")
            print("  Create it with the following SQL:")
            print("""
CREATE TABLE IF NOT EXISTS grid_proposed_generators (
  id UUID DEFAULT gen_random_uuid() PRIMARY KEY,
  source_record_id TEXT NOT NULL UNIQUE,
  plant_name TEXT,
  plant_id INTEGER,
  generator_id TEXT,
  state TEXT NOT NULL,
  county TEXT,
  latitude NUMERIC(10,7),
  longitude NUMERIC(11,7),
  capacity_mw NUMERIC(10,2),
  technology TEXT,
  tech_category TEXT,               -- 'natural_gas', 'battery_storage', 'solar'
  energy_source TEXT,
  developer TEXT,
  entity_id INTEGER,
  status TEXT,
  planned_operation_date DATE,
  fips_code TEXT,
  data_source_id UUID REFERENCES grid_data_sources(id),
  created_at TIMESTAMPTZ DEFAULT NOW()
);

CREATE INDEX IF NOT EXISTS idx_grid_proposed_state ON grid_proposed_generators(state);
CREATE INDEX IF NOT EXISTS idx_grid_proposed_tech ON grid_proposed_generators(tech_category);
CREATE INDEX IF NOT EXISTS idx_grid_proposed_capacity ON grid_proposed_generators(capacity_mw DESC);
CREATE INDEX IF NOT EXISTS idx_grid_proposed_fips ON grid_proposed_generators(fips_code);

-- Also add proposed capacity columns to grid_county_data if not present:
ALTER TABLE grid_county_data ADD COLUMN IF NOT EXISTS proposed_gas_mw NUMERIC(10,2) DEFAULT 0;
ALTER TABLE grid_county_data ADD COLUMN IF NOT EXISTS proposed_solar_mw NUMERIC(10,2) DEFAULT 0;
ALTER TABLE grid_county_data ADD COLUMN IF NOT EXISTS proposed_storage_mw NUMERIC(10,2) DEFAULT 0;
ALTER TABLE grid_county_data ADD COLUMN IF NOT EXISTS proposed_total_mw NUMERIC(10,2) DEFAULT 0;
            """)
            return False
        raise


def get_existing_ids():
    """Load existing source_record_ids from grid_proposed_generators."""
    existing = set()
    offset = 0
    while True:
        result = supabase_request(
            'GET',
            f'grid_proposed_generators?select=source_record_id&limit=1000&offset={offset}'
        )
        if not result:
            break
        for r in result:
            if r.get('source_record_id'):
                existing.add(r['source_record_id'])
        if len(result) < 1000:
            break
        offset += 1000
    return existing


def load_county_fips():
    """Load FIPS code lookup from grid_county_data."""
    fips_map = {}  # (state, county_name_lower) -> fips_code
    offset = 0
    while True:
        result = supabase_request(
            'GET',
            f'grid_county_data?select=fips_code,state,county_name&limit=1000&offset={offset}'
        )
        if not result:
            break
        for r in result:
            key = (r.get('state', ''), r.get('county_name', '').lower())
            fips_map[key] = r.get('fips_code')
        if len(result) < 1000:
            break
        offset += 1000
    return fips_map


def enrich_county_data(records, dry_run=False):
    """Aggregate proposed capacity by county FIPS and update grid_county_data."""
    # Aggregate by FIPS code
    county_agg = {}  # fips -> {gas, solar, storage}
    for r in records:
        fips = r.get('fips_code')
        if not fips:
            continue
        if fips not in county_agg:
            county_agg[fips] = {'gas': 0, 'solar': 0, 'storage': 0}
        cap = r.get('capacity_mw') or 0
        cat = r.get('tech_category', '')
        if cat == 'natural_gas':
            county_agg[fips]['gas'] += cap
        elif cat == 'solar':
            county_agg[fips]['solar'] += cap
        elif cat == 'battery_storage':
            county_agg[fips]['storage'] += cap

    print(f"\n  Counties with proposed capacity: {len(county_agg)}")
    if not county_agg:
        return

    if dry_run:
        print(f"  [DRY RUN] Would update {len(county_agg)} county records")
        top_counties = sorted(county_agg.items(), key=lambda x: sum(x[1].values()), reverse=True)[:10]
        for fips, vals in top_counties:
            total = sum(vals.values())
            print(f"    FIPS {fips}: {total:.0f} MW (gas={vals['gas']:.0f}, solar={vals['solar']:.0f}, storage={vals['storage']:.0f})")
        return

    # Update grid_county_data
    patched = 0
    errors = 0
    for fips, vals in county_agg.items():
        total = vals['gas'] + vals['solar'] + vals['storage']
        try:
            encoded_fips = urllib.parse.quote(fips, safe='')
            supabase_request('PATCH', f'grid_county_data?fips_code=eq.{encoded_fips}', {
                'proposed_gas_mw': round(vals['gas'], 2),
                'proposed_solar_mw': round(vals['solar'], 2),
                'proposed_storage_mw': round(vals['storage'], 2),
                'proposed_total_mw': round(total, 2),
            })
            patched += 1
        except Exception as e:
            errors += 1
            if errors <= 5:
                print(f"    Error patching FIPS {fips}: {e}")

    print(f"  County enrichment: {patched} patched, {errors} errors")


def main():
    print("=" * 60)
    print("GridScout EIA-860M Proposed Generator Ingestion")
    print("=" * 60)

    dry_run = '--dry-run' in sys.argv
    skip_download = '--skip-download' in sys.argv
    custom_file = None
    if '--file' in sys.argv:
        idx = sys.argv.index('--file')
        if idx + 1 < len(sys.argv):
            custom_file = sys.argv[idx + 1]

    if dry_run:
        print("[DRY RUN MODE]")

    # Download/load EIA-860M
    print("\nLoading EIA-860M data...")
    filepath = download_eia860m(skip_download=skip_download, custom_file=custom_file)
    if not filepath:
        sys.exit(1)

    # Parse proposed generators
    print("\nParsing proposed generators...")
    records = parse_proposed_generators(filepath)
    if not records:
        print("No relevant proposed generators found.")
        return

    # Summary
    print(f"\n{'=' * 60}")
    print(f"Proposed Generator Summary")
    print(f"{'=' * 60}")

    # By technology category
    tech_counts = {}
    tech_capacity = {}
    for r in records:
        cat = r.get('tech_category', 'unknown')
        tech_counts[cat] = tech_counts.get(cat, 0) + 1
        tech_capacity[cat] = tech_capacity.get(cat, 0) + (r.get('capacity_mw') or 0)

    print(f"\nBy technology:")
    for cat in sorted(tech_counts.keys()):
        print(f"  {cat}: {tech_counts[cat]} generators, {tech_capacity[cat]:,.0f} MW")

    # By state (top 15)
    state_counts = {}
    state_capacity = {}
    for r in records:
        s = r.get('state', '??')
        state_counts[s] = state_counts.get(s, 0) + 1
        state_capacity[s] = state_capacity.get(s, 0) + (r.get('capacity_mw') or 0)

    print(f"\nTop states by proposed capacity:")
    for s, cap in sorted(state_capacity.items(), key=lambda x: -x[1])[:15]:
        print(f"  {s}: {cap:,.0f} MW ({state_counts[s]} generators)")

    # By developer (top 15)
    dev_counts = {}
    for r in records:
        d = r.get('developer') or 'Unknown'
        dev_counts[d] = dev_counts.get(d, 0) + 1

    print(f"\nTop developers:")
    for d, c in sorted(dev_counts.items(), key=lambda x: -x[1])[:15]:
        print(f"  {d}: {c}")

    # Total capacity
    total_cap = sum(r.get('capacity_mw', 0) for r in records)
    print(f"\nTotal proposed capacity: {total_cap:,.0f} MW across {len(records)} generators")

    if dry_run:
        print(f"\n[DRY RUN] Would insert up to {len(records)} proposed generator records.")

    # Check table exists
    if not dry_run:
        if not ensure_table_exists():
            print("\n  Run the SQL above in Supabase SQL Editor, then re-run this script.")
            sys.exit(1)

    # FIPS code enrichment via county name lookup
    print("\nLooking up FIPS codes...")
    fips_map = load_county_fips()
    print(f"  Loaded {len(fips_map)} county FIPS codes")

    matched_fips = 0
    for r in records:
        county = r.get('county')
        state = r.get('state')
        if county and state:
            # Try exact match
            key = (state, county.lower())
            fips = fips_map.get(key)
            if not fips:
                # Try with "County" suffix/prefix removed
                county_clean = county.lower().replace(' county', '').replace('county ', '').strip()
                for (s, cn), f in fips_map.items():
                    if s == state and county_clean in cn:
                        fips = f
                        break
            if fips:
                r['fips_code'] = fips
                matched_fips += 1

    print(f"  FIPS matched: {matched_fips}/{len(records)} ({100 * matched_fips / len(records):.1f}%)")

    # Enrich county data
    print("\nEnriching county data with proposed capacity...")
    enrich_county_data(records, dry_run=dry_run)

    if dry_run:
        return

    # Get or create data source
    ds = supabase_request('GET', 'grid_data_sources?name=eq.eia860m_proposed&select=id')
    data_source_id = ds[0]['id'] if ds else None
    if not data_source_id:
        print("\nCreating eia860m_proposed data source...")
        result = supabase_request('POST', 'grid_data_sources', [{
            'name': 'eia860m_proposed',
            'url': 'https://www.eia.gov/electricity/data/eia860m/',
            'description': 'EIA-860M Monthly proposed generators (natural gas, battery storage, solar)',
        }], {'Prefer': 'return=representation'})
        if result:
            data_source_id = result[0]['id']

    # Load existing records
    print("\nLoading existing records...")
    existing_ids = get_existing_ids()
    print(f"  {len(existing_ids)} existing records in DB")

    # Filter new records
    new_records = []
    for r in records:
        if r['source_record_id'] in existing_ids:
            continue
        # Build insertable record (only table columns)
        rec = {
            'source_record_id': r['source_record_id'],
            'plant_name': r.get('plant_name'),
            'plant_id': r.get('plant_id'),
            'generator_id': r.get('generator_id'),
            'state': r['state'],
            'county': r.get('county'),
            'latitude': r.get('latitude'),
            'longitude': r.get('longitude'),
            'capacity_mw': r.get('capacity_mw'),
            'technology': r.get('technology'),
            'tech_category': r.get('tech_category'),
            'energy_source': r.get('energy_source'),
            'developer': r.get('developer'),
            'entity_id': r.get('entity_id'),
            'status': r.get('status'),
            'planned_operation_date': r.get('planned_operation_date'),
            'fips_code': r.get('fips_code'),
            'created_at': datetime.now(timezone.utc).isoformat(),
        }
        if data_source_id:
            rec['data_source_id'] = data_source_id
        new_records.append(rec)

    print(f"  {len(new_records)} new records to insert ({len(records) - len(new_records)} already exist)")

    if not new_records:
        print("\nNo new records to insert.")
        return

    # Insert in batches
    created = 0
    errors = 0
    for i in range(0, len(new_records), BATCH_SIZE):
        batch = new_records[i:i + BATCH_SIZE]
        try:
            supabase_request('POST', 'grid_proposed_generators', batch, {'Prefer': 'return=minimal'})
            created += len(batch)
            if created % 200 == 0 or i + BATCH_SIZE >= len(new_records):
                print(f"  Inserted {created}/{len(new_records)}...")
        except Exception as e:
            print(f"  Batch error at {i}: {e}")
            # Try one-by-one
            for rec in batch:
                try:
                    supabase_request('POST', 'grid_proposed_generators', [rec], {'Prefer': 'return=minimal'})
                    created += 1
                except Exception as e2:
                    errors += 1
                    if errors <= 10:
                        print(f"  Record error ({rec['source_record_id']}): {e2}")

    # Update data source count
    if data_source_id:
        supabase_request('PATCH', f'grid_data_sources?id=eq.{data_source_id}', {
            'record_count': len(existing_ids) + created,
            'last_import': datetime.now(timezone.utc).isoformat()
        })

    print(f"\n{'=' * 60}")
    print(f"EIA-860M Proposed Generator Ingestion Complete")
    print(f"  Created: {created}")
    print(f"  Skipped (existing): {len(records) - len(new_records)}")
    print(f"  Errors: {errors}")
    print(f"  Total in DB: {len(existing_ids) + created}")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
