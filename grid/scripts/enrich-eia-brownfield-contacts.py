#!/usr/bin/env python3
"""
Enrich grid_brownfield_sites with operator contact info from EIA-860 Plant file.
Matches brownfield sites that have eia_plant_id to EIA Plant records for operator name+address.
Also enriches from EIA Utility file for utility company details.
"""

import os
import sys
import json
import time
import urllib.request
import urllib.error
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env.local'))

SUPABASE_URL = os.environ.get('SUPABASE_URL') or os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

# Data files
PLANT_FILE = '/tmp/2___Plant_Y2024.xlsx'
UTILITY_FILE = os.path.join(os.path.dirname(__file__), '..', '..', 'solar', 'data', 'eia860_2024', '1___Utility_Y2024.xlsx')


def supabase_request(method, path, data=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
    }
    if headers_extra:
        headers.update(headers_extra)
    body = json.dumps(data, allow_nan=False).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                text = resp.read().decode()
                return json.loads(text) if text else None
        except urllib.error.HTTPError as e:
            error_body = e.read().decode() if e.fp else ''
            if e.code in (500, 502, 503) and attempt < 2:
                time.sleep(2 ** attempt)
                continue
            print(f"  HTTP {e.code}: {error_body[:500]}")
            raise
        except Exception:
            if attempt < 2:
                time.sleep(2 ** attempt)
                continue
            raise


def load_eia_plants():
    """Load EIA-860 Plant data: plant_code -> {utility_name, address, city, state, zip}"""
    import openpyxl
    print(f"  Loading EIA Plant file: {PLANT_FILE}")
    wb = openpyxl.load_workbook(PLANT_FILE, read_only=True)
    ws = wb.active

    plants = {}
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        # Row 2 is header, data starts row 3
        if not row or not row[2]:  # Plant Code at col 2
            continue
        try:
            plant_code = int(row[2])
        except (ValueError, TypeError):
            continue

        plants[plant_code] = {
            'utility_name': str(row[1]).strip() if row[1] else None,
            'address': str(row[4]).strip() if row[4] else None,
            'city': str(row[5]).strip() if row[5] else None,
            'state': str(row[6]).strip() if row[6] else None,
            'zip': str(row[7]).strip() if row[7] else None,
        }

    wb.close()
    print(f"  {len(plants)} plants loaded")
    return plants


def load_eia_utilities():
    """Load EIA-860 Utility data: utility_id -> {name, address, city, state, zip}"""
    import openpyxl
    if not os.path.exists(UTILITY_FILE):
        print(f"  Utility file not found: {UTILITY_FILE}")
        return {}

    print(f"  Loading EIA Utility file: {UTILITY_FILE}")
    wb = openpyxl.load_workbook(UTILITY_FILE, read_only=True)
    ws = wb.active

    utilities = {}
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        if not row or not row[0]:
            continue
        try:
            util_id = int(row[0])
        except (ValueError, TypeError):
            continue

        utilities[util_id] = {
            'name': str(row[1]).strip() if row[1] else None,
            'address': str(row[2]).strip() if row[2] else None,
            'city': str(row[3]).strip() if row[3] else None,
            'state': str(row[4]).strip() if row[4] else None,
            'zip': str(row[5]).strip() if row[5] else None,
        }

    wb.close()
    print(f"  {len(utilities)} utilities loaded")
    return utilities


def main():
    print("=" * 60)
    print("GridScout: EIA-860 Brownfield Contact Enrichment")
    print("=" * 60)

    dry_run = '--dry-run' in sys.argv

    # Check if Plant file exists, download if needed
    if not os.path.exists(PLANT_FILE):
        print(f"\nPlant file not found at {PLANT_FILE}")
        print("Download it first: cd /tmp && curl -sL 'https://www.eia.gov/electricity/data/eia860/xls/eia8602024.zip' -o eia8602024.zip && unzip -o eia8602024.zip 2___Plant_Y2024.xlsx")
        return

    plants = load_eia_plants()
    utilities = load_eia_utilities()

    # Load brownfield sites with eia_plant_id
    print("\nLoading brownfield sites...")
    brownfields = []
    offset = 0
    while True:
        batch = supabase_request('GET',
            f'grid_brownfield_sites?select=id,eia_plant_id,name,operator_name&limit=1000&offset={offset}')
        if not batch:
            break
        brownfields.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    print(f"  {len(brownfields)} brownfield sites in DB")

    with_eia = [b for b in brownfields if b.get('eia_plant_id')]
    print(f"  {len(with_eia)} have eia_plant_id")

    # Enrich brownfield sites
    patched = 0
    errors = 0
    for bf in with_eia:
        plant_id = bf['eia_plant_id']
        plant = plants.get(plant_id)
        if not plant:
            continue

        patch = {}
        if plant.get('utility_name') and not bf.get('operator_name'):
            patch['operator_name'] = plant['utility_name']

        addr_parts = []
        if plant.get('address'):
            addr_parts.append(plant['address'])
        if plant.get('city'):
            addr_parts.append(plant['city'])
        if plant.get('state'):
            addr_parts.append(plant['state'])
        if plant.get('zip'):
            addr_parts.append(plant['zip'])
        if addr_parts:
            patch['operator_address'] = ', '.join(addr_parts)

        if not patch:
            continue

        if not dry_run:
            try:
                supabase_request('PATCH',
                    f'grid_brownfield_sites?id=eq.{bf["id"]}', patch)
                patched += 1
            except Exception as e:
                errors += 1
                if errors <= 5:
                    print(f"  Error patching {bf['id']}: {e}")
        else:
            patched += 1

    print(f"\n{'=' * 60}")
    print(f"EIA Brownfield Enrichment Complete")
    print(f"  Brownfields with EIA ID: {len(with_eia)}")
    print(f"  Matched to EIA Plant:    {patched}")
    print(f"  Errors:                  {errors}")
    if dry_run:
        print("  [DRY RUN — no changes made]")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
