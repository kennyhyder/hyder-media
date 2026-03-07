#!/usr/bin/env python3
"""
Enrich grid_brownfield_sites and grid_dc_sites with utility contact info from EIA-860 Utility file.
Maps utility_id (from EIA Plant file) to utility company address/contact details.
Also enriches IXP facilities and datacenters that have operator_name matching a utility.

Extends enrich-eia-brownfield-contacts.py by doing utility-level lookups.
"""

import os
import sys
import json
import time
import urllib.request
import urllib.error
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env.local'))

SUPABASE_URL = os.environ.get('SUPABASE_URL') or os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')

PLANT_FILE = '/tmp/2___Plant_Y2024.xlsx'
UTILITY_FILE = os.path.join(os.path.dirname(__file__), '..', '..', 'solar', 'data', 'eia860_2024', '1___Utility_Y2024.xlsx')


def supabase_request(method, path, data=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
    }
    if headers_extra:
        headers.update(headers_extra)
    body = json.dumps(data, allow_nan=False).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                text = resp.read().decode()
                return json.loads(text) if text else None
        except urllib.error.HTTPError as e:
            error_body = e.read().decode() if e.fp else ''
            if e.code in (500, 502, 503) and attempt < 2:
                time.sleep(2 ** attempt)
                continue
            print(f"  HTTP {e.code}: {error_body[:500]}")
            raise
        except Exception:
            if attempt < 2:
                time.sleep(2 ** attempt)
                continue
            raise


def load_eia_utilities():
    """Load EIA-860 Utility data: utility_id -> full contact details."""
    import openpyxl
    if not os.path.exists(UTILITY_FILE):
        print(f"  Utility file not found: {UTILITY_FILE}")
        return {}, {}

    print(f"  Loading EIA Utility file: {UTILITY_FILE}")
    wb = openpyxl.load_workbook(UTILITY_FILE, read_only=True)
    ws = wb.active

    # Read header row to find columns
    header = None
    for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
        header = [str(c).strip().lower() if c else '' for c in row]
        break

    if not header:
        wb.close()
        return {}, {}

    # Find column indices
    col_map = {}
    for i, h in enumerate(header):
        if 'utility_id' in h or h == 'utility id':
            col_map['id'] = i
        elif 'utility_name' in h or h == 'utility name':
            col_map['name'] = i
        elif h in ('street_address', 'street address'):
            col_map['address'] = i
        elif h == 'city':
            col_map['city'] = i
        elif h == 'state':
            col_map['state'] = i
        elif h == 'zip':
            col_map['zip'] = i
        elif h == 'zip5':
            col_map['zip'] = i
        elif 'phone' in h and 'phone' not in col_map:
            col_map['phone'] = i
        elif 'entity_type' in h or h == 'entity type':
            col_map['entity_type'] = i

    utilities_by_id = {}
    utilities_by_name = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[col_map.get('id', 0)]:
            continue
        try:
            util_id = int(row[col_map['id']])
        except (ValueError, TypeError):
            continue

        name = str(row[col_map.get('name', 1)]).strip() if row[col_map.get('name', 1)] else None
        address = str(row[col_map.get('address', 2)]).strip() if col_map.get('address') and row[col_map['address']] else None
        city = str(row[col_map.get('city', 3)]).strip() if col_map.get('city') and row[col_map['city']] else None
        state = str(row[col_map.get('state', 4)]).strip() if col_map.get('state') and row[col_map['state']] else None
        zipcode = str(row[col_map.get('zip', 5)]).strip() if col_map.get('zip') and row[col_map['zip']] else None
        phone = str(row[col_map.get('phone', -1)]).strip() if col_map.get('phone') and row[col_map['phone']] else None

        rec = {
            'name': name,
            'address': address,
            'city': city,
            'state': state,
            'zip': zipcode,
            'phone': phone,
        }

        utilities_by_id[util_id] = rec
        if name:
            norm = name.lower().strip()
            utilities_by_name[norm] = rec

    wb.close()
    print(f"  {len(utilities_by_id)} utilities loaded ({len(utilities_by_name)} unique names)")
    return utilities_by_id, utilities_by_name


def load_eia_plants():
    """Load EIA Plant file to get utility_id for each plant_code."""
    import openpyxl
    if not os.path.exists(PLANT_FILE):
        return {}

    print(f"  Loading EIA Plant file: {PLANT_FILE}")
    wb = openpyxl.load_workbook(PLANT_FILE, read_only=True)
    ws = wb.active

    plant_to_utility = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[2]:
            continue
        try:
            plant_code = int(row[2])
            utility_id = int(row[0])
            plant_to_utility[plant_code] = utility_id
        except (ValueError, TypeError):
            continue

    wb.close()
    print(f"  {len(plant_to_utility)} plant→utility mappings loaded")
    return plant_to_utility


def format_address(rec):
    """Build formatted address string from utility record."""
    parts = []
    if rec.get('address'):
        parts.append(rec['address'])
    if rec.get('city'):
        parts.append(rec['city'])
    if rec.get('state'):
        parts.append(rec['state'])
    if rec.get('zip'):
        parts.append(rec['zip'])
    return ', '.join(parts) if parts else None


def main():
    print("=" * 60)
    print("GridScout: EIA Utility Contact Enrichment")
    print("=" * 60)

    dry_run = '--dry-run' in sys.argv

    utilities_by_id, utilities_by_name = load_eia_utilities()
    plant_to_utility = load_eia_plants()

    if not utilities_by_id:
        print("No utility data loaded. Exiting.")
        return

    # Phase 1: Brownfield sites — use eia_plant_id → utility_id → utility contact
    print("\n--- Phase 1: Brownfield Sites ---")
    brownfields = []
    offset = 0
    while True:
        batch = supabase_request('GET',
            f'grid_brownfield_sites?select=id,eia_plant_id,name,operator_name,operator_address,operator_phone'
            f'&limit=1000&offset={offset}')
        if not batch:
            break
        brownfields.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    print(f"  {len(brownfields)} brownfield sites loaded")

    bf_patched = 0
    bf_errors = 0
    for bf in brownfields:
        plant_id = bf.get('eia_plant_id')
        if not plant_id:
            continue

        utility_id = plant_to_utility.get(plant_id)
        if not utility_id:
            continue

        util = utilities_by_id.get(utility_id)
        if not util:
            continue

        patch = {}
        if util.get('name') and not bf.get('operator_name'):
            patch['operator_name'] = util['name']

        addr = format_address(util)
        if addr and not bf.get('operator_address'):
            patch['operator_address'] = addr

        # Always try to fill phone — it's the main gap
        if util.get('phone') and not bf.get('operator_phone'):
            patch['operator_phone'] = util['phone']

        if not patch:
            continue

        if not dry_run:
            try:
                supabase_request('PATCH', f'grid_brownfield_sites?id=eq.{bf["id"]}', patch)
                bf_patched += 1
            except Exception as e:
                bf_errors += 1
                if bf_errors <= 5:
                    print(f"  Error: {e}")
        else:
            bf_patched += 1
            if bf_patched <= 5:
                print(f"  [DRY] {bf.get('name', '?')}: {util['name']} — {addr}")

    print(f"  Brownfields patched: {bf_patched}, errors: {bf_errors}")

    # Phase 2: IXP facilities — match org_name to utility name
    print("\n--- Phase 2: IXP Facilities ---")
    ixps = []
    offset = 0
    while True:
        batch = supabase_request('GET',
            f'grid_ixp_facilities?select=id,name,org_name,sales_phone'
            f'&sales_phone=is.null&limit=1000&offset={offset}')
        if not batch:
            break
        ixps.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    print(f"  {len(ixps)} IXPs missing phone")

    ixp_patched = 0
    for ixp in ixps:
        op = ixp.get('org_name')
        if not op:
            continue
        norm = op.lower().strip()
        util = utilities_by_name.get(norm)
        if not util:
            continue

        patch = {}
        if util.get('phone') and not ixp.get('sales_phone'):
            patch['sales_phone'] = util['phone']

        if not patch:
            continue

        if not dry_run:
            try:
                supabase_request('PATCH', f'grid_ixp_facilities?id=eq.{ixp["id"]}', patch)
                ixp_patched += 1
            except Exception:
                pass
        else:
            ixp_patched += 1

    print(f"  IXPs patched: {ixp_patched}")

    # Phase 3: Datacenters — match name to utility name
    print("\n--- Phase 3: Datacenters ---")
    dcs = []
    offset = 0
    while True:
        batch = supabase_request('GET',
            f'grid_datacenters?select=id,name,sales_phone'
            f'&sales_phone=is.null&limit=1000&offset={offset}')
        if not batch:
            break
        dcs.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    print(f"  {len(dcs)} DCs missing phone")

    dc_patched = 0
    for dc in dcs:
        op = dc.get('name')
        if not op:
            continue
        norm = op.lower().strip()
        util = utilities_by_name.get(norm)
        if not util:
            continue

        patch = {}
        if util.get('phone') and not dc.get('sales_phone'):
            patch['sales_phone'] = util['phone']

        if not patch:
            continue

        if not dry_run:
            try:
                supabase_request('PATCH', f'grid_datacenters?id=eq.{dc["id"]}', patch)
                dc_patched += 1
            except Exception:
                pass
        else:
            dc_patched += 1

    print(f"  DCs patched: {dc_patched}")

    print(f"\n{'=' * 60}")
    print(f"EIA Utility Contact Enrichment Complete")
    print(f"  Brownfields: {bf_patched}")
    print(f"  IXPs:        {ixp_patched}")
    print(f"  Datacenters: {dc_patched}")
    if dry_run:
        print("  [DRY RUN — no changes made]")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
