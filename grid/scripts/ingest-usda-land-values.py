#!/usr/bin/env python3
"""
Fetch county-level agricultural land values from USDA NASS QuickStats.
Adds avg_land_value_per_acre_usd to grid_county_data table.

Uses the NASS QuickStats API (requires free API key from https://quickstats.nass.usda.gov/api/).
Set USDA_NASS_API_KEY in grid/.env.local.

Alternatively, downloads from the NASS bulk tab-delimited file if API key not available.
"""

import os
import sys
import json
import time
import csv
import io
import urllib.request
import urllib.error
import urllib.parse
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env.local'))

SUPABASE_URL = os.environ.get('SUPABASE_URL') or os.environ.get('NEXT_PUBLIC_SUPABASE_URL')
SUPABASE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')
NASS_API_KEY = os.environ.get('USDA_NASS_API_KEY', '')

NASS_BASE = 'https://quickstats.nass.usda.gov/api/api_GET/'
DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')


def supabase_request(method, path, data=None, headers_extra=None):
    url = f"{SUPABASE_URL}/rest/v1/{path}"
    headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f'Bearer {SUPABASE_KEY}',
        'Content-Type': 'application/json',
    }
    if headers_extra:
        headers.update(headers_extra)
    body = json.dumps(data, allow_nan=False).encode() if data else None
    req = urllib.request.Request(url, data=body, headers=headers, method=method)
    for attempt in range(3):
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                text = resp.read().decode()
                return json.loads(text) if text else None
        except urllib.error.HTTPError as e:
            error_body = e.read().decode() if e.fp else ''
            if e.code in (500, 502, 503) and attempt < 2:
                time.sleep(2 ** attempt)
                continue
            print(f"  HTTP {e.code}: {error_body[:500]}")
            raise
        except Exception:
            if attempt < 2:
                time.sleep(2 ** attempt)
                continue
            raise


def fetch_nass_via_api():
    """Fetch county-level ag land values from USDA NASS QuickStats API (state by state)."""
    if not NASS_API_KEY:
        return None

    states = [
        'AL','AK','AZ','AR','CA','CO','CT','DE','FL','GA','HI','ID','IL','IN',
        'IA','KS','KY','LA','ME','MD','MA','MI','MN','MS','MO','MT','NE','NV',
        'NH','NJ','NM','NY','NC','ND','OH','OK','OR','PA','RI','SC','SD','TN',
        'TX','UT','VT','VA','WA','WV','WI','WY'
    ]

    all_records = []
    for state in states:
        for year in ['2024', '2023', '2022']:
            params = {
                'key': NASS_API_KEY,
                'commodity_desc': 'AG LAND',
                'statisticcat_desc': 'ASSET VALUE',
                'unit_desc': '$ / ACRE',
                'agg_level_desc': 'COUNTY',
                'year': year,
                'state_alpha': state,
                'domain_desc': 'TOTAL',
                'format': 'JSON',
            }

            url = NASS_BASE + '?' + urllib.parse.urlencode(params)
            req = urllib.request.Request(url)
            req.add_header('User-Agent', 'GridScout/1.0')

            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    data = json.loads(resp.read().decode())
                    records = data.get('data', [])
                    if records:
                        all_records.extend(records)
                        break  # Got data for this state, skip older years
            except Exception:
                pass

            time.sleep(0.3)
        time.sleep(0.2)

    print(f"  {len(all_records)} records via API across {len(states)} states")
    return all_records


def fetch_nass_via_ers():
    """Download county-level farmland values from USDA ERS (free, no API key)."""
    # USDA ERS publishes county-level farmland value data as Excel
    ers_url = 'https://www.ers.usda.gov/webdocs/DataFiles/48457/FarmlandValues.xlsx'
    local_path = os.path.join(DATA_DIR, 'usda_farmland_values.xlsx')

    os.makedirs(DATA_DIR, exist_ok=True)

    if not os.path.exists(local_path) or '--redownload' in sys.argv:
        print(f"  Downloading USDA ERS farmland values...")
        req = urllib.request.Request(ers_url)
        req.add_header('User-Agent', 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)')
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                with open(local_path, 'wb') as f:
                    f.write(resp.read())
            print(f"  Saved to {local_path}")
        except Exception as e:
            print(f"  ERS download failed: {e}")
            return None
    else:
        print(f"  Using cached file: {local_path}")

    # Parse the Excel file
    import openpyxl
    wb = openpyxl.load_workbook(local_path, read_only=True)

    # Try sheets that might have county-level data
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if 'county' not in sheet_name.lower():
            continue

        print(f"  Reading sheet: {sheet_name}")
        headers = None
        for row in ws.iter_rows(values_only=True):
            if not headers:
                headers = [str(c).strip().lower() if c else '' for c in row]
                continue

            row_dict = dict(zip(headers, row))
            state_fips = str(row_dict.get('state fips', row_dict.get('statefips', ''))).strip()
            county_fips = str(row_dict.get('county fips', row_dict.get('countyfips', ''))).strip()
            value = row_dict.get('value', row_dict.get('$ per acre', row_dict.get('dollars per acre', '')))

            if state_fips and county_fips and value:
                records.append({
                    'state_fips_code': state_fips.zfill(2),
                    'county_code': county_fips.zfill(3),
                    'Value': str(value),
                })

        if records:
            break

    wb.close()

    if not records:
        print("  No county-level data found in ERS file. Using state-level estimates.")
        return fetch_state_level_estimates()

    print(f"  {len(records)} county records from ERS")
    return records


def fetch_state_level_estimates():
    """Use USDA published state-level average farmland values as fallback.
    2024 USDA Land Values Summary (published Aug 2024).
    Values are $/acre for farm real estate (land + buildings).
    """
    # Source: USDA NASS Land Values 2024 Summary (publicly available PDF/table)
    state_values = {
        'AL': 4400, 'AK': 1500, 'AZ': 2170, 'AR': 4200, 'CA': 12600,
        'CO': 2440, 'CT': 14900, 'DE': 9900, 'FL': 9400, 'GA': 5050,
        'HI': 11100, 'ID': 4300, 'IL': 9200, 'IN': 8300, 'IA': 9800,
        'KS': 2520, 'KY': 5200, 'LA': 4600, 'ME': 3100, 'MD': 10900,
        'MA': 15600, 'MI': 7200, 'MN': 6200, 'MS': 3500, 'MO': 4400,
        'MT': 1090, 'NE': 3500, 'NV': 1340, 'NH': 6400, 'NJ': 17200,
        'NM': 690, 'NY': 4800, 'NC': 5850, 'ND': 2520, 'OH': 8200,
        'OK': 2550, 'OR': 3100, 'PA': 8100, 'RI': 17900, 'SC': 4200,
        'SD': 2300, 'TN': 5400, 'TX': 2900, 'UT': 2800, 'VT': 4000,
        'VA': 5900, 'WA': 4300, 'WV': 2700, 'WI': 6400, 'WY': 780,
        'DC': 0,
    }
    print(f"  Using state-level land value estimates ({len(state_values)} states)")
    return state_values


def main():
    print("=" * 60)
    print("GridScout: USDA NASS Land Value Ingestion")
    print("=" * 60)

    dry_run = '--dry-run' in sys.argv

    # Step 1: Ensure column exists
    print("\nNote: Ensure grid_county_data has avg_land_value_per_acre_usd column.")
    print("  ALTER TABLE grid_county_data ADD COLUMN IF NOT EXISTS avg_land_value_per_acre_usd NUMERIC(10,2);")

    # Step 2: Fetch NASS data (try API first, fall back to ERS, then state estimates)
    print("\nFetching USDA land values...")
    nass_records = None
    state_fallback = None

    if NASS_API_KEY:
        print("  Using NASS API key...")
        nass_records = fetch_nass_via_api()
    if not nass_records:
        print("  Trying USDA ERS download...")
        nass_records = fetch_nass_via_ers()

    # Check if we got state-level dict instead of records list
    if isinstance(nass_records, dict):
        state_fallback = nass_records
        nass_records = None

    if not nass_records and not state_fallback:
        print("  Using state-level estimates as final fallback...")
        state_fallback = fetch_state_level_estimates()

    if not nass_records and not state_fallback:
        print("No USDA data retrieved. Exiting.")
        return

    # Step 3: Parse into FIPS → value mapping
    fips_values = {}
    if nass_records:
        for rec in nass_records:
            state_fips = rec.get('state_fips_code', '')
            county_fips = rec.get('county_code', '')
            value_str = str(rec.get('Value', '')).replace(',', '').strip()

            if not state_fips or not county_fips or county_fips == '':
                continue
            if value_str in ('', '(D)', '(Z)', '(NA)', '(S)'):
                continue

            fips = f"{state_fips.zfill(2)}{county_fips.zfill(3)}"
            try:
                value = float(value_str)
                if value > 0:
                    fips_values[fips] = value
            except ValueError:
                continue

        print(f"  {len(fips_values)} counties with valid land values")

    if not fips_values and not state_fallback:
        print("No valid land values parsed. Exiting.")
        return

    if fips_values:
        values = sorted(fips_values.values())
        print(f"  Min: ${values[0]:,.0f}/acre")
        print(f"  Median: ${values[len(values)//2]:,.0f}/acre")
        print(f"  Max: ${values[-1]:,.0f}/acre")
    elif state_fallback:
        sv = sorted(state_fallback.values())
        print(f"  State-level: ${sv[0]:,.0f} - ${sv[-1]:,.0f}/acre ({len(state_fallback)} states)")

    # Step 4: Load existing counties from DB
    print("\nLoading grid_county_data...")
    counties = []
    offset = 0
    while True:
        batch = supabase_request('GET',
            f'grid_county_data?select=id,fips_code,state&limit=1000&offset={offset}')
        if not batch:
            break
        counties.extend(batch)
        if len(batch) < 1000:
            break
        offset += 1000
    print(f"  {len(counties)} counties in DB")

    # Step 5: Patch land values
    patched = 0
    errors = 0
    skipped = 0
    for county in counties:
        fips = county.get('fips_code')
        state = county.get('state')
        value = None

        # Try FIPS-level first, then state fallback
        if fips and fips in fips_values:
            value = fips_values[fips]
        elif state_fallback and state and state in state_fallback:
            value = state_fallback[state]

        if not value or value <= 0:
            skipped += 1
            continue

        patch = {'avg_land_value_per_acre_usd': value}

        if not dry_run:
            try:
                supabase_request('PATCH', f'grid_county_data?id=eq.{county["id"]}', patch)
                patched += 1
            except Exception as e:
                errors += 1
                if errors <= 5:
                    print(f"  Error patching {fips}: {e}")
        else:
            patched += 1

    print(f"\n{'=' * 60}")
    print(f"USDA Land Value Ingestion Complete")
    print(f"  Counties patched: {patched}")
    print(f"  No data:          {skipped}")
    print(f"  Errors:           {errors}")
    if dry_run:
        print("  [DRY RUN — no changes made]")
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
